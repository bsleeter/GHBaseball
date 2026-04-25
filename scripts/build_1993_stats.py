#!/usr/bin/env python3
"""Build 1993 Gig Harbor Varsity season stats xlsx."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1993" / "1993_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1993, [
        ("Individual Records", "Typed source labeled 'Hall of Fame — 1993 Records'.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source labeled 'Highlights of 1993 Team'.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No schedule photo provided.", "NOT PROVIDED"),
        ("Team Batting", "Hand-written stats sheet transcribed. Some breakdown cells inconsistent (TB doesn't always sum cleanly to 1B/2B/3B/HR) — preserved as printed.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Hand-written stats sheet transcribed. Team: 12-6, 7 SV, 2.57 ERA.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Joel Miller",        10),
        (2,  "Andy Cherbas",       11),
        (3,  "Ryan Sawyers",       11),
        (4,  "Tom Friedman",       10),
        (5,  "Jason Lippert",      10),
        (6,  "Paul Baurichter",    10),
        (7,  "Derek Vitcovich",    10),
        (8,  "Chad Schuller",      11),
        (9,  "Sandz Hinrichs",     11),
        (10, "Dale Cox",           12),
        (11, "Kevin Graybill",     10),
        (12, "Christian Lindmark", 11),
        (13, "John Culbertson",    12),
        (14, "Kevin Feltus",       10),
    ]
    write_roster_sheet(ws_r, "1993 Gig Harbor Varsity — Roster", roster_rows,
                       coaches=["Peter Jansen", "Kevin Miller", "Bob Toigo"])

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB",
        "HBP", "K", "SAC", "ERR", "TB", "SB", "AVG", "OBP",
    ]
    bat_rows = [
        ("J. Miller",     55, 15, 18, 17, 3, 0, 0,  9, 3,  8, 0, 4, 33,  7, ".327", ".498"),
        ("Cherbas",       58, 21, 29, 17, 4, 1, 0,  9, 5,  9, 0, 7, 37,  7, ".500", ".597"),
        ("Sawyers",       55, 15, 21, 13, 4, 1, 0,  8, 3, 11, 0, 7, 37,  7, ".382", ".506"),
        ("Friedman",      59,  6, 14,  7, 3, 0, 0,  5, 0,  6, 0, 2, 22,  1, ".237", ".333"),
        ("Lippert",       43, 12, 14,  7, 3, 0, 0, 12, 0,  8, 3, 9, 30,  3, ".326", ".473"),
        ("Baurichter",    43, 11,  7, 14, 2, 0, 0,  3, 0,  3, 0, 1,  9,  2, ".172", ".250"),
        ("Vitcovich",     29,  6,  5,  5, 1, 0, 0,  5, 0,  1, 0, 0, 14,  1, ".172", ".250"),
        ("Schuller",      41, 20, 16,  7, 1, 0, 0, 10, 2, 12, 3, 3, 17, 19, ".390", ".526"),
        ("Hinrichs",      24,  4,  5,  5, 4, 3, 0,  5, 0,  7, 0, 1, 14,  0, ".208", ".367"),
        ("Cox",           18,  6,  7,  2, 0, 0, 0,  4, 0,  7, 0, 0,  7,  2, ".389", ".488"),
        ("Graybill",      36, 10, 14, 12, 1, 0, 0,  5, 2,  9, 0, 3, 12,  5, ".200", ".360"),
        ("Lindmark",      20,  6,  4,  2, 1, 0, 0,  1, 1,  7, 0, 0,  5,  0, ".200", ".100"),
        ("Culbertson",    10,  0,  1,  2, 0, 0, 0,  2, 0,  0, 0, 0,  8,  0, ".455", ".538"),
        ("Feltus",        11,  5,  5,  2, 1, 0, 0,  0, 0,  0, 0, 0,  6,  1, ".000", ".000"),
        ("Langford",       1,  0,  0,  0, 0, 0, 0,  1, 0,  0, 0, 0,  0,  1, ".000", "1.000"),
        ("Ford",           0,  0,  0,  0, 0, 0, 0,  0, 1,  0, 0, 0,  0,  0, ".000", ".000"),
        ("Ratliff",        1,  0,  0,  0, 0, 0, 0,  0, 0,  0, 0, 0,  0,  0, ".000", ".000"),
        ("TEAM",         495,137,164,118,26, 3, 0, 81,16, 80,13,45,297, 52, ".331", ".448"),
    ]
    write_table(
        ws2,
        title="1993 Gig Harbor Varsity — Team Batting (Overall 18 games)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Hand-written stats sheet. Some per-player TB and breakdown "
            "cells don't sum cleanly — preserved as printed. Trust Individual "
            "Records sheet for top numbers."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF", "R", "H", "K", "BB", "ER",
        "HBP", "G", "ERA", "OPPBA",
    ]
    pit_rows = [
        ("Cherbas",   0, 1, 3, "15.1",  58,  8,  9,  9, 11,  4, 0,  5,  "1.83", ".155"),
        ("Sawyers",   4, 1, 0, "23.0",  92, 18, 21, 18, 10,  7, 6,  6,  "2.13", ".228"),
        ("Friedman",  0, 1, 0,  "1.2",   8,  2,  3,  1,  2,  2, 0,  1,  "8.40", ".375"),
        ("Lippert",   1, 2, 2, "24.0",  97, 20, 23, 25, 25, 12, 0,  6,  "3.50", ".237"),
        ("Hinrichs",  0, 0, 2, "11.2",  50, 10, 17,  8,  5,  7, 0,  6,  "4.20", ".340"),
        ("Lindmark",  7, 1, 0, "45.1", 187, 28, 45, 50, 16,  8, 5, 10,  "1.24", ".241"),
        ("Ratliff",   0, 0, 0,  "1.0",   7,  5,  3,  2,  1,  4, 0,  1, "28.00", ".429"),
        ("TEAM",     12, 6, 7, "120.0",499, 91,120,113, 70, 44,11, 35,  "2.57", ".241"),
    ]
    write_table(
        ws3,
        title="1993 Gig Harbor Varsity — Team Pitching (Overall 18 games)",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    write_table(ws_h, "Highlights of 1993 Team",
        ["Highlight", "Value", "Context"],
        [
            ("Most Runs (Game)",              15, "vs. Enumclaw, Tahoma"),
            ("Most Runs (Inning)",             8, "vs. Charles Wright, Fife (×2)"),
            ("Most Runs Allowed (Game)",      14, "vs. Tahoma"),
            ("Widest Margin of Victory",      14, "vs. Enumclaw"),
            ("One Run Games Won",              1, "vs. Tahoma"),
            ("One Run Games Lost",             1, "vs. White River"),
            ("Most K's (by GH Pitchers)",     10, "vs. Tahoma, Enumclaw, White River, Washington"),
            ("Most K's (by Opponents)",       10, "vs. Enumclaw"),
            ("Most Hits (Game)",              14, "vs. Tahoma, Enumclaw"),
            ("Most Hits Allowed (Game)",      19, "vs. Tahoma"),
            ("Most Walks (Game for GH)",      11, "vs. Fife"),
            ("Most Walks Allowed",             7, "vs. Fife"),
            ("Most Singles (Game)",           13, "vs. Tahoma"),
            ("Most Doubles (Game)",            5, "vs. Enumclaw"),
            ("Most Triples (Game)",            1, "vs. Charles Wright, Yelm, Franklin Pierce"),
            ("Most Steals (Game)",             7, "vs. Washington"),
            ("Most Errors (Game)",             7, "vs. Peninsula, Washington"),
            ("Longest Winning Streak",         9, ""),
            ("Longest Losing Streak",          3, ""),
        ],
        col_widths=[34, 10, 48], left_align_cols={1, 3})

    ws4 = wb.create_sheet("Individual Records")
    write_table(ws4, "1993 Records (Hall of Fame)",
        ["Record", "Holder", "Value", "Qualifier / Note"],
        [
            ("Highest Average",        "Andy Cherbas",                               ".500", ""),
            ("Most Hits",              "Andy Cherbas",                                29,    ""),
            ("Most At Bats",           "Tom Friedman",                                59,    ""),
            ("Most Runs Scored",       "Andy Cherbas",                                21,    ""),
            ("Most HBP",               "Andy Cherbas",                                 5,    ""),
            ("Most Doubles",           "Andy Cherbas, Ryan Sawyers",                   4,    ""),
            ("Most Triples",           "Andy Cherbas, Ryan Sawyers, Tom Friedman",     1,    ""),
            ("Most Walks",             "Jason Lippert",                               12,    ""),
            ("Most Stolen Bases",      "Chad Schuller",                               19,    ""),
            ("Most Total Bases",       "Andy Cherbas",                                37,    ""),
            ("Most RBIs",              "Andy Cherbas, Joel Miller",                   17,    ""),
            ("Most Wins Pitching",     "Christian Lindmark",                           7,    ""),
            ("Most Innings Pitched",   "Christian Lindmark",                        "45.1",  ""),
            ("Most K's",               "Christian Lindmark",                          50,    ""),
            ("Lowest ERA",             "Christian Lindmark",                        "1.24",  ""),
            ("Most Saves",             "Andy Cherbas",                                 3,    "from pitching sheet"),
            ("Best On-Base Average",   "Andy Cherbas",                              ".597",  ""),
            ("Longest Hitting Streak", "Andy Cherbas",                                 5,    ""),
        ],
        col_widths=[26, 46, 12, 22], left_align_cols={2, 4})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
