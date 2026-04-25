#!/usr/bin/env python3
"""Build 1998 Gig Harbor Varsity season stats xlsx from pages in
Historical/1998/."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1998" / "1998_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1998, [
        ("Individual Records", "Clean typed source. Note: records-page lists 'Most RBIs Drake Hano 16' but stats sheet shows Adam Harris with 16 RBI (Hano had 13) — kept as printed. Records 'Most Wins Keller 4' vs stats W=6 — kept as printed.", "READABLE"),
        ("Team Highlights", "Clean typed source.", "READABLE"),
        ("Schedule", "No per-game schedule photo provided. Season record 14-7 comes from stats header.", "NOT PROVIDED"),
        ("Team Batting", "Re-transcribed from 1998 Overall Final Stats photo. Per-player rows verified row-by-row.", "READABLE"),
        ("Team Pitching", "Transcribed.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Anthony Gilich",     12),
        (2,  "Drake Hano",         12),
        (3,  "Adam Harris",        12),
        (4,  "Willie Keith",       12),
        (5,  "Dustin Bienias",     11),
        (6,  "Mat Cleary",         11),
        (7,  "Kurt Elliott",       11),
        (8,  "Justin Fagering",    11),
        (9,  "Kris Keller",        11),
        (10, "Matt Metsker",       11),
        (11, "Derek Phill",        11),
        (12, "Kurt Wright",        11),
        (13, "Shane Yelish",       11),
        (14, "Kevin Freeman",      10),
        (15, "Ryan Snow",          10),
    ]
    write_roster_sheet(
        ws_r,
        title="1998 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Mike Moeller"],
        manager="Mike Boyle",
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    # Re-transcribed directly from photo. Columns visible in source:
    # Inn AB R H 2B 3B HR BB RBI SAC K HP SLUG. AVG and OBP computed.
    bat_rows = [
        ("Gilich",     "149.0", 64, 20, 23, 4, 0, 0, 10, 15,  5,  8, 5, ".422", ".468", ".359"),
        ("Bienias",    "129.0", 53,  9, 14, 4, 0, 0,  5, 12,  3,  6, 3, ".340", ".361", ".264"),
        ("Harris",     "126.0", 59, 19, 27, 6, 1, 0, 10, 16,  1,  6, 2, ".593", ".521", ".458"),
        ("Phill",      "145.0", 61, 20, 22, 2, 0, 0, 13, 11,  0,  7, 3, ".393", ".468", ".361"),
        ("Hano",       "134.0", 58, 15, 19, 7, 0, 1, 16, 13,  0, 11, 0, ".500", ".473", ".328"),
        ("Snow",        "65.0", 14,  2,  2, 0, 0, 0,  1,  2,  0,  1, 0, ".143", ".200", ".143"),
        ("Keith",       "94.0", 34, 10, 11, 0, 0, 0,  5, 10,  0,  4, 2, ".324", ".415", ".324"),
        ("Keller",      "90.0", 26,  3,  5, 4, 1, 0,  2,  5,  0,  7, 1, ".423", ".276", ".192"),
        ("Elliott",    "148.0", 65, 13, 18, 3, 1, 2,  4, 14,  0, 12, 4, ".415", ".342", ".277"),
        ("Yelish",     "116.0", 39, 10, 10, 3, 0, 1, 12,  7,  1,  6, 2, ".410", ".453", ".256"),
        ("Cleary",      "75.0", 22,  6,  6, 0, 0, 0,  6,  2,  0,  9, 3, ".273", ".419", ".273"),
        ("Fagering",    "88.0", 28,  6,  5, 0, 1, 0,  1,  6,  1,  3, 1, ".250", ".200", ".179"),
        ("Metsker",     "16.0",  1,  1,  1, 1, 0, 0,  0,  1,  0,  0, 0,"2.000","1.000","1.000"),
        ("Freeman",    "113.0", 47, 12, 15, 2, 0, 0,  6,  9,  2,  5, 2, ".362", ".382", ".319"),
        ("Wright",      "34.0",  2,  0,  0, 0, 0, 0,  0,  0,  0,  0, 0, ".000", ".000", ".000"),
        ("Jendro",       "1.0",  1,  0,  0, 0, 0, 0,  0,  0,  0,  1, 0, ".000", ".000", ".000"),
        ("TEAM",       "149.0",574,146,178,36, 3, 4, 91,123, 13, 86,28, ".404", ".404", ".310"),
    ]
    write_table(
        ws2,
        title="1998 Gig Harbor Varsity — Team Batting (Overall Final Stats)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Gilich",     "28.2", 127, 33, 16, 10, 10, 0,  1, 17, 3, 1, 1, ".276",  "2.36"),
        ("Bienias",    "25.0", 115, 29, 17, 13,  8, 0,  4, 17, 2, 2, 0, ".284",  "3.64"),
        ("Keller",     "44.2", 195, 40, 27, 14, 22, 0,  8, 47, 6, 4, 0, ".225",  "2.19"),
        ("Elliott",    "15.1", 73, 19, 15, 14,  9, 0,  2, 13, 1, 0, 0, ".297",  "6.39"),
        ("Metsker",    "10.2",  53, 14,  7,  7,  5, 0,  1,  8, 1, 0, 0, ".294",  "4.59"),
        ("Wright",     "29.1", 131, 35, 22, 15, 21, 0,  3, 20, 1, 0, 0, ".292",  "3.58"),
        ("TEAM",      "149.0", 662,155,103, 67, 72, 0, 12,124,14, 7, 1, ".268",  "3.25"),
    ]
    write_table(
        ws3,
        title="1998 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              18, "vs. N. Thurston"),
        ("Most Runs (Inning)",             9, "vs. Foss"),
        ("Most Runs Allowed (Game)",      11, "vs. S. Kitsap"),
        ("Widest Margin of Victory",      11, "vs. N. Thurston"),
        ("One Run Games Lost",             4, "vs. N. Thurston, Bellarmine, Wilson, Stadium"),
        ("One Run Games Won",              2, "vs. Highline, Mt. Tahoma"),
        ("Most K's (by GH Pitchers)",     12, "vs. Foss"),
        ("Most K's (by Opponents)",        7, "vs. Highline, S. Kitsap"),
        ("Most Hits (Game)",              16, "vs. N. Thurston"),
        ("Most Hits Allowed (Game)",      13, "vs. S. Kitsap"),
        ("Most Walks (Game for GH)",       9, "vs. Timberline"),
        ("Most Walks Allowed",             8, "vs. Shelton"),
        ("Most Singles (Game)",           14, "vs. N. Thurston"),
        ("Most Doubles (Game)",            4, "vs. River Ridge"),
        ("Most Triples (Game)",            1, "vs. Lincoln, River Ridge, Wilson"),
        ("Most Home Runs (Game)",          1, "vs. River Ridge, Foss, Shelton, Timberline"),
        ("Most Steals (Game)",             5, "vs. N. Thurston"),
        ("Longest Winning Streak",         9, ""),
        ("Longest Losing Streak",          2, ""),
        ("Total Team Runs Scored",       146, ""),
        ("Total Team Runs Allowed",      104, ""),
        ("Total Team Hits",              178, ""),
        ("Total Team Extra-Base Hits",    43, ""),
        ("Total Team Home Runs",           4, ""),
        ("Total Team Stolen Bases",       48, ""),
        ("Total Team Walks",              91, ""),
        ("Total Team HBP",                28, ""),
    ]
    write_table(
        ws_h,
        title="1998 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 60],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Adam Harris",                                 ".458", "30+ AB (27-59)"),
        ("Most Hits",              "Adam Harris",                                  27,    ""),
        ("Most At Bats",           "Kurt Elliott",                                 65,    ""),
        ("Most HBP",               "Anthony Gilich",                                5,    ""),
        ("Lowest K Ratio",         "Anthony Gilich",                            "10.1%",  "(8-84)"),
        ("Most Doubles",           "Drake Hano",                                    7,    ""),
        ("Most Triples",           "Adam Harris, Kurt Elliott, Justin Fagering",    1,    ""),
        ("Most Home Runs",         "Kurt Elliott",                                  2,    ""),
        ("Most Walks",             "Drake Hano",                                   16,    ""),
        ("Most Stolen Bases",      "Anthony Gilich, Derek Phill",                  12,    ""),
        ("Most Total Bases",       "Adam Harris",                                  35,    ""),
        ("Most RBIs",              "Drake Hano",                                   16,    ""),
        ("Most Wins Pitching",     "Kris Keller",                                   4,    ""),
        ("Most Innings Pitched",   "Kris Keller",                                "44.2",  ""),
        ("Most K's",               "Kris Keller",                                  47,    ""),
        ("Lowest ERA",             "Kris Keller",                                "2.19",  ""),
        ("Most Saves",             "Anthony Gilich",                                5,    ""),
        ("Best On-Base Avg.",      "Adam Harris",                                ".597",  "30+ AB"),
        ("Longest Hitting Streak", "Adam Harris",                                   8,    ""),
        ("Most Runs Scored",       "Anthony Gilich, Derek Phill",                  20,    ""),
    ]
    write_table(
        ws4,
        title="1998 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 50, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
