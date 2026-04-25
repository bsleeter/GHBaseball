#!/usr/bin/env python3
"""Build 2007 Gig Harbor Varsity season stats xlsx from pages in
Historical/2007/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2007" / "2007_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def _write_readme(ws, year, needs):
    """needs: list of (sheet, issue, status) tuples."""
    ws["A1"] = f"{year} Season Stats — Manual Update Required"
    ws.merge_cells("A1:C1")
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 24
    for c, h in enumerate(("SHEET", "WHAT NEEDS ATTENTION", "STATUS"), start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 22
    for i, (sheet, issue, status) in enumerate(needs):
        for c, val in enumerate((sheet, issue, status), start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            cell.font = Font(name="Arial", size=10, color=NAVY,
                             bold=(c == 3 and status in ("NEEDS MANUAL UPDATE", "NOT PROVIDED")))
            cell.alignment = Alignment(
                horizontal="left" if c != 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = BORDER
            if c == 3 and status == "NEEDS MANUAL UPDATE":
                cell.fill = PatternFill("solid", start_color=FLAG)
            elif c == 3 and status == "NOT PROVIDED":
                cell.fill = PatternFill("solid", start_color="F8D7DA")
            elif i % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT)
        ws.row_dimensions[4 + i].height = 48
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 0: README ────────────────────────────────────────────
    ws_n = wb.active
    ws_n.title = "README"
    _write_readme(ws_n, 2007, [
        ("Individual Records",
         "Source page listed records cleanly; 'Most Triples' had no value in the source. Verify all records against Team Batting/Pitching sheets before publishing.",
         "NEEDS MANUAL UPDATE"),
        ("Team Highlights",
         "No Team Highlights page was provided for 2007.",
         "NOT PROVIDED"),
        ("Schedule",
         "No per-game schedule photo was provided for 2007. Season record (9-12) comes from the pitching report's W-L totals.",
         "NOT PROVIDED"),
        ("Team Batting",
         "CoachStat report transcribed. Per-player 1B/2B/3B/HR breakdowns sum cleanly to published TB. Records page lists Buckles 23 RBI, matching stats sheet.",
         "READABLE"),
        ("Team Pitching",
         "CoachStat report transcribed. Team totals match source (9-12, 5.01 ERA).",
         "READABLE"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Nick Mareno",          12),
        (2,  "Jason Van Skike",      12),
        (3,  "Derek Speigner",       12),
        (4,  "Michael Gaube",        12),
        (5,  "Ryan Holsten",         12),
        (6,  "Ryan Buckles",         12),
        (7,  "Brandon Rohde",        11),
        (8,  "Drew Young",           11),
        (9,  "Kyle Mauren",          11),
        (10, "Steve Savage",         11),
        (11, "Alex Rose",            11),
        (12, "Bubba Brown",          11),
        (13, "Tyler Girdler",        11),
        (14, "Chet Thompson",        10),
        (15, "Cameron Holcomb",      10),
        (16, "Mike Barnett",         10),
        (17, "David Bigelow",         9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2007 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "BB", "R", "H", "1B", "2B", "3B", "HR", "RBI",
        "AVG", "SLG", "OBP", "HBP", "SAC-B", "SAC-F", "SO", "TB",
    ]
    bat_rows = [
        ("Savage",     32,  7, 11,  7,  2, 4, 0, 1,  5, ".219", ".438", ".366", 1, 0, 1, 6, 14),
        ("Mareno",     72,  6, 20, 31, 26, 3, 0, 2, 14, ".431", ".556", ".481", 1, 0, 0, 3, 40),
        ("Thompson",   52,  5, 14, 19, 18, 1, 0, 0, 15, ".365", ".385", ".429", 0, 0, 3,13, 20),
        ("Buckles",    60, 12, 22, 25, 11, 8, 0, 6, 23, ".417", ".850", ".507", 1, 0, 0, 5, 51),
        ("Rose",       39,  2,  5,  8,  7, 1, 0, 0,  6, ".205", ".231", ".262", 0, 0, 0, 0,  9),
        ("Bigelow",    63,  7, 19, 20, 15, 4, 0, 1, 11, ".317", ".429", ".435", 5, 0, 0, 2, 27),
        ("Girdler",    17,  1,  4,  7,  5, 2, 0, 0,  4, ".412", ".529", ".474", 0, 0, 0, 3,  9),
        ("Holsten",    33,  3,  3,  8,  4, 4, 0, 0,  3, ".242", ".364", ".306", 0, 0, 0, 0, 12),
        ("Young",      30,  3,  1,  7,  7, 0, 0, 0,  5, ".233", ".233", ".303", 0, 0, 0, 5,  7),
        ("Munoz",      34,  4, 11,  6,  5, 1, 0, 0,  3, ".176", ".206", ".326", 4, 1, 1, 7,  7),
        ("Gaube",      56,  5, 12, 17, 15, 2, 0, 0,  5, ".304", ".339", ".359", 0, 1, 2, 2, 19),
        ("Rohde",       0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0, 0, 0, 0,  0),
        ("Holcomb",    45,  7,  5, 18, 16, 2, 0, 0,  6, ".400", ".444", ".491", 1, 0, 0, 7, 20),
        ("Speigner",   15,  1,  7,  4,  4, 0, 0, 0,  2, ".267", ".267", ".333", 0, 0, 0, 5,  4),
        ("Brown",       8,  0,  1,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0, 0, 0, 3,  0),
        ("VanSkike",    3,  0,  1,  1,  1, 0, 0, 0,  1, ".333", ".333", ".333", 0, 0, 0, 1,  1),
        ("Barnett",     8,  0,  2,  2,  2, 0, 0, 0,  0, ".250", ".250", ".250", 0, 0, 0, 1,  2),
        ("Mauren",      2,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0, 0, 0, 0,  0),
        ("TEAM",      567, 67,136,184,137,36, 0,11,114, ".325", ".446", ".408",17, 4, 8,92,253),
    ]
    write_table(
        ws2,
        title="2007 Gig Harbor Varsity — Team Batting",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "#BF", "RS", "ER", "BB", "H", "HBP",
        "W", "L", "SV", "BAA", "ERA", "SO",
    ]
    pit_rows = [
        ("Mareno",     "3.2",   28,  8,  6,  7,  5, 1, 0, 1, 0, ".250",  "11.45",  2),
        ("Thompson",   "2.0",   15,  5,  3,  3,  0, 0, 0, 0, 0, ".417",  "10.50",  2),
        ("Buckles",    "8.0",   48, 10,  9, 12, 12, 2, 0, 1, 0, ".250",   "7.88", 10),
        ("Rose",       "0.1",    5,  1,  1,  0,  1, 1, 0, 0, 0, ".182",   "3.94",  4),
        ("Bigelow",    "5.1",   28,  5,  3,  5,  4, 4, 1, 0, 0, ".250",  "14.00",  0),
        ("Holsten",    "1.0",    8,  2,  2,  4,  1, 0, 0, 0, 0, ".250",  "11.45",  0),
        ("Speigner",  "37.2",  189, 23, 23, 27, 47, 8, 3, 3, 2, ".287",   "4.27", 23),
        ("Brown",     "36.0",  168, 24, 20, 25, 32, 5, 4, 2, 0, ".232",   "3.89", 30),
        ("VanSkike",  "37.2",  183, 35, 25, 21, 34, 7, 2, 4, 0, ".221",   "4.65", 23),
        ("Mauren",     "1.0",    6,  3,  3,  1,  1, 0, 0, 0, 0, ".200",  "21.00",  0),
        ("TEAM",     "132.2",  687,132, 95,105,142,25, 9,12, 2, ".255",   "5.01", 92),
    ]
    write_table(
        ws3,
        title="2007 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Nick Mareno",                           ".431", "30+ AB (31-72)"),
        ("Most At Bats",           "Nick Mareno",                            72,    ""),
        ("Most HBP",               "David Bigelow",                           5,    ""),
        ("Lowest K Ratio",         "Nick Mareno",                          "3.8%",  "(3-79)"),
        ("Most Doubles",           "Ryan Buckles",                            8,    ""),
        ("Most Triples",           "(no value recorded)",                    "—",   "source sheet value missing"),
        ("Most Home Runs",         "Ryan Buckles",                            6,    ""),
        ("Most Walks",             "Ryan Buckles",                           12,    ""),
        ("Most Stolen Bases",      "Nick Mareno",                             8,    ""),
        ("Most Total Bases",       "Ryan Buckles",                           51,    ""),
        ("Most RBIs",              "Ryan Buckles",                           23,    ""),
        ("Most Wins Pitching",     "Bubba Brown",                             4,    ""),
        ("Most Innings Pitched",   "Derek Speigner, Jason Van Skike",       "37.2",  ""),
        ("Most K's",               "Bubba Brown",                            30,    ""),
        ("Lowest ERA",             "Bubba Brown",                          "3.89",  ""),
        ("Most Saves",             "Derek Speigner",                          2,    ""),
        ("Best On-Base Avg.",      "Ryan Buckles",                         ".507",  "30+ AB"),
        ("Longest Hitting Streak", "Cameron Holcomb",                         7,    ""),
        ("Most Runs Scored",       "Ryan Buckles",                           22,    ""),
    ]
    write_table(
        ws4,
        title="2007 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 38, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
