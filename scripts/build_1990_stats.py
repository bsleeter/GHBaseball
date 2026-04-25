#!/usr/bin/env python3
"""Build 1990 Gig Harbor Varsity season stats xlsx."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1990" / "1990_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1990, [
        ("Overall Statistics",
         "Clean typed source with batting and pitching stats on one page.",
         "READABLE"),
        ("Team Highlights",
         "No Team Highlights page provided for this year.",
         "NOT PROVIDED"),
        ("Individual Records",
         "No Individual Records page existed for 1990. Records here were derived directly from the Overall Statistics — leaders in each category.",
         "NEEDS SPOT-CHECK"),
        ("Schedule",
         "No schedule photo provided.",
         "NOT PROVIDED"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Jay Brown",          12),   # displayed grade was 11; grades column on sheet was hard to read
        (2,  "Travis Brown",       11),
        (3,  "Jason Dupuis",       11),
        (4,  "Mark Evenson",       12),
        (5,  "Ed Hazel",           12),
        (6,  "Donnie Jones",       12),
        (7,  "Ian Priestley",      12),
        (8,  "Sig Siegmund",       10),
        (9,  "Shelby Zamberlin",   12),
        (10, "John Waage",         12),
        (11, "Scott Crawford",     10),
        (12, "Brad Fazio",         12),
        (13, "Lonnie Ledbetter",   11),
        (14, "Billy Kirk",         10),
        (15, "Jim Peshek",         10),
    ]
    write_roster_sheet(ws_r, "1990 Gig Harbor Varsity — Roster", roster_rows,
                       coaches=["Peter Jansen", "Craig Sugai"],
                       manager="Dave Taylor (Batboy: Jon Erickson)")

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB", "HBP",
        "K", "SB", "SAC", "TB", "E", "AVG", "G",
    ]
    bat_rows = [
        ("J. Brown",    44,  7, 14, 17, 2, 0, 1,  4,  0, 12, 6, 1, 23,  1, ".318", 18),
        ("T. Brown",     6,  1,  1,  0, 0, 0, 0,  2,  0,  5, 0, 0,  3,  0, ".167", 10),
        ("Dupuis",      56, 15, 22,  6, 2, 0, 0,  8,  2,  8, 7, 0, 34,  8, ".393", 17),
        ("Evenson",     13,  4,  2,  1, 0, 0, 0,  4,  0,  4, 4, 0,  7,  0, ".154", 11),
        ("Hazel",       61, 14, 24, 16, 4, 3, 0,  3,  1,  6, 8, 0, 37,  9, ".393", 17),
        ("Jones",       37,  7, 10,  4, 1, 0, 0,  8,  1,  5, 1, 0, 18,  5, ".270", 18),
        ("Priestley",   34,  8,  8, 10, 1, 0, 0,  9,  1,  6, 6, 5, 19,  1, ".235", 18),
        ("Siegmund",    41,  5, 13,  9, 0, 0, 0,  7,  0,  6, 0, 0, 19,  4, ".317", 18),
        ("Zamberlin",   54, 18, 18,  4, 0, 1, 0, 15,  3,  5,15, 1, 39,  4, ".333", 18),
        ("Waage",       58, 10, 18, 10, 2, 0, 0,  6,  0,  8, 1, 1, 23,  4, ".310", 17),
        ("Crawford",    53,  8, 16,  7, 4, 0, 0,  4,  0,  5, 2, 0, 25,  1, ".302", 18),
        ("Fazio",       40,  5,  8,  3, 4, 0, 0,  5,  1, 14, 3, 0, 19,  6, ".200", 18),
        ("Ledbetter",   16,  2,  1,  0, 0, 0, 0,  4,  0,  2, 2, 0,  5,  2, ".063", 15),
        ("Kirk",        12,  3,  1,  1, 1, 0, 0,  3,  0,  2, 0, 0,  6,  0, ".083",  8),
        ("Peschek",      3,  0,  1,  0, 0, 0, 0,  0,  0,  1, 0, 0,  0,  0, ".000",  3),
    ]
    write_table(ws2, "1990 Gig Harbor Varsity — Overall Batting Statistics",
                bat_headers, bat_rows,
                col_widths=[14] + [6] * (len(bat_headers) - 2) + [7],
                left_align_cols={1})

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = ["Player", "W", "L", "IP", "R", "ER", "H", "BB", "K", "HBP", "ERA", "SV", "Games"]
    pit_rows = [
        ("Crawford",   3, 1, "33.1",  22, 16, 39, 15, 38, 4, "3.36", 0, 9),
        ("Fazio",      1, 2, "18.2",  21, 11, 14, 15, 11, 1, "4.12", 0, 6),
        ("Ledbetter",  0, 1, "4.1",    5,  4,  4,  3,  2, 1, "6.47", 0, 2),
        ("Brown",      3, 4, "52.0",  33, 26, 53, 25, 34, 1, "3.50", 1,12),
        ("Siegmund",   2, 1, "16.0",   7,  5, 13,  7,  9, 2, "2.19", 2, 9),
    ]
    write_table(ws3, "1990 Gig Harbor Varsity — Overall Pitching Statistics",
                pit_headers, pit_rows,
                col_widths=[14] + [6] * (len(pit_headers) - 1),
                left_align_cols={1})

    ws4 = wb.create_sheet("Individual Records")
    write_table(ws4, "1990 Gig Harbor Varsity — Individual Records (derived)",
        ["Record", "Holder", "Value", "Qualifier / Note"],
        [
            ("Highest Average",        "Jason Dupuis, Ed Hazel",              ".393", "30+ AB · Dupuis 22/56, Hazel 24/61 (tied)"),
            ("Most Hits",              "Ed Hazel",                             24,    ""),
            ("Most At Bats",           "Ed Hazel",                             61,    ""),
            ("Most Runs Scored",       "Shelby Zamberlin",                     18,    ""),
            ("Most HBP",               "Shelby Zamberlin",                      3,    ""),
            ("Lowest K Ratio",         "Shelby Zamberlin",                   "9.3%",  "(5-54) · 30+ AB"),
            ("Most Doubles",           "Ed Hazel, Scott Crawford, Brad Fazio",  4,    "3-way tie"),
            ("Most Triples",           "Ed Hazel",                              3,    ""),
            ("Most Home Runs",         "Jay Brown",                             1,    "only player with HR"),
            ("Most Walks",             "Shelby Zamberlin",                     15,    ""),
            ("Most Stolen Bases",      "Shelby Zamberlin",                     15,    ""),
            ("Most Total Bases",       "Shelby Zamberlin",                     39,    ""),
            ("Most RBIs",              "Jay Brown",                            17,    ""),
            ("Best On-Base Avg.",      "Shelby Zamberlin",                   ".493",  "30+ AB · computed"),
            ("Most Wins Pitching",     "Scott Crawford, Jay Brown",             3,    "tied"),
            ("Most Innings Pitched",   "Jay Brown",                          "52.0",  ""),
            ("Most K's",               "Scott Crawford",                       38,    ""),
            ("Lowest ERA",             "Sig Siegmund",                       "2.19",  "14+ IP qualifier"),
            ("Most Saves",             "Sig Siegmund",                          2,    ""),
        ],
        col_widths=[26, 38, 12, 50], left_align_cols={2, 4},
        note=(
            "Records derived directly from the 1990 Overall Statistics page "
            "(no dedicated records page existed). Some TB column values in "
            "the source don't match the visible 2B/3B/HR breakdowns — TB "
            "values preserved as printed. Hit streak not derivable from "
            "cumulative stats."
        ))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
