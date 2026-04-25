#!/usr/bin/env python3
"""Build 2006 Gig Harbor Varsity season stats xlsx from pages in
Historical/2006/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2006" / "2006_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def _write_readme(ws, year, needs):
    ws["A1"] = f"{year} Season Stats — Manual Update Required"
    ws.merge_cells("A1:C1")
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 24
    for c, h in enumerate(("SHEET", "WHAT NEEDS ATTENTION", "STATUS"), start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 22
    for i, (sheet, issue, status) in enumerate(needs):
        for c, val in enumerate((sheet, issue, status), start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            cell.font = Font(name="Arial", size=10, color=NAVY,
                             bold=(c == 3 and status in ("NEEDS MANUAL UPDATE", "NOT PROVIDED")))
            cell.alignment = Alignment(
                horizontal="left" if c != 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = BORDER
            if c == 3 and status == "NEEDS MANUAL UPDATE":
                cell.fill = PatternFill("solid", start_color=FLAG)
            elif c == 3 and status == "NOT PROVIDED":
                cell.fill = PatternFill("solid", start_color="F8D7DA")
            elif i % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT)
        ws.row_dimensions[4 + i].height = 48
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 0: README ────────────────────────────────────────────
    ws_n = wb.active
    ws_n.title = "README"
    _write_readme(ws_n, 2006, [
        ("Individual Records",
         "Clean typed source — 18 records transcribed. Verify against Team Batting/Pitching before publishing.",
         "NEEDS MANUAL UPDATE"),
        ("Team Highlights",
         "Clean typed source. Values preserved as printed.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "No per-game schedule photo was provided for 2006. Season record (15-6-1) comes from the CoachStat stats page header.",
         "NOT PROVIDED"),
        ("Team Batting",
         "CoachStat report transcribed. Some per-player breakdowns don't sum cleanly to published AVG/TB — preserved as printed.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "CoachStat report transcribed. Team totals match source.",
         "READABLE"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Troy Burki",            12),
        (2,  "Rob Emmett",             12),
        (3,  "David Benedict",         12),
        (4,  "Owen MacDonald",         12),
        (5,  "Brent Smoots",           12),
        (6,  "Nolan Smith",            12),
        (7,  "Antone Saltvick",        12),
        (8,  "Jordan Harrison",        12),
        (9,  "DJ Myers",               12),
        (10, "Casey Knox",             12),
        (11, "Josiah Ward",            12),
        (12, "Nick Mareno",            11),
        (13, "Derek Speigner",         11),
        (14, "Michael Gaube",          11),
        (15, "Ryan Buckles",           11),
        (16, "Brandon Rohde",          10),
        (17, "Drew Young",             10),
        (18, "Chet Thompson",           9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2006 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Jim Peschek"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "BB", "R", "H", "1B", "2B", "3B", "HR", "RBI",
        "AVG", "SLG", "OBP", "HBP", "SAC-B", "SAC-F", "SO", "TB",
    ]
    bat_rows = [
        ("Mareno",      71, 11, 25, 26, 17, 8, 1, 0, 13, ".366", ".521", ".500", 7, 2, 1, 9, 37),
        ("Benedict",    69,  9, 17, 28, 23, 5, 0, 0, 15, ".406", ".478", ".500", 4, 2, 5, 3, 33),
        ("MacDonald",   58,  7, 16, 19, 17, 2, 0, 0, 17, ".328", ".379", ".400", 6, 3, 0, 3, 23),
        ("Smoots",      70,  6, 13, 18, 17, 1, 0, 0, 17, ".257", ".271", ".325", 1, 1, 0,12, 19),
        ("Rohde",       41,  4, 11,  7,  7, 0, 1, 0,  6, ".171", ".220", ".260", 1, 0, 0, 7,  9),
        ("Emmett",      39,  9, 19, 14,  8, 4, 0, 2, 19, ".359", ".667", ".500", 1, 0, 0, 7, 26),
        ("Smith",       57, 13, 12, 19,  9, 2, 1, 1, 16, ".333", ".474", ".458", 2, 4, 3, 5, 27),
        ("Burki",       29,  7,  9, 12,  6, 2, 1, 3,  7, ".414", ".862", ".514", 0, 1, 0, 0, 25),
        ("Knox",        13,  1,  2,  2,  1, 1, 0, 0,  1, ".154", ".231", ".214", 0, 0, 0, 4,  3),
        ("Myers",       10,  1,  1,  1,  1, 0, 0, 0,  1, ".100", ".100", ".167", 0, 0, 0, 1,  1),
        ("Ward",        53, 11, 21, 24, 16, 5, 0, 3, 12, ".453", ".660", ".538", 5, 4, 0, 4, 35),
        ("Gaube",       33,  6, 11, 11, 11, 0, 0, 0,  6, ".333", ".333", ".447", 1, 1, 0, 5, 11),
        ("Speigner",    13,  3,  2,  2,  2, 0, 0, 0,  0, ".154", ".154", ".313", 1, 0, 0, 4,  2),
        ("Harrison",    10,  3,  3,  2,  2, 0, 0, 0,  1, ".200", ".200", ".385", 1, 0, 0, 3,  2),
        ("Saltvick",    26,  6, 10,  6,  6, 0, 0, 0,  4, ".231", ".231", ".375", 0, 1, 0, 6,  6),
        ("Buckles",      9,  2,  3,  2,  1, 0, 0, 1,  3, ".222", ".556", ".364", 0, 0, 0, 4,  5),
        ("Young",        2,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0, 0, 0, 2,  0),
        ("Thompson",     4,  0,  2,  1,  1, 0, 0, 0,  0, ".250", ".250", ".250", 0, 0, 0, 1,  1),
        ("TEAM",       599, 99,158,198,134,48, 1,13,145, ".331", ".476", ".437",21, 3, 8,23,285),
    ]
    write_table(
        ws2,
        title="2006 Gig Harbor Varsity — Team Batting",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Per-player values transcribed from CoachStat report. Some "
            "hit breakdowns don't sum cleanly to published AVG/TB — preserved "
            "as printed. Trust Individual Records sheet for top numbers."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "#BF", "RS", "ER", "BB", "H", "HR", "HBP",
        "W", "L", "SV", "BAA", "ERA", "GP", "SO",
    ]
    pit_rows = [
        ("Mareno",      "1.0",    5,  1,  1,  1,  1, 0, 0, 0, 0, 0, ".250",  "7.00",  1,  1),
        ("Rohde",      "28.1",  135, 21, 10, 16,  6, 0, 0, 2, 2, 1, ".286",  "2.47", 13, 31),
        ("Emmett",      "6.2",   30,  6,  2,  9,  3, 0, 0, 2, 1, 1, ".200",  "2.10",  4,  6),
        ("Burki",      "43.0",  195, 24, 19, 39, 21, 2, 3, 6, 2, 1, ".170",  "3.09", 13, 68),
        ("Knox",        "6.2",   32,  4,  3,  5,  0, 2, 0, 0, 0, 0, ".280",  "3.15",  3,  5),
        ("Myers",       "9.0",   46, 10,  7,  4,  3, 1, 2, 1, 1, 0, ".308",  "5.44",  6,  4),
        ("Speigner",   "17.0",   76, 12,  4,  3,  1, 1, 1, 3, 1, 0, ".286",  "1.65",  9, 15),
        ("Harrison",   "15.1",   80, 16, 11,  5,  4, 1, 2, 1, 1, 1, ".219",  "5.02",  9, 11),
        ("Saltvick",   "20.0",   88, 11, 10,  7,  6, 2, 1, 1, 1, 1, ".190",  "3.50",  9, 20),
        ("TEAM",      "147.0",  687, 99, 67, 83, 43, 3,14,15, 6, 5, ".233",  "3.19", 67,153),
    ]
    write_table(
        ws3,
        title="2006 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              18, "vs. Bremerton"),
        ("Most Runs (Inning)",            12, "vs. NK"),
        ("Most Runs Allowed (Game)",      14, "vs. NK"),
        ("Widest Margin of Victory",      14, "vs. Bremerton, Mt. Tahoma"),
        ("One Run Games Lost",             2, "vs. SK, Peninsula"),
        ("One Run Games Won",              2, "vs. PA, Olympic"),
        ("Most K's (by GH Pitchers)",     12, "vs. CK, Stadium"),
        ("Most K's (by Opponents)",        9, "vs. CK"),
        ("Most Hits (Game)",              17, "vs. CK"),
        ("Most Hits Allowed (Game)",      12, "vs. SK, NK"),
        ("Most Walks (Game for GH)",      12, "vs. Shelton"),
        ("Most Walks Allowed",             9, "vs. Bellarmine"),
        ("Most Singles (Game)",           12, "vs. CK"),
        ("Most Doubles (Game)",            5, "vs. SK"),
        ("Most Triples (Game)",            1, "vs. Olympic"),
        ("Most Home Runs (Game)",          2, "vs. Bremerton, Shelton"),
        ("Most Steals (Game)",             7, "vs. Bremerton"),
        ("Longest Winning Streak",         9, ""),
        ("Longest Losing Streak",          2, ""),
        ("Total Team Runs Scored",       158, ""),
        ("Total Team Hits",              196, ""),
        ("Total Team Extra-Base Hits",    62, ""),
        ("Total Team Home Runs",          13, ""),
        ("Total Team Stolen Bases",       38, ""),
        ("Total Team Walks",             101, ""),
        ("Total Team HBP",                23, ""),
    ]
    write_table(
        ws_h,
        title="2006 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Josiah Ward",                                                      ".453", "30+ AB (24-53)"),
        ("Most At Bats",           "Nick Mareno",                                                       71,    ""),
        ("Most HBP",               "Nick Mareno",                                                        7,    ""),
        ("Lowest K Ratio",         "David Benedict",                                                  "4.6%",  "(4-87)"),
        ("Most Doubles",           "Nick Mareno",                                                        8,    ""),
        ("Most Triples",           "Troy Burki",                                                         1,    ""),
        ("Most Home Runs",         "Troy Burki, Ryan Buckles",                                           3,    ""),
        ("Most Walks",             "Nolan Smith",                                                       13,    ""),
        ("Most Stolen Bases",      "Josiah Ward",                                                       13,    ""),
        ("Most Total Bases",       "Nick Mareno",                                                       37,    ""),
        ("Most RBIs",              "Brent Smoots, Owen MacDonald",                                      17,    ""),
        ("Most Wins Pitching",     "Troy Burki",                                                         6,    ""),
        ("Most Innings Pitched",   "Troy Burki",                                                        43,    ""),
        ("Most K's",               "Troy Burki",                                                        68,    ""),
        ("Lowest ERA",             "Derek Speigner",                                                  "1.65",  ""),
        ("Most Saves",             "Brandon Rohde, Rob Emmett, DJ Myers, Troy Burki, Antone Saltvick",   1,    "5-way tie"),
        ("Best On-Base Avg.",      "Josiah Ward",                                                    ".538",  "30+ AB"),
        ("Longest Hitting Streak", "Owen MacDonald, Brent Smoots, Nick Mareno, Josiah Ward",             5,    "4-way tie"),
        ("Most Runs Scored",       "Nick Mareno",                                                       25,    ""),
    ]
    write_table(
        ws4,
        title="2006 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 68, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
