#!/usr/bin/env python3
"""Build 2017 Gig Harbor Varsity season stats xlsx from scanned records
in Historical/2017/. 2017 was the state championship year. Includes an
extra Roster sheet with player grades."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2017" / "2017_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
GOLD = "D4AF37"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_section_header(cell, gold=False):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE, italic=True)
    cell.fill = PatternFill("solid", start_color=GOLD if gold else CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def write_table(ws, title, headers, rows, team_row_index=None,
                section_rows=None, col_widths=None, left_align_cols=None):
    """section_rows: dict of {row_index_in_rows: (label, is_gold)}"""
    section_rows = section_rows or {}
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20

    r = 4
    for i, row in enumerate(rows):
        if i in section_rows:
            label, is_gold = section_rows[i]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            sc = ws.cell(row=r, column=1, value=label)
            style_section_header(sc, gold=is_gold)
            r += 1
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Schedule"
    schedule_rows = [
        ("Capital",     "0-2"),
        ("Shelton",     "10-0"),
        ("O'Dea",       "8-7"),
        ("Timberline",  "3-5"),
        ("Timberline",  "13-1"),
        ("Curtis",      "9-2"),
        ("Wilson",      "4-1"),
        ("C. Kitsap",   "5-4"),
        ("Capital",     "4-2"),
        ("C. Kitsap",   "13-6"),
        ("Bellevue",    "3-8"),
        ("Peninsula",   "11-0"),
        ("Peninsula",   "7-1"),
        ("S. Kitsap",   "1-0"),
        ("Yelm",        "11-1"),
        ("Yelm",        "4-1"),
        ("Shelton",     "11-0"),
        ("N. Thurston", "6-1"),
        ("Bellevue",    "4-0"),
        ("Bellarmine",  "0-1"),
        ("Lincoln",     "3-0"),
        ("Kelso",       "10-7"),
        ("Mt. View",    "4-1"),
        ("Timberline",  "1-0"),
        # State Tournament
        ("Bainbridge",    "6-3"),
        ("Ed.-Woodway",   "5-0"),
        ("Southridge",    "5-1"),
        ("Season Record", "24-4 · STATE CHAMPIONS"),
    ]
    write_table(
        ws,
        title="2017 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        section_rows={24: ("WA 3A State Tournament", True)},
        col_widths=[22, 26],
        left_align_cols={1},
    )

    # ─── Sheet 2: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Logan Gerling",        12),
        (2,  "Jacob Bonnell",        12),
        (3,  "Cameron MacIntosh",    12),
        (4,  "Chad Stevens",         12),
        (5,  "Avery Jones",          12),
        (6,  "Patrick Fredrickson",  12),
        (7,  "Cameron Brooks",       12),
        (8,  "Johnny Schmidt",       12),
        (9,  "Tanner Hardy",         12),
        (10, "RJ Green",             12),
        (11, "Mason Gibson",         11),
        (12, "Jordan Haworth",       11),
        (13, "Austin Dempewolf",     11),
        (14, "Kirin Peterson",       11),
        (15, "Bodi Tisch",           11),
        (16, "Shannon Milbourn",     11),
        (17, "Cage Hardy",           10),
        (18, "Cole Smith",           10),
        (19, "Cade Dessert",          9),
    ]
    write_table(
        ws_r,
        title="2017 Gig Harbor Varsity — Roster",
        headers=["#", "Player", "Grade"],
        rows=roster_rows,
        col_widths=[6, 28, 10],
        left_align_cols={2},
    )
    # Coaches note below the roster
    last_row = 3 + len(roster_rows) + 2
    ws_r.cell(row=last_row, column=1,
              value="COACHES").font = Font(name="Arial", size=10, bold=True, color=NAVY)
    for i, coach in enumerate(["Pete Jansen", "Larry Carlson", "Larry Roehr", "Oscar Ortiz"]):
        c = ws_r.cell(row=last_row + 1 + i, column=1, value=coach)
        c.font = Font(name="Arial", size=10, color=NAVY)

    # ─── Sheet 3: Hitting & Fielding ────────────────────────────────
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Bonnell",      2,  1,  1,  1,  1, 0, 0, 0,  1,  ".500", ".500", ".500",  0,  0, 0, 0,   2,   0,  0, "1.000",  0),
        ("Brooks",       83, 17, 21, 12, 19, 1, 1, 0, 24,  ".333", ".381", ".382",  9,  5, 1, 0,  22,   3,  1,  ".962",  1),
        ("Dempewolf",    70, 11, 17, 16, 12, 2, 0, 3, 28,  ".243", ".400", ".369", 15, 14, 2, 2, 149,   8,  6,  ".963",  1),
        ("Dessert",      26,  4,  6,  3,  5, 1, 0, 0,  7,  ".231", ".269", ".310",  6,  3, 0, 0,   0,   0,  0, "1.000",  1),
        ("Fredrickson",   0,  0,  0,  0,  0, 0, 0, 0,  0,  ".000", ".000", ".000",  0,  0, 0, 0,   7,   0,  1,  ".875",  0),
        ("Gerling",       1,  0,  1,  0,  1, 0, 0, 0,  1, "1.000","1.000","1.000",  0,  0, 0, 0,   4,   8,  3,  ".800",  0),
        ("Gibson",       11,  3,  2,  2,  2, 1, 0, 0,  3,  ".182", ".273", ".400",  3,  4, 0, 0,   8,   4,  2,  ".857",  1),
        ("Green",        91, 21, 31, 25, 21, 5, 1, 0, 38,  ".341", ".418", ".464",  2, 21, 2, 4,  40,   0,  0, "1.000", 12),
        ("Hardy, C",      4,  0,  1,  1,  1, 0, 0, 0,  1,  ".250", ".250", ".250",  2,  0, 0, 0,   0,   0,  0,  ".000",  0),
        ("Hardy, T",     73, 14, 19, 11, 16, 3, 0, 0, 22,  ".260", ".301", ".349", 10,  8, 2, 3,  25,   1,  1,  ".963",  6),
        ("Haworth",      73, 21, 27, 17, 20, 5, 1, 1, 37,  ".370", ".507", ".471",  6,  8, 6, 3,  13,  36,  6,  ".891",  1),
        ("Jones",         0,  0,  0,  0,  0, 0, 0, 0,  0,  ".000", ".000", ".000",  0,  0, 0, 0,  12,   0,  0, "1.000",  0),
        ("MacIntosh",    88, 21, 33, 19, 20,10, 3, 0, 49,  ".375", ".557", ".450",  2, 11, 1, 1,   6,  28,  3,  ".919",  9),
        ("Milbourn",     13,  4,  3,  3,  3, 0, 0, 0,  3,  ".231", ".231", ".444",  5,  4, 1, 0,   0,   0,  0, "1.000",  2),
        ("Peterson",      8,  2,  1,  2,  0, 0, 0, 1,  4,  ".125", ".500", ".364",  4,  2, 1, 0,   7,   4,  0, "1.000",  0),
        ("Schmidt",      30,  3, 10,  4, 10, 0, 0, 0, 10,  ".333", ".333", ".412",  7,  3, 1, 0,  21,   1,  0, "1.000",  0),
        ("Smith",        51, 10, 15,  7, 11, 4, 0, 0, 19,  ".294", ".314", ".455", 10, 12, 3, 0,  21,  32,  3,  ".954",  5),
        ("Stevens",      80, 18, 25, 21, 16, 5, 3, 0, 39,  ".313", ".488", ".415",  8, 11, 3, 3,  24,  70,  6,  ".940",  5),
        ("Tisch",        37,  8,  7,  4,  6, 1, 0, 0,  8,  ".189", ".216", ".302", 10,  6, 0, 0,  27,   2,  1,  ".967",  2),
        ("TEAM",        721,158,220,144,170,35, 6, 8,291,  ".305", ".404", ".411", 99,112,18,18, 374, 219, 35,  ".944", 45),
    ]
    write_table(
        ws2,
        title="2017 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=[14] + [7] * (len(hit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Pitching ──────────────────────────────────────────
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF/AB", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Bonnell",      1, 1, 0, "9.0",  44,  6, 10,  6,  7, 4,  8, 2, 0, 0, 0, "3.111"),
        ("Brooks",       0, 0, 0, "5.1",  23,  2,  6,  2,  2, 2,  4, 2, 0, 0, 0, "2.745"),
        ("Dempewolf",    3, 0, 0, "13.0", 58,  5,  6, 15, 10, 4,  3, 3, 0, 0, 1, "2.154"),
        ("Fredrickson",  5, 2, 2, "50.1", 201, 14, 38, 42, 14,12, 30, 5, 3, 0, 2, "1.677"),
        ("Gerling",      6, 0, 0, "41.2", 165, 10, 21, 39, 18, 8, 17, 4, 0, 0, 5, "1.359"),
        ("Gibson",       1, 1, 0, "8.2",  41,  6,  6,  7, 11, 4,  5, 0, 0, 1, 0, "3.415"),
        ("Jones",        7, 0, 2, "48.1", 194, 11, 35, 63,  8, 7, 27, 6, 1, 1, 2, "1.019"),
        ("Schmidt",      1, 0, 1, "8.2",  38,  4,  7, 14,  4, 4,  5, 1, 0, 1, 0, "3.415"),
        ("TEAM",        24, 4, 5, "185.0", 764, 58, 129, 188, 74, 45, 99, 23, 4, 3, 10, "1.703"),
    ]
    write_table(
        ws3,
        title="2017 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Cameron MacIntosh",              ".375", "30+ AB (33-88)"),
        ("Most Plate Appearances", "RJ Green",                        114,   "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",         "RJ Green",                       "1.8%", "(2-114)"),
        ("Most Doubles",           "Cameron MacIntosh",               10,    ""),
        ("Most Triples",           "Chad Stevens, Cameron MacIntosh",  3,    ""),
        ("Most Home Runs",         "Austin Dempewolf",                 3,    ""),
        ("Most Walks",             "RJ Green",                        21,    ""),
        ("Most Stolen Bases",      "RJ Green",                        12,    ""),
        ("Most Total Bases",       "Cameron MacIntosh",               49,    ""),
        ("Most RBIs",              "RJ Green",                        25,    ""),
        ("Best On-Base Avg.",      "Jordan Haworth",                 ".471", "30+ AB"),
        ("Longest Hitting Streak", "RJ Green",                         6,    ""),
        ("Most Runs Scored",       "Cameron MacIntosh, RJ Green",     21,    ""),
        ("Most Wins Pitching",     "Avery Jones",                      7,    ""),
        ("Most Innings Pitched",   "Patrick Fredrickson",             50.1,  ""),
        ("Most K's",               "Avery Jones",                     63,    ""),
        ("Lowest ERA",             "Avery Jones",                    "1.02", "21+ IP"),
        ("Most Saves",             "Logan Gerling, Patrick Fredrickson", 2, ""),
    ]
    write_table(
        ws4,
        title="2017 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 36, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
