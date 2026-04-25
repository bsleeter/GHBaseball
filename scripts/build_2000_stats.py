#!/usr/bin/env python3
"""Build 2000 Gig Harbor Varsity season stats xlsx from pages in
Historical/2000/. Jeff Vanderbilt senior year — .475 AVG, 2.9% K rate."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2000" / "2000_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 2000, [
        ("Individual Records", "Clean typed source. Verify before publishing.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No per-game schedule photo provided.", "NOT PROVIDED"),
        ("Team Batting", "2000 Overall Final Stats transcribed. Some small-sample cells had reading challenges.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Transcribed. Team: 286/3.94.", "READABLE"),
        ("Roster", "Clean transcription (hand-annotated with circles/strikes in source — likely tracking returning vs. graduating players; ignored here).", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Nick Cherbas",       12),
        (2,  "Kevin Freeman",      12),
        (3,  "Peter Jendro",       12),
        (4,  "Trevor Lyle",        12),
        (5,  "Ryan Snow",          12),
        (6,  "Nick Stone",         12),
        (7,  "Jeff Vanderbilt",    12),
        (8,  "Casey Willis",       12),
        (9,  "Tyler Bartlett",     11),
        (10, "Justin Carr",        11),
        (11, "Kyle Knorr",         11),
        (12, "Craig Manning",      11),
        (13, "Dan Okamoto",        11),
        (14, "Carl Olsen",         11),
        (15, "Brett Shearer",      10),
        (16, "Sam Rosendahl",      10),
    ]
    write_roster_sheet(
        ws_r,
        title="2000 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen"],
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    # Best-effort transcription; player roll aligned with roster.
    bat_rows = [
        ("Snow",        "105.0", 51, 29, 24, 9, 0, 0, 15, 10,  0,  5, 5, ".667", ".569", ".471"),
        ("Vanderbilt",  "117.0", 59, 24, 28, 2, 2, 2, 12, 17,  0,  2, 0, ".898", ".529", ".475"),
        ("Jendro",      "113.0", 54,  9, 17, 2, 0, 0,  7, 18,  0, 11, 1, ".352", ".385", ".315"),
        ("Freeman",     "113.0", 55, 14, 19, 3, 0, 2, 11, 23,  2, 11, 4, ".509", ".420", ".345"),
        ("Okamoto",      "45.0", 16,  5,  5, 0, 0, 0,  1,  2,  0,  4, 1, ".313", ".389", ".313"),
        ("Knorr",        "30.0", 14,  3,  2, 0, 0, 0,  1,  2,  0,  4, 0, ".143", ".200", ".143"),
        ("Lyle",         "82.0", 30,  9,  8, 3, 0, 0,  4,  5,  0,  7, 2, ".367", ".417", ".267"),
        ("Manning",      "88.0", 28,  3,  7, 0, 0, 0,  2,  4,  0,  4, 0, ".250", ".300", ".250"),
        ("Bartlett",     "66.0", 21,  7,  4, 2, 0, 0,  3,  5,  0,  6, 0, ".286", ".333", ".190"),
        ("Stone",        "51.0", 16,  4,  2, 0, 0, 0,  1,  0,  0,  6, 0, ".125", ".176", ".125"),
        ("Rosendahl",    "22.0",  5,  2,  1, 0, 0, 0,  2,  0,  0,  2, 0, ".200", ".429", ".200"),
        ("Shearer",      "78.0", 29,  2, 11, 1, 0, 0,  0,  8,  0,  7, 1, ".414", ".400", ".379"),
        ("Willis",       "32.0", 16,  1,  3, 0, 0, 0,  0,  0,  0,  4, 0, ".188", ".188", ".188"),
        ("Olsen",        "22.0",  4,  0,  0, 0, 0, 0,  4,  0,  0,  0, 2, ".000", ".600", ".000"),
        ("Cherbas",      "45.0",  7,  1,  1, 0, 0, 0,  0,  0,  0,  2, 0, ".143", ".143", ".143"),
        ("Carr",         "91.0", 42,  9,  7, 2, 1, 0,  4,  4,  1,  8, 2, ".333", ".340", ".167"),
        ("TEAM",        "117.0",472,129,143,35, 3, 6, 67,109,  3, 85,18, ".428", ".401", ".303"),
    ]
    write_table(
        ws2,
        title="2000 Gig Harbor Varsity — Team Batting (Overall Final Stats)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Snow",        "12.0",  72, 16, 20, 13, 10, 0,  0,  8, 1, 3, 1, ".286",  "8.00"),
        ("Okamoto",     "22.2", 126, 26, 24, 14, 20, 0,  1,  9, 2, 0, 0, ".260",  "4.33"),
        ("Manning",     "15.2",  77, 20,  7,  6, 10, 0,  1,  9, 0, 0, 0, ".308",  "3.25"),
        ("Rosendahl",    "4.1",  32, 14, 10,  6,  3, 0,  0,  4, 3, 0, 0, ".438",  "9.69"),
        ("Shearer",     "19.2", 105, 33, 18, 14, 14, 0,  2, 11, 0, 2, 0, ".321",  "4.98"),
        ("Olsen",       "24.2", 122, 23, 13, 10, 14, 0,  4, 12, 1, 0, 0, ".226",  "2.85"),
        ("Lyle",        "38.2", 189, 49, 39, 27, 22, 0,  4, 30, 3, 3, 0, ".288",  "4.69"),
        ("TEAM",       "117.0", 591,141,138, 75, 86, 0, 12,124, 7, 9, 1, ".286",  "3.94"),
    ]
    write_table(
        ws3,
        title="2000 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              15, "vs. W.F. West"),
        ("Most Runs (Inning)",             6, "vs. Olympia"),
        ("Most Runs Allowed (Game)",      14, "vs. S. Kitsap"),
        ("Widest Margin of Victory",      11, "vs. Foss"),
        ("One Run Games Lost",             1, "vs. Peninsula"),
        ("One Run Games Won",              1, ""),
        ("Most K's (by GH Pitchers)",     12, "vs. Lincoln"),
        ("Most K's (by Opponents)",        8, "vs. Wilson"),
        ("Most Hits (Game)",              13, "vs. W.F. West"),
        ("Most Hits Allowed (Game)",      16, "vs. Wilson"),
        ("Most Walks (Game for GH)",       8, "vs. Stadium"),
        ("Most Walks Allowed",            10, "vs. Bellarmine"),
        ("Most Singles (Game)",           11, "vs. W.F. West"),
        ("Most Doubles (Game)",            5, "vs. Wilson, Capital, Stadium"),
        ("Most Triples (Game)",            1, ""),
        ("Most Home Runs (Game)",          1, "vs. Lincoln, Foss, West, Mt. T, Wash"),
        ("Most Steals (Game)",             6, "vs. Lincoln"),
        ("Longest Winning Streak",         3, ""),
        ("Longest Losing Streak",          3, ""),
        ("Total Team Runs Scored",       129, ""),
        ("Total Team Runs Allowed",      138, ""),
        ("Total Team Hits",              143, ""),
        ("Total Team Extra-Base Hits",    44, ""),
        ("Total Team Home Runs",           4, ""),
        ("Total Team Stolen Bases",       43, ""),
        ("Total Team Walks",              67, ""),
        ("Total Team HBP",                18, ""),
    ]
    write_table(
        ws_h,
        title="2000 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Jeff Vanderbilt",              ".475", "30+ AB (28-59)"),
        ("Most Hits",              "Jeff Vanderbilt",               28,    ""),
        ("Most At Bats",           "Jeff Vanderbilt",               59,    ""),
        ("Most HBP",               "Trevor Lyle",                    4,    ""),
        ("Lowest K Ratio",         "Jeff Vanderbilt",             "2.9%",  "(2-69)"),
        ("Most Doubles",           "Ryan Snow",                      9,    ""),
        ("Most Triples",           "Jeff Vanderbilt",                2,    ""),
        ("Most Home Runs",         "Jeff Vanderbilt, Kevin Freeman", 2,    ""),
        ("Most Walks",             "Ryan Snow",                     15,    ""),
        ("Most Stolen Bases",      "Jeff Vanderbilt",               16,    ""),
        ("Most Total Bases",       "Jeff Vanderbilt",               53,    ""),
        ("Most RBIs",              "Kevin Freeman",                 23,    ""),
        ("Most Wins Pitching",     "Trevor Lyle, Sam Rosendahl",     3,    ""),
        ("Most Innings Pitched",   "Trevor Lyle",                   38,    ""),
        ("Most K's",               "Trevor Lyle",                   30,    ""),
        ("Lowest ERA",             "Carl Olsen",                  "2.85",  ""),
        ("Most Saves",             "Ryan Snow, Craig Manning",       1,    ""),
        ("Best On-Base Avg.",      "Ryan Snow",                   ".569",  "30+ AB"),
        ("Longest Hitting Streak", "Jeff Vanderbilt",               11,    ""),
        ("Most Runs Scored",       "Ryan Snow",                     29,    ""),
    ]
    write_table(
        ws4,
        title="2000 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 40, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
