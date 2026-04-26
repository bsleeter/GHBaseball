#!/usr/bin/env python3
"""Aggregate every Historical/{year}/{year}_Season_Stats.xlsx into a single
program-history JSON consumed by the Next.js site.

Outputs:
  gh-baseball-site/src/data/programHistory.json

Schema (high level):
  {
    meta: { yearsAvailable, bbcorEraStart, generatedAt, ... },
    seasons: { [year]: Season },
    players: { [playerId]: Player }
  }

Run:
  python3 scripts/build_master_data.py
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
HISTORICAL = ROOT / "Historical"
OUT_JSON = ROOT / "gh-baseball-site" / "src" / "data" / "programHistory.json"
ALIASES_FP = Path(__file__).resolve().parent / "data" / "name_aliases.json"
SEASONS_FP = Path(__file__).resolve().parent / "data" / "team_seasons.json"

BBCOR_ERA_START = 2012  # NFHS adopted BBCOR for HS baseball

# ─────────────────────────────────────────────────────────────────────────────
# Header normalization
# ─────────────────────────────────────────────────────────────────────────────

# Map source header → canonical key. Multiple source headers may map to the
# same canonical key; we prefer the most recent value seen.
BAT_HEADER_MAP: dict[str, str] = {
    "Player": "player",
    "G": "G", "GP": "G", "Games": "G", "Inn": "INN",
    "AB": "AB", "R": "R", "H": "H",
    "1B": "1B", "2B": "2B", "3B": "3B", "HR": "HR",
    "RBI": "RBI",
    "BB": "BB",
    "HBP": "HBP", "HP": "HBP",
    "K": "K", "SO": "K",
    "SAC": "SAC", "SAC-B": "SACB", "SACB": "SACB",
    "SAC-F": "SACF", "SACF": "SACF", "SF": "SACF",
    "TB": "TB",
    "SB": "SB",
    "AVG": "AVG", "BA": "AVG",
    "OBP": "OBP",
    "SLG": "SLG",
    "OPS": "OPS",
    "ERR": "E", "E": "E",
    "K%": "_skip",  # display-only
    "PO": "_skip", "A": "_skip", "FLD%": "_skip",
    "CS": "CS",
}

PIT_HEADER_MAP: dict[str, str] = {
    "Player": "player",
    "W": "W", "L": "L",
    "SV": "SV", "S": "SV",
    "G": "G", "GP": "G", "Games": "G",
    "GS": "_skip", "CG": "_skip", "SHO": "_skip", "SVO": "_skip",
    "CG/SHO": "_skip", "SV/SVO": "_skip",
    "IP": "IP",
    "BF": "BF", "#BF": "BF", "BF/AB": "BF", "AB": "BF",
    "R": "R", "RS": "R",
    "H": "H",
    "ER": "ER",
    "BB": "BB",
    "K": "K", "SO": "K",
    "HBP": "HBP", "HB": "HBP",
    "HR": "HR", "HRA": "HR",
    "ERA": "ERA",
    "OBA": "OPPBA", "BAA": "OPPBA", "OPPBA": "OPPBA",
    "WHIP": "WHIP",
    "WP": "_skip",
    "1B": "_skip", "2B": "_skip", "3B": "_skip",
    "SACB": "_skip", "SACF": "_skip",
}


def parse_num(v: Any) -> float | int | None:
    """Coerce cell values to a number where possible. Strings like '.327' become 0.327."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    if s in ("", "—", "-", "?"):
        return None
    # Remove trailing %
    s = s.rstrip("%")
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return None


def parse_ip(v: Any) -> float | None:
    """Convert baseball IP notation (5.1 = 5⅓ innings, 5.2 = 5⅔) to true decimal IP.
    Stores 5.1 as 5.333..., 5.2 as 5.667."""
    n = parse_num(v)
    if n is None:
        return None
    whole = int(n)
    frac = round((n - whole) * 10)
    if frac == 1:
        return whole + 1 / 3
    if frac == 2:
        return whole + 2 / 3
    return float(whole) if frac == 0 else float(n)


def find_header_row(rows: list[tuple], must_have: list[str]) -> tuple[int, list[str]] | None:
    for i, r in enumerate(rows):
        if not r:
            continue
        cells = [str(c).strip() if c is not None else "" for c in r]
        if all(any(m == c for c in cells) for m in must_have):
            return i, cells
    return None


def normalize_row(row: tuple, headers: list[str], header_map: dict[str, str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for h, v in zip(headers, row):
        key = header_map.get(h)
        if key is None or key == "_skip":
            continue
        if key == "player":
            out["player"] = str(v).strip() if v is not None else ""
        else:
            out[key] = v
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Per-section extractors
# ─────────────────────────────────────────────────────────────────────────────


def extract_roster(wb) -> tuple[list[dict], list[str], str | None]:
    """Returns (roster_rows, coaches, manager)."""
    if "Roster" not in wb.sheetnames:
        return [], [], None
    ws = wb["Roster"]
    rows = list(ws.iter_rows(values_only=True))
    roster: list[dict] = []
    coaches: list[str] = []
    manager: str | None = None
    in_coaches = False
    for r in rows:
        if not r or not any(c for c in r):
            in_coaches = False
            continue
        first = r[0]
        if isinstance(first, str):
            up = first.strip().upper()
            if up == "COACHES" or up.startswith("COACH"):
                in_coaches = True
                continue
            if up.startswith("MANAGER"):
                # "MANAGER: Mike Boyle" or just "MANAGER"
                m = re.match(r"MANAGER:?\s*(.+)?", first.strip(), re.IGNORECASE)
                if m and m.group(1):
                    manager = m.group(1).strip()
                in_coaches = False
                continue
            if "Roster" in first or first == "#":
                continue
        if isinstance(first, int):
            # roster row: (#, Player, Grade)
            num = first
            name = r[1] if len(r) > 1 else None
            grade = r[2] if len(r) > 2 else None
            if name:
                roster.append({"num": num, "player": str(name).strip(),
                                "grade": grade if isinstance(grade, int) else None})
        elif isinstance(first, str) and in_coaches:
            coaches.append(first.strip())
    return roster, coaches, manager


def _extract_table(wb, sheet_candidates: list[str], must_have: list[str],
                   header_map: dict[str, str], canonical_keys_required: list[str]):
    """Return (data_rows, team_total_row_or_None) from the first sheet matching."""
    sheet = None
    for n in wb.sheetnames:
        if any(c in n for c in sheet_candidates):
            sheet = wb[n]; break
    if sheet is None:
        return [], None
    rows = list(sheet.iter_rows(values_only=True))
    found = find_header_row(rows, must_have)
    if not found:
        return [], None
    hi, headers = found
    data_rows: list[dict] = []
    team_row: dict | None = None
    for r in rows[hi + 1:]:
        if not r or not any(c is not None and c != "" for c in r):
            continue
        norm = normalize_row(r, headers, header_map)
        player = norm.get("player", "").strip()
        if not player or player.lower().startswith("note") or player.startswith("—"):
            continue
        if player.upper() == "TEAM" or player.upper().startswith("TEAM "):
            team_row = norm
            continue
        # Require at least one of the required canonical keys to be present and numeric
        if not any(parse_num(norm.get(k)) is not None for k in canonical_keys_required):
            continue
        data_rows.append(norm)
    return data_rows, team_row


def extract_batting(wb) -> tuple[list[dict], dict | None]:
    rows, team = _extract_table(
        wb,
        sheet_candidates=["Batting", "Hitting"],
        must_have=["AB", "H"],
        header_map=BAT_HEADER_MAP,
        canonical_keys_required=["AB"],
    )
    # Post-process: coerce numbers, compute PA where possible
    def shape(r):
        out = {"player": r.get("player", "")}
        for k in ["G", "AB", "R", "H", "1B", "2B", "3B", "HR", "RBI", "BB",
                  "HBP", "K", "SAC", "SACB", "SACF", "TB", "SB", "E", "CS"]:
            n = parse_num(r.get(k))
            if n is not None:
                out[k] = n
        for k in ["AVG", "OBP", "SLG", "OPS"]:
            n = parse_num(r.get(k))
            if n is not None:
                out[k] = n
        # Combine SACB/SACF into SAC if SAC missing
        if "SAC" not in out:
            sb = out.pop("SACB", None)
            sf = out.pop("SACF", None)
            if sb is not None or sf is not None:
                out["SAC"] = (sb or 0) + (sf or 0)
        else:
            out.pop("SACB", None); out.pop("SACF", None)
        # Compute PA = AB + BB + HBP + SAC (high-school convention)
        ab = out.get("AB"); bb = out.get("BB", 0); hbp = out.get("HBP", 0); sac = out.get("SAC", 0)
        if isinstance(ab, (int, float)):
            out["PA"] = int(ab) + int(bb) + int(hbp) + int(sac)
        # Backfill 1B = H - 2B - 3B - HR when not directly tracked but breakdown is present
        if "1B" not in out:
            h = out.get("H"); d = out.get("2B"); t = out.get("3B"); hr = out.get("HR")
            if all(isinstance(v, (int, float)) for v in (h, d, t, hr)):
                singles = int(h) - int(d) - int(t) - int(hr)
                if singles >= 0:
                    out["1B"] = singles
        return out
    return [shape(r) for r in rows], (shape(team) if team else None)


def extract_pitching(wb) -> tuple[list[dict], dict | None]:
    rows, team = _extract_table(
        wb,
        sheet_candidates=["Pitching"],
        must_have=["IP"],
        header_map=PIT_HEADER_MAP,
        canonical_keys_required=["IP"],
    )
    def shape(r):
        out = {"player": r.get("player", "")}
        for k in ["W", "L", "SV", "G", "BF", "R", "H", "ER", "BB", "K", "HBP", "HR"]:
            n = parse_num(r.get(k))
            if n is not None:
                out[k] = n
        # IP — keep both display string and decimal
        ip_raw = r.get("IP")
        if ip_raw is not None and ip_raw != "":
            out["IPDisplay"] = str(ip_raw)
            ip_dec = parse_ip(ip_raw)
            if ip_dec is not None:
                out["IP"] = round(ip_dec, 4)
        for k in ["ERA", "OPPBA", "WHIP"]:
            n = parse_num(r.get(k))
            if n is not None:
                out[k] = n
        # ── Derived pitching stats (overwrite any source value for consistency)
        # WHIP = (BB + H) / IP; K7 = K × 7 / IP; KBB = K / BB
        ip_val = out.get("IP")
        if isinstance(ip_val, (int, float)) and ip_val > 0:
            bb = out.get("BB"); h = out.get("H"); k = out.get("K")
            if all(isinstance(v, (int, float)) for v in (bb, h)):
                out["WHIP"] = round((bb + h) / ip_val, 4)
            if isinstance(k, (int, float)):
                out["K7"] = round(k * 7 / ip_val, 4)
            if isinstance(k, (int, float)) and isinstance(bb, (int, float)) and bb > 0:
                out["KBB"] = round(k / bb, 4)
        return out
    return [shape(r) for r in rows], (shape(team) if team else None)


def extract_schedule(wb) -> list[dict]:
    if "Schedule" not in wb.sheetnames:
        return []
    ws = wb["Schedule"]
    rows = list(ws.iter_rows(values_only=True))
    # Find header
    hi = None
    for i, r in enumerate(rows):
        if not r:
            continue
        cells = [str(c).strip() if c else "" for c in r]
        if "Opponent" in cells:
            hi = i
            headers = cells
            break
    if hi is None:
        return []
    out = []
    for r in rows[hi + 1:]:
        if not r or not any(c is not None and c != "" for c in r):
            continue
        rec = {}
        for h, v in zip(headers, r):
            if v is None or v == "":
                continue
            key = h.lower().replace("/", "_").replace(" ", "_") or "col"
            rec[key] = v
        if rec.get("opponent"):
            out.append(rec)
    return out


def extract_individual_records(wb) -> list[dict]:
    sheet = None
    for n in wb.sheetnames:
        if "Records" in n or "Hall" in n:
            sheet = wb[n]; break
    if not sheet:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    # First row that looks like a header
    hi = None
    for i, r in enumerate(rows):
        if r and r[0] and isinstance(r[0], str) and r[0].strip() == "Record":
            hi = i; break
    if hi is None:
        return []
    out = []
    for r in rows[hi + 1:]:
        if not r or not r[0]:
            continue
        rec = {
            "stat": str(r[0]).strip(),
            "holder": str(r[1]).strip() if r[1] else "",
            "value": r[2] if r[2] is not None else None,
            "qualifier": str(r[3]).strip() if r[3] else "",
        }
        out.append(rec)
    return out


def extract_highlights(wb) -> list[dict]:
    sheet = None
    for n in wb.sheetnames:
        if "Highlight" in n:
            sheet = wb[n]; break
    if not sheet:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    hi = None
    for i, r in enumerate(rows):
        if r and r[0] == "Highlight":
            hi = i; break
    if hi is None:
        return []
    return [{"highlight": str(r[0]).strip(),
             "value": r[1] if len(r) > 1 else None,
             "context": str(r[2]).strip() if len(r) > 2 and r[2] else ""}
            for r in rows[hi + 1:] if r and r[0]]


def extract_readme(wb) -> dict[str, dict]:
    if "README" not in wb.sheetnames:
        return {}
    ws = wb["README"]
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    for r in rows:
        if not r or not r[0]:
            continue
        s = str(r[0]).strip()
        if s in ("SHEET", "Note") or "—" in s and len(r) < 3:
            continue
        if len(r) >= 3 and r[2]:
            status = str(r[2]).strip()
            if status in ("READABLE", "NEEDS MANUAL UPDATE", "NEEDS SPOT-CHECK", "NOT PROVIDED"):
                note = str(r[1]).strip() if r[1] else ""
                out[s] = {"status": status, "note": note}
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Player ID resolution
# ─────────────────────────────────────────────────────────────────────────────


class PlayerResolver:
    def __init__(self, aliases_data: dict, rosters_by_year: dict[int, list[dict]] | None = None):
        # canonical_id -> set(aliases)
        self.aliases: dict[str, set[str]] = {}
        for canonical, names in aliases_data.items():
            if canonical.startswith("_"):
                continue
            self.aliases[canonical] = set(n.strip().lower() for n in names) | {canonical.lower()}
        # Build reverse lookup
        self.alias_to_id: dict[str, str] = {}
        for canonical, alts in self.aliases.items():
            for alt in alts:
                self.alias_to_id[alt] = canonical
        # Cross-year roster index: year -> list of fully-named roster entries.
        # Used to resolve callups whose stat-line uses just a last name in
        # year X but appear on the roster in year X±1.
        self.rosters_by_year: dict[int, list[dict]] = rosters_by_year or {}
        # Track unrecognized names per year (diagnostic).
        self.unmatched: dict[tuple[int, str], int] = {}

    def slug(self, name: str) -> str:
        s = re.sub(r"[^A-Za-z0-9]+", "-", name.strip().lower()).strip("-")
        return s or "unknown"

    def _normalize_lastfirst(self, n: str) -> str:
        """Convert 'Henckel, M' or 'Frame, C' to a 'Last' fallback we can match on."""
        if "," in n:
            last, *rest = [s.strip() for s in n.split(",")]
            return last
        return n

    def _find_in_year(self, year: int, last: str) -> set[str]:
        """Return the set of fully-named roster entries from `year` whose last
        name matches `last`. Empty if no roster or no match."""
        roster = self.rosters_by_year.get(year, [])
        return {
            p["player"]
            for p in roster
            if " " in p["player"] and p["player"].split()[-1].lower() == last
        }

    def resolve(self, raw_name: str, year: int, roster_for_year: list[dict] | None = None) -> str:
        """Return a stable playerId for a raw name (best effort).

        Resolution order:
          1. Current-year roster lookup for ambiguous forms (e.g. 'Last',
             'Last, F'). Best for disambiguating families that overlap.
          2. Full alias map (handles 'TW. Friedman' → 'Tom Friedman').
          3. Last-only alias map.
          4. Cross-year roster lookup — expand window ±1, ±2, ±3 to find a
             unique fully-named candidate. Catches JV callups whose stat
             line appears in year X but who are rostered in year X±1.
          5. Fall back to slug of last name; record for diagnostic.
        """
        n = raw_name.strip()
        if not n:
            return "unknown"
        last_only = self._normalize_lastfirst(n)
        is_ambiguous = " " not in n or "," in n  # 'Last', 'Last, F', etc.

        # 1. Current-year roster
        if is_ambiguous and roster_for_year:
            last = last_only.split()[-1].lower()
            cands = [p["player"] for p in roster_for_year
                     if p["player"].split()[-1].lower() == last
                     and " " in p["player"]]
            if len(cands) == 1:
                return self.slug(cands[0])
        # 2. Alias map (full string)
        key = n.lower()
        if key in self.alias_to_id:
            return self.slug(self.alias_to_id[key])
        # 3. Alias map (last-only form)
        if last_only.lower() in self.alias_to_id:
            return self.slug(self.alias_to_id[last_only.lower()])
        # 4. Cross-year roster lookup for ambiguous names
        if is_ambiguous and self.rosters_by_year:
            last = last_only.split()[-1].lower()
            for window in (1, 2, 3):
                cands: set[str] = set()
                for y in range(year - window, year + window + 1):
                    if y == year:
                        continue  # already tried in step 1
                    cands |= self._find_in_year(y, last)
                if len(cands) == 1:
                    return self.slug(next(iter(cands)))
                if len(cands) > 1:
                    # Multiple candidates appeared — ambiguous, stop widening.
                    break
        # 5. Fall back to slug of last name; record diagnostic
        self.unmatched[(year, n)] = self.unmatched.get((year, n), 0) + 1
        return self.slug(last_only)


# ─────────────────────────────────────────────────────────────────────────────
# Main build
# ─────────────────────────────────────────────────────────────────────────────


def head_coach_for(year: int, eras: list[dict]) -> str | None:
    for e in eras:
        frm = e.get("from"); to = e.get("to")
        if frm is not None and year >= frm and (to is None or year <= to):
            return e.get("coach")
    return None


def main() -> None:
    aliases_data = json.loads(ALIASES_FP.read_text())
    seasons_meta = json.loads(SEASONS_FP.read_text())
    head_coach_eras = seasons_meta.get("_headCoachEras", [])

    year_dirs: list[tuple[int, Path]] = []
    for year_dir in sorted(HISTORICAL.iterdir()):
        if not year_dir.is_dir() or not year_dir.name.isdigit():
            continue
        year = int(year_dir.name)
        fp = year_dir / f"{year}_Season_Stats.xlsx"
        if not fp.exists():
            continue
        year_dirs.append((year, fp))

    # ── Pre-pass: collect rosters from every year so the resolver can
    # peek at adjacent years when a stat-line name is ambiguous.
    rosters_by_year: dict[int, list[dict]] = {}
    for year, fp in year_dirs:
        wb = load_workbook(fp, data_only=True)
        roster, _, _ = extract_roster(wb)
        if not roster:
            # Synthesize from stat lines when no Roster sheet (2023-2025+ GC files)
            batting, _ = extract_batting(wb)
            pitching, _ = extract_pitching(wb)
            seen: set[str] = set()
            for line in batting + pitching:
                nm = line.get("player", "").strip()
                if nm and nm not in seen:
                    seen.add(nm)
                    roster.append({"num": None, "player": nm, "grade": None})
        rosters_by_year[year] = roster
    resolver = PlayerResolver(aliases_data, rosters_by_year)

    seasons: dict[str, dict] = {}
    players: dict[str, dict] = {}
    years_available: list[int] = []

    for year, fp in year_dirs:
        wb = load_workbook(fp, data_only=True)

        roster, coaches, manager = extract_roster(wb)
        batting, team_batting = extract_batting(wb)
        pitching, team_pitching = extract_pitching(wb)
        schedule = extract_schedule(wb)
        records = extract_individual_records(wb)
        highlights = extract_highlights(wb)
        readme = extract_readme(wb)

        # Synthesize roster from stats lines when Roster sheet missing (e.g. 2023-2025)
        if not roster:
            seen2: set[str] = set()
            for line in batting + pitching:
                nm = line.get("player", "").strip()
                if nm and nm not in seen2:
                    seen2.add(nm)
                    roster.append({"num": None, "player": nm, "grade": None})

        # Resolve player IDs
        for line in batting:
            line["playerId"] = resolver.resolve(line["player"], year, roster)
        for line in pitching:
            line["playerId"] = resolver.resolve(line["player"], year, roster)
        for entry in roster:
            entry["playerId"] = resolver.resolve(entry["player"], year, roster)

        # Update player registry — display name picks the most "complete" form seen
        # (longer string with a space wins over short 'Last' or 'Last, F')
        def name_score(nm: str) -> int:
            return len(nm) + (10 if " " in nm and "," not in nm else 0)

        for entry in roster:
            pid = entry["playerId"]
            p = players.setdefault(pid, {
                "playerId": pid,
                "displayName": entry["player"],
                "years": [],
                "grades": {},
            })
            if name_score(entry["player"]) > name_score(p["displayName"]):
                p["displayName"] = entry["player"]
            if year not in p["years"]:
                p["years"].append(year)
            if entry.get("grade") is not None:
                p["grades"][str(year)] = entry["grade"]

        meta = seasons_meta.get(str(year), {})

        # Pull team record from pitching team-row if available; meta override wins.
        record = None
        if team_pitching:
            w = team_pitching.get("W"); l = team_pitching.get("L")
            if isinstance(w, (int, float)) and isinstance(l, (int, float)):
                record = {"W": int(w), "L": int(l)}
        meta_record = meta.get("record")
        if isinstance(meta_record, dict) and "W" in meta_record and "L" in meta_record:
            record = {"W": int(meta_record["W"]), "L": int(meta_record["L"])}

        # Append meta-supplied highlights (for years without a Team Highlights sheet)
        meta_highlights = meta.get("highlights")
        if isinstance(meta_highlights, list):
            for h in meta_highlights:
                if isinstance(h, dict) and "highlight" in h:
                    highlights.append({
                        "highlight": str(h["highlight"]).strip(),
                        "value": h.get("value"),
                        "context": str(h.get("context", "")).strip(),
                    })

        # Apply meta-supplied grade overrides (useful for GameChanger years
        # whose synthesized roster has no grade field).
        meta_grades = meta.get("rosterGrades")
        if isinstance(meta_grades, dict):
            grade_lookup = {k.strip().lower(): v for k, v in meta_grades.items()}
            for entry in roster:
                if entry.get("grade") is None:
                    g = grade_lookup.get(entry["player"].strip().lower())
                    if isinstance(g, int):
                        entry["grade"] = g

        # Append meta-supplied schedule entries (for years built from
        # GameChanger CSVs that lack a schedule sheet).
        meta_schedule = meta.get("schedule")
        if isinstance(meta_schedule, list):
            for g in meta_schedule:
                if isinstance(g, dict) and g.get("opponent"):
                    schedule.append({
                        k: g[k]
                        for k in ("date", "loc", "opponent", "w_l", "score",
                                  "result_gh_opp", "r", "h", "e", "notes")
                        if k in g
                    })

        seasons[str(year)] = {
            "year": year,
            "era": "bbcor" if year >= BBCOR_ERA_START else "pre-bbcor",
            "league": meta.get("league"),
            "leagueChamp": bool(meta.get("leagueChamp")),
            "districtChamp": bool(meta.get("districtChamp")),
            "stateChamp": bool(meta.get("stateChamp")),
            "statePlace": meta.get("statePlace"),
            "record": record,
            "headCoach": head_coach_for(year, head_coach_eras),
            "coaches": coaches,
            "manager": manager,
            "roster": roster,
            "schedule": schedule,
            "battingLines": batting,
            "pitchingLines": pitching,
            "teamBatting": team_batting,
            "teamPitching": team_pitching,
            "individualRecords": records,
            "highlights": highlights,
            "sourceQuality": readme,
        }
        years_available.append(year)
        print(f"  {year}: roster={len(roster)} bat={len(batting)} pit={len(pitching)} "
              f"sched={len(schedule)} records={len(records)}")

    # Finalize player display names (prefer most-recent, longest non-shortened form)
    for pid, p in players.items():
        # If display name is just last name, leave as-is for now
        p["years"].sort()

    output = {
        "meta": {
            "generatedAt": datetime.utcnow().isoformat() + "Z",
            "bbcorEraStart": BBCOR_ERA_START,
            "yearsAvailable": sorted(years_available),
            "playerCount": len(players),
            "schemaVersion": 1,
        },
        "seasons": seasons,
        "players": players,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(output, indent=2, default=str) + "\n")

    # Show unmatched names so we can grow the alias file
    if resolver.unmatched:
        print("\nUnmatched names (consider adding to name_aliases.json):")
        by_count = sorted(resolver.unmatched.items(), key=lambda x: -x[1])
        for (year, n), c in by_count[:25]:
            print(f"  {year}: {n!r} ({c}x)")
        if len(by_count) > 25:
            print(f"  ... +{len(by_count) - 25} more")

    sz = OUT_JSON.stat().st_size / 1024
    print(f"\nWrote {OUT_JSON}  ({sz:.1f} KB)")
    print(f"  {len(years_available)} seasons, {len(players)} unique players (best-effort)")


if __name__ == "__main__":
    main()
