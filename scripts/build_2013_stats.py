#!/usr/bin/env python3
"""Build 2013 Gig Harbor Varsity season stats xlsx from CoachStat Baseball
Report pages in Historical/2013/. Format differs from other years — the
CoachStat report has per-game schedule with pitchers and its own column
set for team batting/pitching stats."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2013" / "2013_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_section_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE, italic=True)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None, section_rows=None,
                col_widths=None, left_align_cols=None, note=None):
    section_rows = section_rows or {}
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        if i in section_rows:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            sc = ws.cell(row=r, column=1, value=section_rows[i])
            style_section_header(sc)
            r += 1
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule (CoachStat format) ───────────────────────
    ws = wb.active
    ws.title = "Schedule"
    sched_headers = ["Date", "Loc", "Opponent", "W/L", "Score", "R", "H", "E", "Winning Pitcher"]
    sched_rows = [
        # Conference 7-7
        ("Mar 25, 2013", "Home", "Stadium 1",          "L", "7-4",  4,  9, 1, ""),
        ("Mar 27, 2013", "Away", "South Kitsap 1",     "W", "2-0",  2, 11, 1, "Nick Gagliardi"),
        ("Mar 30, 2013", "Away", "Central Kitsap 1",   "L", "4-3",  3, 13, 6, ""),
        ("Apr 2, 2013",  "Home", "Yelm 1",             "W", "3-2",  3,  6, 2, "Nick Gagliardi"),
        ("Apr 9, 2013",  "Away", "Bellarmine 1",       "L", "2-0",  0,  2, 2, ""),
        ("Apr 13, 2013", "Away", "Olympia 1",          "L", "6-2",  2,  7, 0, ""),
        ("Apr 15, 2013", "Home", "South Kitsap 2",     "W", "4-3",  4, 10, 0, "Kody Davis"),
        ("Apr 16, 2013", "Home", "Central Kitsap 2",   "W", "4-3",  4, 10, 0, "Kody Davis"),
        ("Apr 22, 2013", "Away", "Yelm 2",             "W", "13-1",13, 16, 3, "Nick Gagliardi"),
        ("Apr 24, 2013", "Home", "Bellarmine 2",       "L", "4-0",  0,  2, 1, ""),
        ("Apr 26, 2013", "Home", "Olympia 2",          "L", "4-1",  1,  5, 1, ""),
        ("Apr 29, 2013", "Away", "Stadium 2",          "W", "2-1",  2,  4, 2, "Nick Gagliardi"),
        ("May 3, 2013",  "Away", "Bellarmine 3",       "W", "5-4",  5, 10, 1, "Conor Scanlan"),
        ("May 3, 2013",  "Home", "CK 3",               "W", "5-2",  5,  8, 1, "Matt Henckel"),
        # Non-Conference 4-2
        ("Mar 12, 2013", "Away", "Curtis",             "W", "8-2",  8,  9, 1, "Kody Davis"),
        ("Mar 21, 2013", "Away", "Timberline",         "W", "10-7",10, 14, 3, "Aidan O'Neill"),
        ("Mar 22, 2013", "Away", "N. Thurston",        "W", "12-1",12,  8, 0, "Dean Hassan"),
        ("Apr 1, 2013",  "Away", "Wilson",             "W", "19-6",19, 15, 2, "Hunter Johnson"),
        ("Apr 21, 2013", "Away", "Summit 2",           "L", "10-8", 8,  7, 3, ""),
        ("Apr 21, 2013", "Away", "Summit 1",           "L", "7-6",  6,  7, 3, ""),
        # Sectional 1-1
        ("May 9, 2013",  "Home", "Graham Kapowsin",    "W", "4-2",  4,  9, 0, "Nick Gagliardi"),
        ("May 9, 2013",  "Away", "Kentridge",          "L", "3-2",  2,  6, 0, ""),
        # Totals
        ("Season Record", "", "22 games", "W 12-10", "RS 113 · RA 80", "", "", "", ""),
    ]
    write_table(
        ws,
        title="2013 Gig Harbor Varsity — Schedule",
        headers=sched_headers,
        rows=sched_rows,
        team_row_index=len(sched_rows) - 1,
        section_rows={14: "Non-Conference (4-2)", 20: "Sectional (1-1)"},
        col_widths=[14, 8, 20, 6, 12, 6, 6, 6, 20],
        left_align_cols={1, 3, 9},
    )

    # ─── Sheet 2: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Kody Davis",          12),
        (2,  "Drew Frame",          12),
        (3,  "Nick Gagliardi",      12),
        (4,  "Garrett Gallinger",   12),
        (5,  "Owen Guenther",       12),
        (6,  "Hunter Johnson",      12),
        (7,  "Dan Koberstein",      12),
        (8,  "Colten Miller",       12),
        (9,  "Drew Nordi",          12),
        (10, "Aidan O'Neill",       12),
        (11, "Haak Wagner",         12),
        (12, "Drew Barnett",        11),
        (13, "Sterling Brown",      11),
        (14, "Quintin Carlson",     11),
        (15, "Chad Glover",         11),
        (16, "Dean Hassan",         11),
        (17, "Conor Scanlan",       11),
        (18, "Mason Selby",         11),
        (19, "Mark Sluys",          11),
        (20, "Matt Henckel",        10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2013 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
    )

    # ─── Sheet 3: Team Batting (CoachStat format) ───────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "H", "1B", "2B", "3B", "HR", "RBI", "TB",
        "SO", "SB", "CS", "OBP", "SLG", "AVG",
    ]
    # CoachStat report values transcribed as shown. The source sheet has
    # some internal inconsistencies (breakdown sums vs. TB/AVG); values
    # preserved as printed. Trust the Individual Records sheet for
    # authoritative top numbers.
    bat_rows = [
        ("Barnett, D",    30,  6,  4, 1, 1, 0,  7,  9,  6, 0, 0, ".273", ".233", ".200"),
        ("Brown, S",      75, 28, 13, 8, 0, 0,  8, 38, 15, 8, 0, ".447", ".507", ".373"),
        ("Carlson, Q",    25,  5,  2, 1, 0, 0, 11, 11,  5, 3, 0, ".333", ".280", ".200"),
        ("Davis, K",       1,  0,  0, 0, 0, 0,  0,  0,  0, 0, 0, ".374", ".400", ".500"),
        ("Frame, C",      82, 17, 15, 2, 0, 0, 16, 19,  7, 2, 0, ".333", ".232", ".207"),
        ("Gagliardi, N",  34,  5,  6, 0, 0, 0,  4,  6,  9, 0, 0, ".500", ".176", ".176"),
        ("Gallinger, G",  12,  2,  3, 0, 0, 2,  4,  2,  7, 0, 0, ".500", ".500", ".250"),
        ("Glover, C",     12,  1,  3, 0, 0, 0,  0,  3,  7, 0, 0, ".500", ".250", ".250"),
        ("Guenther, O",   12,  1,  2, 0, 0, 0,  0,  2,  6, 0, 0, ".231", ".167", ".167"),
        ("Hassan, D",      0,  0,  0, 0, 0, 0,  0,  0,  0, 0, 0, ".000", ".000", ".000"),
        ("Henckel, M",    45, 10,  6, 3, 1, 0,  8, 16, 15, 4, 0, ".333", ".356", ".222"),
        ("Johnson, H",    67, 20, 13, 7, 0, 1, 11, 30, 14, 6, 0, ".382", ".448", ".299"),
        ("Koberstein, D",  7,  3,  1, 2, 0, 0,  1,  5,  3, 1, 0, ".500", ".714", ".429"),
        ("Miller, C",     56, 14,  6, 2, 0, 0,  7, 16,  3, 5, 0, ".500", ".286", ".250"),
        ("Nordi, D",      40,  6,  7, 2, 0, 0, 16, 11,  9, 3, 0, ".333", ".275", ".175"),
        ("O'Neill, A",    41,  8,  3, 2, 1, 0,  7, 15,  5, 5, 0, ".280", ".366", ".195"),
        ("Scanlan, C",    42,  6,  5, 1, 0, 0,  4,  7,  5, 2, 0, ".306", ".167", ".143"),
        ("Selby, M",      30,  3,  3, 0, 0, 0,  3,  3,  5, 0, 0, ".167", ".100", ".100"),
        ("Sluys, M",      40,  8,  4, 2, 0, 0,  2, 10,  2, 2, 0, ".302", ".250", ".200"),
        ("Sutton, G",     10,  3,  2, 1, 0, 0,  3,  4,  3, 0, 0, ".300", ".300", ".300"),
        ("Wagner, H",     20,  4,  2, 1, 0, 0,  5,  7,  2, 3, 0, ".300", ".350", ".200"),
        ("TEAM",         681,183,113, 37, 3, 5,103,241, 92,105,29, ".372", ".354", ".269"),
    ]
    write_table(
        ws2,
        title="2013 Gig Harbor Varsity — Team Batting (CoachStat)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [7] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: CoachStat Baseball Report format. Per-player breakdown sums "
            "do not always match TB/AVG in the source sheet — values preserved "
            "as printed. Trust the Individual Records sheet for authoritative "
            "top-line numbers."
        ),
    )

    # ─── Sheet 4: Team Pitching (CoachStat format) ──────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "W", "L", "ERA", "G", "GS", "CG/SHO", "SV/SVO",
        "IP", "H", "R", "ER", "HR", "BB", "SO",
    ]
    pit_rows = [
        ("Davis, K",       4, 2, "3.82",  8, 0, 0, 0, "20.0",  22, 14, 12, 0, 18, 20),
        ("Gagliardi, N",   5, 2, "1.86",  8, 2, 0, 0, "24.0",  22, 13,  6, 0, 18, 23),
        ("Hassan, D",      2, 1, "2.33",  4, 0, 0, 0, "15.0",  12,  6,  5, 1,  6,  8),
        ("Henckel, M",     0, 1, "2.17",  4, 0, 0, 0, "6.1",    5,  5,  2, 0,  3,  3),
        ("Johnson, H",     0, 1, "2.40",  6, 1, 0, 0, "10.0",   8,  4,  3, 0,  4,  6),
        ("Nordi, D",       0, 0, "36.75", 3, 0, 0, 0, "2.0",    4,  8,  7, 0,  6,  2),
        ("O'Neill, A",     1, 1, "10.50", 2, 0, 0, 0, "6.0",    9,  7,  7, 1,  5,  3),
        ("Scanlan, C",     0, 0, "3.04",  2, 0, 0, 0, "1.1",    5,  5,  4, 0,  4,  1),
        ("Selby, M",       0, 0, "0.00",  2, 0, 0, 0, "1.0",    0,  0,  0, 0,  0,  0),
        ("Sutton, G",      0, 0, "0.00",  1, 0, 0, 0, "2.1",    1,  0,  0, 0,  0,  2),
        ("Wagner, H",      0, 0, "0.00",  1, 0, 0, 0, "1.0",    0,  0,  0, 0,  0,  1),
        ("TEAM",          12,10, "3.25", 22, 3, 0, 0, "153.0", 124, 80, 71, 0, 14, 85),
    ]
    write_table(
        ws3,
        title="2013 Gig Harbor Varsity — Team Pitching (CoachStat)",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Sterling Brown",                 ".373", "30+ AB (28-75)"),
        ("Most Plate Appearances", "Drew Frame",                      89,    "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",         "Colten Miller",                  "5.3%", "(3-56)"),
        ("Most Doubles",           "Sterling Brown",                   8,    ""),
        ("Most Triples",           "Garrett Gallinger, Sterling Brown", 1,   ""),
        ("Most Home Runs",         "Garrett Gallinger",                2,    ""),
        ("Most Walks",             "Sterling Brown",                   8,    ""),
        ("Most Stolen Bases",      "Sterling Brown",                   8,    ""),
        ("Most Total Bases",       "Sterling Brown",                  38,    ""),
        ("Most RBIs",              "Drew Nordi, Drew Frame",          16,    ""),
        ("Best On-Base Avg.",      "Sterling Brown",                 ".447", "30+ AB"),
        ("Longest Hitting Streak", "Sterling Brown",                   9,    ""),
        ("Most Runs Scored",       "Drew Frame",                      18,    ""),
        ("Most Wins Pitching",     "Nick Gagliardi",                   5,    ""),
        ("Most Innings Pitched",   "Nick Gagliardi",                  41,    ""),
        ("Most K's",               "Nick Gagliardi",                  57,    ""),
        ("Lowest ERA",             "Nick Gagliardi",                 "1.88", "14+ IP"),
        ("Most Saves",             "Haak Wagner",                      3,    ""),
    ]
    write_table(
        ws4,
        title="2013 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 38, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
