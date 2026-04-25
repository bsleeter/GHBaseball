#!/usr/bin/env python3
"""Build 2018 Gig Harbor Varsity season stats xlsx from scanned records
in Historical/2018/. Some cells are best-effort transcriptions from low-
resolution images and may need manual verification."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2018" / "2018_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"  # light amber for cells to verify

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    """Writes a roster sheet with player number, name, grade, and coaches below."""
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20

    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10

    # Coaches note
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None, col_widths=None,
                left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 32
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Schedule"
    schedule_rows = [
        ("Bellevue",    "9-6"),
        ("O'Dea",       "0-10"),
        ("Bainbridge",  "1-0"),
        ("Capital",     "0-2 (9 inn)"),
        ("Capital",     "14-0"),
        ("Shelton",     "4-8"),
        ("Shelton",     "5-6 (10 inn)"),
        ("Timberline",  "0-5"),
        ("Timberline",  "2-1"),
        ("Peninsula",   "5-2"),
        ("Peninsula",   "7-2"),
        ("C. Kitsap",   "3-2"),
        ("Yelm",        "14-2"),
        ("Yelm",        "17-8"),
        ("C. Kitsap",   "12-9"),
        ("N. Thurston", "7-6 (8 inn)"),
        ("S. Kitsap",   "1-4"),
        ("N. Thurston", "1-0"),
        ("Curtis",      "2-5 (8 inn)"),
        ("Bethel",      "13-0"),
        ("Kelso",       "9-5"),
        ("Shelton",     "7-1"),
        ("Capital",     "6-7 (10 inn)"),
        ("Kelso",       "4-7"),
        ("Season Record", "15-9"),
    ]
    write_table(
        ws,
        title="2018 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        col_widths=[22, 18],
        left_align_cols={1},
    )

    # ─── Sheet 2: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Jordan Haworth",      12),
        (2,  "Austin Dempewolf",    12),
        (3,  "Kirin Peterson",      12),
        (4,  "Bodi Tisch",          12),
        (5,  "Shannon Milbourn",    12),
        (6,  "Max Kein",            12),
        (7,  "Will Fernan",         12),
        (8,  "Ike Semmler",         12),
        (9,  "Cage Hardy",          11),
        (10, "Cole Smith",          11),
        (11, "Jake Bruess",         11),
        (12, "Chad Sorenson",       11),
        (13, "Eric Underwood",      11),
        (14, "Grant Hassan",        11),
        (15, "Max Sparrow",         11),
        (16, "Sebastian Toglia",    11),
        (17, "Owen Wild",           10),
        (18, "Luke Deschenes",      10),
        (19, "Luke Finnegan",       10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2018 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Larry Roehr", "Oscar Ortiz"],
    )

    # ─── Sheet 3: Hitting & Fielding ────────────────────────────────
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Dempewolf", 24, 3, 3, 1, 2, 1, 0, 0, 4, ".125", ".167", ".364", 12, 9, 0, 0, 35, 7, 1, ".977", 0),
        ("Dessert",   15, 1, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".063", 7, 1, 0, 0, 0, 0, 0, ".000", 0),
        ("Feenstra",   2, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 2, 0, 0, 0, 0, 0, 0, ".000", 0),
        ("Hardy",     51, 1, 14, 11, 10, 2, 1, 1, 21, ".275", ".412", ".422", 4, 7, 4, 6, 3, 7, 1, ".909", 0),
        ("Hassan",     0, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 0, 0, 0, 0, 1, 1, 0, "1.000", 0),
        ("Haworth",   86, 23, 37, 28, 18, 10, 7, 2, 67, ".430", ".779", ".479", 18, 6, 2, 1, 23, 48, 5, ".930", 4),
        ("Kein",      31, 1, 4, 1, 4, 0, 0, 0, 4, ".129", ".129", ".206", 5, 2, 1, 0, 1, 2, 1, ".750", 1),
        ("Milbourn",  30, 8, 13, 3, 11, 2, 0, 0, 15, ".433", ".500", ".500", 5, 4, 0, 0, 1, 0, 0, "1.000", 0),
        ("Peterson",  76, 25, 26, 25, 17, 5, 1, 3, 42, ".342", ".553", ".451", 14, 18, 0, 3, 37, 50, 6, ".935", 1),
        ("Semmler",   58, 11, 14, 9, 13, 1, 0, 0, 15, ".241", ".259", ".323", 5, 6, 1, 0, 7, 13, 2, ".882", 3),
        ("Smith",     68, 13, 18, 4, 16, 2, 0, 0, 20, ".265", ".294", ".383", 15, 10, 3, 2, 25, 4, 5, ".853", 1),
        ("Sorenson",  45, 9, 10, 10, 10, 0, 0, 0, 10, ".222", ".222", ".375", 16, 10, 1, 2, 1, 1, 2, ".500", 0),
        ("Sparrow",    9, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 3, 0, 0, 1, 2, 4, 1, ".800", 0),
        ("Tisch",     60, 14, 22, 11, 20, 2, 0, 0, 24, ".367", ".400", ".500", 8, 5, 0, 3, 144, 3, 2, ".986", 3),
        ("Toglia",    60, 16, 20, 7, 17, 2, 1, 0, 23, ".333", ".383", ".333", 11, 3, 1, 2, 18, 1, 0, ".955", 0),
        ("Underwood",  0, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 0, 0, 0, 0, 0, 0, 0, "1.000", 0),
        ("Wild",      51, 14, 17, 10, 13, 3, 1, 0, 23, ".333", ".451", ".385", 13, 12, 5, 1, 21, 17, 2, ".950", 1),
        ("TEAM",      657, 157, 191, 120, 142, 31, 10, 8, 266, ".291", ".405", ".392", 135, 93, 16, 25, 358, 207, 36, ".940", 15),
    ]
    write_table(
        ws2,
        title="2018 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=[13] + [7] * (len(hit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 3: Pitching ──────────────────────────────────────────
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Bruess",    0, 0, 0, "0.2",  5,  4,  1,  0,  3, 3,  1,  0, 0, 0, 1, "31.500"),
        ("Dempewolf", 7, 0, 0, "55.0", 234, 24, 29, 72, 26, 18, 23, 5, 1, 3, 3, "2.291"),
        ("Dessert",   0, 1, 0, "5.0",  27, 5,  6,  5,  5, 5,  3,  3, 0, 0, 3, "7.000"),
        ("Fernan",    0, 0, 0, "8.1",  44, 7,  10, 3,  7, 2,  7,  3, 0, 0, 0, "1.680"),
        ("Finnigan",  0, 1, 0, "4.1",  21, 3,  5,  2,  2, 2,  3,  2, 0, 0, 0, "3.233"),
        ("Hassan",    0, 0, 0, "7.1",  42, 13, 9,  3,  9, 5,  5,  3, 0, 0, 5, "8.594"),
        ("Haworth",   1, 1, 0, "8.0",  34, 3,  5,  9,  5, 3,  3,  2, 0, 0, 0, "2.625"),
        ("Peterson",  0, 1, 0, "2.2",  12, 2,  4,  2,  1, 2,  3,  1, 0, 0, 0, "6.364"),
        ("Sorenson",  0, 0, 1, "1.0",  2,  0,  0,  1,  1, 0,  1,  0, 0, 0, 0, "0.000"),
        ("Sparrow",   3, 3, 0, "21.0", 97, 13, 23, 10, 11, 9, 18, 3, 0, 2, 2, "3.857"),
        ("Underwood", 0, 2, 2, "4.2",  28, 5,  9,  3,  4, 2,  6,  1, 2, 0, 2, "2.997"),
        ("Wild",      4, 0, 1, "50.2", 219, 28, 49, 51, 12, 22, 35, 11, 2, 1, 4, "3.068"),
        ("TEAM",     15, 9, 4, "169.2", 772, 107, 151, 180, 86, 77, 111, 31, 5, 4, 22, "3.186"),
    ]
    write_table(
        ws3,
        title="2018 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[13] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Shannon McLoughlin",    ".433", "30+ AB (13-30)"),
        ("Most Plate Appearances", "Jordan Haworth",         97,    "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",         "Cage Hardy",            "6.0%", "(4-67)"),
        ("Most Doubles",           "Jordan Haworth",         10,    ""),
        ("Most Triples",           "Kirby Peterson",          3,    ""),
        ("Most Home Runs",         "Kirby Peterson",          7,    ""),
        ("Most Walks",             "Jordan Haworth",         16,    ""),
        ("Most Stolen Bases",      "Shannon McLoughlin",      6,    ""),
        ("Most Total Bases",       "Jordan Haworth",         67,    ""),
        ("Most RBIs",              "Jordan Haworth",         28,    ""),
        ("Best On-Base Avg.",      "Shannon McLoughlin",    ".500", "30+ AB"),
        ("Longest Hitting Streak", "Jordan Haworth",         10,    ""),
        ("Most Runs Scored",       "Kirby Peterson",         25,    ""),
        ("Most Wins Pitching",     "Austin Dempewolf",        7,    ""),
        ("Most Innings Pitched",   "Austin Dempewolf",       55,    ""),
        ("Most K's",               "Austin Dempewolf",       72,    ""),
        ("Lowest ERA",             "Owen Wild",             "2.29", "21+ IP"),
        ("Most Saves",             "Austin Underwood",        2,    ""),
    ]
    write_table(
        ws4,
        title="2018 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 30, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
