#!/usr/bin/env python3
"""Build per-season xlsx files for the GameChanger-era seasons (2023-2025)
matching the structure of the 2019 Historical stats file: Hitting & Fielding,
Pitching, and Individual Records. Game-by-game schedule isn't in the source
CSVs, so that sheet is omitted."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
STATS_DIR = ROOT / "Stats"
OUT_DIR = ROOT / "Historical"

# Each value is a path relative to ROOT — most live under Stats/ but recent
# in-progress seasons can also be placed directly in Historical/{year}/.
SEASONS = {
    2023: "Stats/Gig Harbor Varsity Tides Spring 2023 Stats.csv",
    2024: "Stats/Gig Harbor Tides Varsity Spring 2024 Stats.csv",
    2025: "Stats/Gig Harbor Varsity Tides Spring 2025 Stats.csv",
    2026: "Historical/2026/Gig Harbor Varsity Tides Spring 2026 Stats.csv",
}

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── Parsing ──────────────────────────────────────────────────────────────

def load_season(path: Path):
    with path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    section_row = rows[0]
    header_row = rows[1]
    current = ""
    sections = []
    for s in section_row:
        if s.strip():
            current = s.strip()
        sections.append(current)
    keys = []
    for sec, name in zip(sections, header_row):
        name = name.strip()
        if sec and name not in ("Number", "Last", "First"):
            keys.append(f"{sec}:{name}")
        else:
            keys.append(name)
    players = []
    totals = None
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        if row[0] == "Totals":
            totals = dict(zip(keys, row))
            continue
        if row[0] == "Glossary":
            break
        players.append(dict(zip(keys, row)))
    return players, totals


def to_num(v, default=0.0):
    if v is None:
        return default
    v = str(v).strip()
    if v in ("", "-", "N/A", "Inf"):
        return default
    try:
        return float(v)
    except ValueError:
        return default


def ip_frac_to_thirds(ip_raw):
    whole, _, frac = str(ip_raw).partition(".")
    try:
        t = int(whole or 0) * 3
        if frac:
            t += int(frac[0])
        return t
    except ValueError:
        return 0


def ip_thirds_to_str(t):
    return f"{int(t // 3)}.{int(t % 3)}"


def full_name(rec):
    return f"{rec.get('First','').strip()} {rec.get('Last','').strip()}".strip()


def fmt_avg(x):
    if x is None or x <= 0:
        return ".000"
    s = f"{x:.3f}"
    if s.startswith("0"):
        s = s[1:]
    elif s.startswith("-0"):
        s = "-" + s[2:]
    return s


# ─── Styling ──────────────────────────────────────────────────────────────

def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def write_table(ws, title, headers, rows, team_row_index=None, col_widths=None,
                left_align_cols=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20

    r = 4
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(
                    cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols)
                )
        r += 1

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ─── Sheet builders ───────────────────────────────────────────────────────

HIT_HEADERS = [
    "Player", "GP", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
    "AVG", "SLG", "OBP", "OPS", "SO", "BB", "HBP", "SAC", "SF", "SB",
    "PO", "A", "E", "FLD%",
]


def build_hit_row(p):
    ab = int(to_num(p.get("Batting:AB")))
    h = int(to_num(p.get("Batting:H")))
    tb = int(to_num(p.get("Batting:TB")))
    return [
        full_name(p),
        int(to_num(p.get("Batting:GP"))),
        ab,
        int(to_num(p.get("Batting:R"))),
        h,
        int(to_num(p.get("Batting:RBI"))),
        int(to_num(p.get("Batting:1B"))),
        int(to_num(p.get("Batting:2B"))),
        int(to_num(p.get("Batting:3B"))),
        int(to_num(p.get("Batting:HR"))),
        tb,
        fmt_avg(h / ab) if ab else ".000",
        fmt_avg(tb / ab) if ab else ".000",
        p.get("Batting:OBP", ".000").strip() or ".000",
        p.get("Batting:OPS", ".000").strip() or ".000",
        int(to_num(p.get("Batting:SO"))),
        int(to_num(p.get("Batting:BB"))),
        int(to_num(p.get("Batting:HBP"))),
        int(to_num(p.get("Batting:SAC"))),
        int(to_num(p.get("Batting:SF"))),
        int(to_num(p.get("Batting:SB"))),
        int(to_num(p.get("Fielding:PO"))),
        int(to_num(p.get("Fielding:A"))),
        int(to_num(p.get("Fielding:E"))),
        (p.get("Fielding:FPCT", ".000").strip() or ".000"),
    ]


def build_hit_sheet(ws, year, players, totals):
    rows = [build_hit_row(p) for p in players if to_num(p.get("Batting:PA")) > 0 or to_num(p.get("Batting:AB")) > 0 or to_num(p.get("Fielding:TC")) > 0]
    # Sort by AB descending, then name
    rows.sort(key=lambda r: (-r[2], r[0]))
    # Team totals
    if totals:
        ab = int(to_num(totals.get("Batting:AB")))
        h = int(to_num(totals.get("Batting:H")))
        tb = int(to_num(totals.get("Batting:TB")))
        team_row = [
            "TEAM",
            int(to_num(totals.get("Batting:GP"))),
            ab,
            int(to_num(totals.get("Batting:R"))),
            h,
            int(to_num(totals.get("Batting:RBI"))),
            int(to_num(totals.get("Batting:1B"))),
            int(to_num(totals.get("Batting:2B"))),
            int(to_num(totals.get("Batting:3B"))),
            int(to_num(totals.get("Batting:HR"))),
            tb,
            fmt_avg(h / ab) if ab else ".000",
            fmt_avg(tb / ab) if ab else ".000",
            totals.get("Batting:OBP", ".000").strip() or ".000",
            totals.get("Batting:OPS", ".000").strip() or ".000",
            int(to_num(totals.get("Batting:SO"))),
            int(to_num(totals.get("Batting:BB"))),
            int(to_num(totals.get("Batting:HBP"))),
            int(to_num(totals.get("Batting:SAC"))),
            int(to_num(totals.get("Batting:SF"))),
            int(to_num(totals.get("Batting:SB"))),
            int(to_num(totals.get("Fielding:PO"))),
            int(to_num(totals.get("Fielding:A"))),
            int(to_num(totals.get("Fielding:E"))),
            (totals.get("Fielding:FPCT", ".000").strip() or ".000"),
        ]
        rows.append(team_row)
    write_table(
        ws,
        title=f"{year} Gig Harbor Varsity — Hitting & Fielding",
        headers=HIT_HEADERS,
        rows=rows,
        team_row_index=len(rows) - 1 if totals else None,
        col_widths=[14] + [7] * (len(HIT_HEADERS) - 1),
        left_align_cols={1},
    )


PIT_HEADERS = [
    "Player", "W", "L", "SV", "IP", "BF", "H", "R", "ER", "SO", "BB", "HBP",
    "HR", "ERA", "WHIP",
]


def build_pit_row(p):
    ip_t = ip_frac_to_thirds(p.get("Pitching:IP", "0"))
    return [
        full_name(p),
        int(to_num(p.get("Pitching:W"))),
        int(to_num(p.get("Pitching:L"))),
        int(to_num(p.get("Pitching:SV"))),
        ip_thirds_to_str(ip_t),
        int(to_num(p.get("Pitching:BF"))),
        int(to_num(p.get("Pitching:H"))),
        int(to_num(p.get("Pitching:R"))),
        int(to_num(p.get("Pitching:ER"))),
        int(to_num(p.get("Pitching:SO"))),
        int(to_num(p.get("Pitching:BB"))),
        int(to_num(p.get("Pitching:HBP"))),
        int(to_num(p.get("Pitching:HR"))),
        (p.get("Pitching:ERA", "0.00").strip() or "0.00"),
        (p.get("Pitching:WHIP", "0.00").strip() or "0.00"),
    ]


def build_pit_sheet(ws, year, players, totals):
    rows = [
        build_pit_row(p) for p in players
        if ip_frac_to_thirds(p.get("Pitching:IP", "0")) > 0
    ]
    rows.sort(key=lambda r: -ip_frac_to_thirds(r[4]))
    if totals:
        ip_t = ip_frac_to_thirds(totals.get("Pitching:IP", "0"))
        team_row = [
            "TEAM",
            int(to_num(totals.get("Pitching:W"))),
            int(to_num(totals.get("Pitching:L"))),
            int(to_num(totals.get("Pitching:SV"))),
            ip_thirds_to_str(ip_t),
            int(to_num(totals.get("Pitching:BF"))),
            int(to_num(totals.get("Pitching:H"))),
            int(to_num(totals.get("Pitching:R"))),
            int(to_num(totals.get("Pitching:ER"))),
            int(to_num(totals.get("Pitching:SO"))),
            int(to_num(totals.get("Pitching:BB"))),
            int(to_num(totals.get("Pitching:HBP"))),
            int(to_num(totals.get("Pitching:HR"))),
            (totals.get("Pitching:ERA", "0.00").strip() or "0.00"),
            (totals.get("Pitching:WHIP", "0.00").strip() or "0.00"),
        ]
        rows.append(team_row)
    write_table(
        ws,
        title=f"{year} Gig Harbor Varsity — Pitching",
        headers=PIT_HEADERS,
        rows=rows,
        team_row_index=len(rows) - 1 if totals else None,
        col_widths=[14] + [7] * (len(PIT_HEADERS) - 1),
        left_align_cols={1},
    )


REC_HEADERS = ["Record", "Holder", "Value", "Qualifier / Note"]


def build_records_sheet(ws, year, players):
    """Compute the single-season records from the roster. Matches the 2019
    Historical sheet's record list (omits Longest Hitting Streak since it
    needs game-by-game data)."""
    rows = []

    def leader(extractor, qualifier=None, lowest=False):
        picks = []
        for p in players:
            if qualifier and not qualifier(p):
                continue
            v = extractor(p)
            if v is None:
                continue
            picks.append((v, full_name(p)))
        if not picks:
            return None, None
        picks.sort(key=lambda x: x[0], reverse=not lowest)
        best = picks[0][0]
        winners = [name for v, name in picks if v == best]
        return best, ", ".join(winners)

    def counting(stat_key):
        return lambda p: int(to_num(p.get(stat_key))) if to_num(p.get(stat_key)) > 0 else None

    def highest_avg(p):
        ab = to_num(p.get("Batting:AB"))
        if ab < 30:
            return None
        h = to_num(p.get("Batting:H"))
        return h / ab

    def best_obp(p):
        ab = to_num(p.get("Batting:AB"))
        if ab < 30:
            return None
        s = (p.get("Batting:OBP", "") or "").strip()
        if not s or s in ("-", "N/A"):
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def k_rate(p):
        ab = to_num(p.get("Batting:AB"))
        if ab < 30:
            return None
        so = to_num(p.get("Batting:SO"))
        return so / ab

    def lowest_era(p):
        t = ip_frac_to_thirds(p.get("Pitching:IP", "0"))
        if t < 21 * 3:  # 21 IP minimum, matches 2019 qualifier
            return None
        er = to_num(p.get("Pitching:ER"))
        ip = t / 3
        return (er * 7) / ip

    # Compute & gather
    v, who = leader(highest_avg)
    if v:
        # find AB and H for context
        avg_player = next(p for p in players if full_name(p) == who.split(",")[0].strip())
        ab = int(to_num(avg_player.get("Batting:AB")))
        h = int(to_num(avg_player.get("Batting:H")))
        rows.append(("Highest Average", who, fmt_avg(v), f"30+ AB ({h}-{ab})"))

    v, who = leader(counting("Batting:AB"))
    if v:
        rows.append(("Most At Bats", who, v, ""))

    v, who = leader(k_rate, lowest=True)
    if v is not None:
        target = next(p for p in players if full_name(p) == who.split(",")[0].strip())
        so = int(to_num(target.get("Batting:SO")))
        ab = int(to_num(target.get("Batting:AB")))
        rows.append(("Lowest K Ratio", who, f"{v*100:.1f}%", f"({so}-{ab})"))

    for label, key in [
        ("Most Doubles", "Batting:2B"),
        ("Most Triples", "Batting:3B"),
        ("Most Home Runs", "Batting:HR"),
        ("Most Walks", "Batting:BB"),
        ("Most Stolen Bases", "Batting:SB"),
        ("Most Total Bases", "Batting:TB"),
        ("Most RBIs", "Batting:RBI"),
    ]:
        v, who = leader(counting(key))
        if v:
            rows.append((label, who, v, ""))

    v, who = leader(best_obp)
    if v:
        rows.append(("Best On-Base Avg.", who, fmt_avg(v), "30+ AB"))

    v, who = leader(counting("Batting:R"))
    if v:
        rows.append(("Most Runs Scored", who, v, ""))

    for label, key in [
        ("Most Wins Pitching", "Pitching:W"),
        ("Most Innings Pitched", "Pitching:IP"),
        ("Most K's", "Pitching:SO"),
    ]:
        if key == "Pitching:IP":
            v, who = leader(lambda p: ip_frac_to_thirds(p.get(key, "0")) or None)
            if v:
                rows.append((label, who, ip_thirds_to_str(v), ""))
        else:
            v, who = leader(counting(key))
            if v:
                rows.append((label, who, v, ""))

    v, who = leader(lowest_era, lowest=True)
    if v is not None:
        rows.append(("Lowest ERA", who, f"{v:.2f}", "21+ IP"))

    v, who = leader(counting("Pitching:SV"))
    if v:
        rows.append(("Most Saves", who, v, ""))

    write_table(
        ws,
        title=f"{year} Gig Harbor Varsity — Individual Records",
        headers=REC_HEADERS,
        rows=rows,
        col_widths=[26, 30, 12, 22],
        left_align_cols={2, 4},
    )


# ─── Main ─────────────────────────────────────────────────────────────────

def build_year(year, csv_path):
    # Allow csv_path to be relative to ROOT or, for back-compat, a bare filename
    # in the Stats/ directory.
    p = (ROOT / csv_path) if "/" in str(csv_path) else (STATS_DIR / csv_path)
    players, totals = load_season(p)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Hitting & Fielding"
    build_hit_sheet(ws1, year, players, totals)

    ws2 = wb.create_sheet("Pitching")
    build_pit_sheet(ws2, year, players, totals)

    ws3 = wb.create_sheet("Individual Records")
    build_records_sheet(ws3, year, players)

    out_dir = OUT_DIR / str(year)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{year}_Season_Stats.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    for year, csv_name in SEASONS.items():
        build_year(year, csv_name)


if __name__ == "__main__":
    main()
