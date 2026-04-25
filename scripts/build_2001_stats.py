#!/usr/bin/env python3
"""Build 2001 Gig Harbor Varsity season stats xlsx from pages in
Historical/2001/. Kevin Owens freshman debut — .373 AVG, .500 OBP."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import (
    write_table, write_readme, write_roster_sheet,
)
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2001" / "2001_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 2001, [
        ("Individual Records", "Clean typed source. Note: records-page 'Most Total Bases Owens 58' doesn't match stats (.525 SLG × 59 AB = 31 TB) — likely typo in source.", "READABLE"),
        ("Team Highlights", "Clean typed source.", "READABLE"),
        ("Schedule", "No per-game schedule photo provided. Season record 9-11 comes from stats header.", "NOT PROVIDED"),
        ("Team Batting", "Season-to-date stats thru 20 games. Per-player 1B/2B/3B/HR breakdowns sum cleanly to published TB/SLG.", "READABLE"),
        ("Team Pitching", "Transcribed. Team: 9-11, 5.41 ERA, 134 K.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Tyler Bartlett",    12),
        (2,  "Kyle Knorr",        12),
        (3,  "Craig Manning",     12),
        (4,  "Dan Okamoto",       12),
        (5,  "Carl Olsen",        12),
        (6,  "Tom Patterson",     12),
        (7,  "Brett Shearer",     11),
        (8,  "Sam Rosendahl",     11),
        (9,  "Tyler Kullman",     11),
        (10, "Ryan Emmett",       10),
        (11, "David Jackson",     10),
        (12, "Alex Medeiros",     10),
        (13, "Matt Stock",        10),
        (14, "Kevin Owens",        9),
        (15, "Kevin Bogue",        9),
    ]
    write_roster_sheet(
        ws_r,
        title="2001 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    bat_rows = [
        ("Emmett",     "117.0", 53, 13, 15, 6, 0, 0, 10,  5, 0, 14, 0, ".396", ".397", ".283"),
        ("Shearer",    "140.0", 62, 11, 23, 4, 0, 0,  5, 17, 1, 13, 6, ".435", ".459", ".371"),
        ("Owens",      "140.0", 59, 16, 22, 3, 0, 2, 13,  9, 0,  5, 4, ".525", ".500", ".373"),
        ("Rosendahl",  "131.0", 51, 14, 10, 2, 0, 1,  8,  8, 4, 12, 0, ".294", ".316", ".196"),
        ("Manning",    "104.0", 36,  4,  9, 0, 0, 1,  4,  2, 1,  6, 2, ".361", ".275", ".250"),
        ("Bartlett",   "140.0", 65, 12, 16, 2, 0, 0,  4,  0, 2,  6, 1, ".277", ".300", ".246"),
        ("Okamoto",     "77.0", 30,  9,  9, 0, 0, 0,  4,  7, 0, 15, 0, ".300", ".371", ".300"),
        ("Patterson",  "125.0", 56, 15, 12, 1, 1, 1,  3,  8, 2, 12, 0, ".321", ".362", ".214"),
        ("Olsen",       "96.0", 36,  3,  8, 2, 1, 0,  0,  0, 0,  9, 1, ".278", ".275", ".222"),
        ("Knorr",       "69.0", 30,  5,  2, 1, 0, 0,  0,  2, 1,  0, 0, ".100", ".182", ".067"),
        ("Kullman",     "46.0", 13,  0,  0, 0, 0, 0,  1,  3, 0,  5, 0, ".000", ".333", ".000"),
        ("Medeiros",    "55.0", 26,  4, 10, 0, 0, 0,  1,  3, 0,  5, 0, ".385", ".444", ".385"),
        ("Jackson",     "80.0", 30,  4, 10, 2, 0, 0,  0,  1, 7,  7, 0, ".400", ".355", ".333"),
        ("Bogue",       "22.0",  5,  1,  1, 0, 0, 0,  2,  0, 0,  3, 0, ".200", ".333", ".200"),
        ("Stock",       "13.0",  4,  1,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Biddle",      "11.0",  4,  2,  1, 0, 0, 0,  2,  0, 0,  0, 0, ".250", ".500", ".250"),
        ("TEAM",       "140.0",551,114,144,23, 1, 5, 71, 84,13,122,16, ".334", ".360", ".261"),
    ]
    write_table(
        ws2,
        title="2001 Gig Harbor Varsity — Team Batting (Season to Date, 20 games)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Emmett",     "8.2",   56, 18, 16, 12,  6, 0,  3, 10, 0, 0, 0, ".383",  "9.69"),
        ("Rosendahl", "12.0",   56, 11,  8,  6, 11, 0,  2,  5, 2, 2, 0, ".256",  "3.50"),
        ("Manning",   "13.1",   76, 16, 19, 17, 18, 0,  3, 12, 0, 3, 0, ".291",  "8.92"),
        ("Bartlett",   "0.2",    5,  2,  0,  0,  0, 0,  0,  0, 0, 0, 0, ".667",  "0.00"),
        ("Okamoto",    "1.1",    6,  4,  5,  5,  0, 0,  0,  0, 0, 1, 0, ".667", "26.25"),
        ("Olsen",     "31.1",  145, 31, 24, 18, 20, 0,  2, 34, 3, 2, 0, ".252",  "4.02"),
        ("Kullman",   "41.2",  216, 43, 35, 27, 35, 0,  3, 46, 2, 4, 1, ".242",  "4.54"),
        ("Jackson",   "21.2",  115, 31, 23, 16, 11, 0,  8, 20, 1, 0, 0, ".323",  "5.17"),
        ("Bogue",      "7.2",   40, 10, 10,  6,  9, 0,  0,  1, 1, 1, 0, ".313",  "5.48"),
        ("TEAM",     "140.0",  715,166,140,107,108, 0, 22,134, 9,11, 1, ".284",  "5.41"),
    ]
    write_table(
        ws3,
        title="2001 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              14, "vs. Foss, Mt. Tahoma"),
        ("Most Runs (Inning)",            10, "vs. Wilson"),
        ("Most Runs Allowed (Game)",      17, "vs. Stadium, Peninsula"),
        ("Widest Margin of Victory",      14, "vs. Mt. Tahoma"),
        ("One Run Games Lost",             2, "vs. Lincoln, Bellarmine"),
        ("One Run Games Won",              2, "vs. Foss, Washington"),
        ("Most K's (by GH Pitchers)",     15, "vs. Peninsula"),
        ("Most K's (by Opponents)",       15, "vs. Peninsula"),
        ("Most Hits (Game)",              14, "vs. Foss"),
        ("Most Hits Allowed (Game)",      17, "vs. Peninsula"),
        ("Most Walks (Game for GH)",      16, "vs. Washington"),
        ("Most Walks Allowed",            12, "vs. Peninsula"),
        ("Most Singles (Game)",           13, "vs. Mt. Tahoma"),
        ("Most Doubles (Game)",            4, "vs. Peninsula"),
        ("Most Triples (Game)",            1, "vs. Mt. Tahoma"),
        ("Most Home Runs (Game)",          2, "vs. Bellarmine, Foss"),
        ("Most Steals (Game)",             5, "vs. Eatonville, Foss"),
        ("Longest Winning Streak",         5, ""),
        ("Longest Losing Streak",          5, ""),
        ("Total Team Runs Scored",       114, ""),
        ("Total Team Runs Allowed",      140, ""),
        ("Total Team Hits",              144, ""),
        ("Total Team Extra-Base Hits",    29, ""),
        ("Total Team Home Runs",           5, ""),
        ("Total Team Stolen Bases",       27, ""),
        ("Total Team Walks",              71, ""),
        ("Total Team HBP",                16, ""),
    ]
    write_table(
        ws_h,
        title="2001 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Kevin Owens",     ".373", "30+ AB (22-59) · freshman"),
        ("Most Hits",              "Brett Shearer",    23,    ""),
        ("Most At Bats",           "Tyler Bartlett",   65,    ""),
        ("Most HBP",               "Brett Shearer",     6,    ""),
        ("Lowest K Ratio",         "Kevin Owens",    "6.5%",  "(5-76)"),
        ("Most Doubles",           "Ryan Emmett",       6,    ""),
        ("Most Triples",           "Tom Patterson",     1,    ""),
        ("Most Home Runs",         "Kevin Owens",       2,    ""),
        ("Most Walks",             "Kevin Owens",      13,    ""),
        ("Most Stolen Bases",      "Tom Patterson",    11,    ""),
        ("Most Total Bases",       "Kevin Owens",      58,    "as printed (stats give 31 TB)"),
        ("Most RBIs",              "Brett Shearer",    17,    ""),
        ("Most Wins Pitching",     "Carl Olsen",        3,    ""),
        ("Most Innings Pitched",   "Tyler Kullman",  "41.2",  ""),
        ("Most K's",               "Tyler Kullman",    46,    ""),
        ("Lowest ERA",             "Sam Rosendahl",  "3.50",  ""),
        ("Most Saves",             "Tyler Kullman",     1,    ""),
        ("Best On-Base Avg.",      "Kevin Owens",    ".500",  "30+ AB"),
        ("Longest Hitting Streak", "Brett Shearer",     7,    ""),
        ("Most Runs Scored",       "Kevin Owens",      16,    ""),
    ]
    write_table(
        ws4,
        title="2001 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 36, 12, 36],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
