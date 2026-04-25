#!/usr/bin/env python3
"""Build 1995 Gig Harbor Varsity season stats xlsx."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1995" / "1995_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1995, [
        ("Individual Records", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No per-game schedule photo provided.", "NOT PROVIDED"),
        ("Team Batting", "Overall Final Batting Stats transcribed from photo.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Final Overall Pitching Stats transcribed. Team: 17-5-1, 1.35 ERA.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1, "Joel Miller",       12),
        (2, "Tom Friedman",      12),
        (3, "Jason Lippert",     12),
        (4, "Paul Baurichter",   12),
        (5, "Derek Vitcovich",   12),
        (6, "Kevin Graybill",    12),
        (7, "Kevin Feltus",      12),
        (8, "Rich Langford",     12),
        (9, "Aaron Ford",        12),
        (10, "Mike Miller",      11),
        (11, "Dan Iverson",      11),
        (12, "Aaron Araujo",     10),
        (13, "Matt Gardner",     10),
        (14, "Tim Friedman",     10),
    ]
    write_roster_sheet(ws_r, "1995 Gig Harbor Varsity — Roster", roster_rows,
                       coaches=["Peter Jansen", "Bob Maguinez", "Tony Anderson"])

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = ["Player", "G", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB", "HBP", "K", "SAC", "TB", "SB", "BA", "OBP"]
    bat_rows = [
        ("J. Miller",     21, 57, 27, 22, 22, 2,  2, 1, 15,  9, 6, 3, 31, 18, ".417", ".570"),
        ("Baurichter",    22, 69, 23, 25, 19, 1,  2, 0, 15,  1, 4, 1, 30, 13, ".368", ".576"),
        ("TW. Friedman",  22, 68, 22, 31, 27, 7,  1, 3, 12,  5, 7, 1, 49, 10, ".456", ".635"),
        ("Graybill",      20, 54, 11, 16, 17, 2,  1, 0,  9,  2, 5, 2, 21,  4, ".296", ".462"),
        ("M. Miller",     16, 39,  7,  9,  4, 1,  0, 0,  4,  0, 3, 1,  9,  3, ".231", ".372"),
        ("Gardner",       19, 19,  2,  1,  1, 0,  0, 0,  4,  2, 6, 0,  1,  0, ".053", ".280"),
        ("TJ. Friedman",  22, 48, 11, 14, 10, 3,  1, 0,  4,  3, 8, 2,  5, 19, ".292", ".382"),
        ("Lippert",       18, 13,  3,  2,  1, 0,  0, 0,  3,  1, 5, 0,  2,  3, ".154", ".375"),
        ("Ford",          18, 41, 14, 14,  8, 3,  0, 0,  5,  1,13, 0,  0, 17, ".341", ".489"),
        ("Feltus",        22, 48, 15, 14,  7, 2,  0, 0, 13,  1, 9, 4,  1,  6, ".292", ".484"),
        ("Langford",      22, 44, 13, 14,  7, 2,  0, 1,  7,  1,12, 1,  6,  3, ".304", ".481"),
        ("Vitcovich",     19,  8,  8,  4, 14, 3,  0, 0,  5,  1, 7, 0,  2, 17, ".292", ".438"),
        ("Araujo",        15,  7,  3,  5,  7, 3,  0, 0,  3,  0, 2, 1,  0,  8, ".294", ".450"),
        ("Iverson",       13, 21,  4,  7,  9, 4,  0, 0,  6,  0, 5, 3,  1,  3, ".333", ".538"),
        ("TEAM",          22,564,183,176,152,33, 10, 5,116, 21,89,19,146,58, ".312", ".479"),
    ]
    write_table(ws2, "1995 Gig Harbor Varsity — Team Batting",
                bat_headers, bat_rows, team_row_index=len(bat_rows) - 1,
                col_widths=[14, 5] + [6] * 14 + [7], left_align_cols={1})

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = ["Player", "W", "L", "SV", "IP", "AB", "R", "H", "K", "BB", "ER", "HBP", "G", "ERA", "OPPBA"]
    pit_rows = [
        ("Lippert",     7, 2, 1, "65.0",  235, 21, 49, 64, 33, 15, 2, 10, "1.62", ".209"),
        ("Gardner",     6, 2, 0, "49.0",  188, 13, 42, 57, 10,  8, 1,  8, "1.14", ".223"),
        ("T. Friedman", 2, 1, 0, "15.0",   49,  4,  6, 12,  5,  1, 1,  4, "0.47", ".122"),
        ("Araujo",      2, 0, 0, "19.0",   74,  7, 16, 15, 10,  3, 1,  6, "1.11", ".216"),
        ("Ford",        0, 0, 0, "0.2",     4,  2,  0,  1,  4,  2, 2,  2, "10.5",".000"),
        ("J. Miller",   0, 0, 0, "1.0",     3,  0,  0,  2,  0,  0, 0,  1, "0.00",".000"),
        ("TEAM",       17, 5, 1, "150.0", 553, 47,113,151, 62, 29, 7, 22, "1.35", ".207"),
    ]
    write_table(ws3, "1995 Gig Harbor Varsity — Team Pitching",
                pit_headers, pit_rows, team_row_index=len(pit_rows) - 1,
                col_widths=[14] + [6] * (len(pit_headers) - 1), left_align_cols={1})

    ws_h = wb.create_sheet("Team Highlights")
    write_table(ws_h, "1995 Gig Harbor Varsity — Team Highlights",
        ["Highlight", "Value", "Context"],
        [
            ("Most Runs (Game)",              16, "vs. Fife"),
            ("Most Runs (Inning)",             8, "vs. Peninsula"),
            ("Most Runs Allowed (Game)",       7, "vs. Sequim"),
            ("Widest Margin of Victory",      16, "vs. Fife"),
            ("One Run Games Won",              1, "vs. Washington"),
            ("One Run Games Lost",             3, "vs. Tahoma (×2), Franklin Pierce"),
            ("Most K's (by GH Pitchers)",     14, "vs. Foss"),
            ("Most K's (by Opponents)",        9, "vs. Washington"),
            ("Most Hits (Game)",              15, "vs. White River"),
            ("Most Hits Allowed (Game)",      11, "vs. N. Mason"),
            ("Most Walks (Game for GH)",      11, "vs. Peninsula"),
            ("Most Walks Allowed",             7, "vs. Enumclaw"),
            ("Most Singles (Game)",           12, "vs. White River"),
            ("Most Doubles (Game)",            6, "vs. Bethel"),
            ("Most Triples (Game)",            2, "vs. F.P."),
            ("Most Home Runs (Game)",          3, "vs. Enumclaw"),
            ("Most Steals (Game)",            10, "vs. White River"),
            ("Longest Winning Streak",        11, ""),
            ("Longest Losing Streak",          3, ""),
        ],
        col_widths=[34, 10, 44], left_align_cols={1, 3})

    ws4 = wb.create_sheet("Individual Records")
    write_table(ws4, "1995 Gig Harbor Varsity — Individual Records",
        ["Record", "Holder", "Value", "Qualifier / Note"],
        [
            ("Highest Average",        "Tom Friedman",                ".456", "30+ AB (31-68)"),
            ("Most Hits",              "Tom Friedman",                 31,    ""),
            ("Most At Bats",           "Tom Friedman",                 68,    ""),
            ("Most HBP",               "Joel Miller",                   9,    ""),
            ("Lowest K Ratio",         "Paul Baurichter",            "4.7%",  ""),
            ("Most Doubles",           "Tom Friedman",                  7,    ""),
            ("Most Triples",           "Joel Miller, Paul Baurichter",  2,    ""),
            ("Most Home Runs",         "Tom Friedman",                  3,    ""),
            ("Most Walks",             "Paul Baurichter, Joel Miller", 15,    ""),
            ("Most Stolen Bases",      "Joel Miller",                  18,    ""),
            ("Most Total Bases",       "Tom Friedman",                 49,    ""),
            ("Most RBIs",              "Tom Friedman",                 27,    ""),
            ("Most Wins Pitching",     "Jason Lippert",                 7,    ""),
            ("Most Innings Pitched",   "Jason Lippert",                65,    ""),
            ("Most K's",               "Jason Lippert",                64,    ""),
            ("Lowest ERA",             "Matt Gardner",              "1.14",   ""),
            ("Best On-Base Avg.",      "Tom Friedman",              ".635",  "30+ AB"),
            ("Longest Hitting Streak", "Tom Friedman",                  9,    ""),
            ("Most Runs Scored",       "Joel Miller",                  27,    ""),
        ],
        col_widths=[26, 36, 12, 22], left_align_cols={2, 4})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
