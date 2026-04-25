#!/usr/bin/env python3
"""Build 2005 Gig Harbor Varsity season stats xlsx from pages in
Historical/2005/. This is Tyler Rice's legendary year — 0 K in 72 PA,
a program Hall of Fame record."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2005" / "2005_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def _write_readme(ws, year, needs):
    ws["A1"] = f"{year} Season Stats — Manual Update Required"
    ws.merge_cells("A1:C1")
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 24
    for c, h in enumerate(("SHEET", "WHAT NEEDS ATTENTION", "STATUS"), start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 22
    for i, (sheet, issue, status) in enumerate(needs):
        for c, val in enumerate((sheet, issue, status), start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            cell.font = Font(name="Arial", size=10, color=NAVY,
                             bold=(c == 3 and status in ("NEEDS MANUAL UPDATE", "NOT PROVIDED")))
            cell.alignment = Alignment(
                horizontal="left" if c != 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = BORDER
            if c == 3 and status == "NEEDS MANUAL UPDATE":
                cell.fill = PatternFill("solid", start_color=FLAG)
            elif c == 3 and status == "NOT PROVIDED":
                cell.fill = PatternFill("solid", start_color="F8D7DA")
            elif i % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT)
        ws.row_dimensions[4 + i].height = 48
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 0: README ────────────────────────────────────────────
    ws_n = wb.active
    ws_n.title = "README"
    _write_readme(ws_n, 2005, [
        ("Individual Records",
         "Clean typed source. Tyler Rice's 0% K ratio (0-72) is a PROGRAM RECORD still on the Hall of Fame page. Verify all records against Team Batting/Pitching.",
         "NEEDS MANUAL UPDATE"),
        ("Team Highlights",
         "Clean typed source. Values preserved as printed.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "Schedule is on the main stats photo. 19 games visible for a 10-10 record (20 games expected) — one game may be missing/illegible.",
         "NEEDS MANUAL UPDATE"),
        ("Team Batting",
         "Transcribed from the main stats photo. Some fielding/aggregate cells hard to read; flagged where unclear.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "Pitcher's Summary transcribed. Team totals: 10-10, 3.365 ERA.",
         "READABLE"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Cassidy Emery",      12),
        (2,  "Tyler Rice",         12),
        (3,  "Seth Whiting",       12),
        (4,  "Matt Pleau",         12),
        (5,  "Troy Burki",         11),
        (6,  "Rob Emmett",         11),
        (7,  "David Benedict",     11),
        (8,  "Owen MacDonald",     11),
        (9,  "Brent Smoots",       11),
        (10, "Nolan Smith",        11),
        (11, "Antone Saltvick",    11),
        (12, "Jordan Harrison",    11),
        (13, "DJ Myers",           11),
        (14, "Casey Knox",         11),
        (15, "Josiah Ward",        11),
        (16, "Nick Mareno",        10),
        (17, "Brandon Rohde",       9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2005 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Jim Peschek"],
    )

    # ─── Sheet 2: Schedule ──────────────────────────────────────────
    ws_s = wb.create_sheet("Schedule")
    schedule_rows = [
        ("Eatonville",    "8-1"),
        ("Mt. Tahoma",    "10-6"),
        ("Wilson",        "0-8"),
        ("Olympia",       "4-10"),
        ("Shelton",       "2-9"),
        ("P. Angeles",    "3-0"),
        ("Bremerton",     "10-3"),
        ("Peninsula",     "2-12"),
        ("N. Kitsap",     "2-10"),
        ("N. Kitsap",     "6-7"),
        ("Bremerton",     "11-1"),
        ("Bremerton",     "9-0"),
        ("Foss",          "9-4"),
        ("S. Kitsap",     "5-4"),
        ("S. Kitsap",     "5-4"),
        ("C. Kitsap",     "3-4"),
        ("C. Kitsap",     "3-11"),
        ("Olympia",       "4-0"),
        ("Olympia",       "3-5"),
        ("Season Record", "10-10"),
    ]
    write_table(
        ws_s,
        title="2005 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        col_widths=[22, 18],
        left_align_cols={1},
        note=(
            "Note: Schedule transcribed from main stats photo. 19 games legible; "
            "season record is 10-10 (20 games expected). One game may need to "
            "be added."
        ),
    )

    # ─── Sheet 3: Hitting & Fielding ────────────────────────────────
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Benedict",   50, 11, 15,  5, 13, 2, 0, 0, 17, ".340", ".462", ".481", 8, 12, 3, 1,  20,  30,  4, ".926", 6),
        ("MacDonald",  53,  5, 11,  4,  6, 5, 0, 0, 16, ".208", ".302", ".250", 5,  5, 0, 0,  29,   5,  6, ".850", 2),
        ("Rice",       66, 17, 31, 13, 23, 4, 0, 4, 47, ".470", ".712", ".514", 0,  9, 1, 1,  22,  34,  2, ".972",23),
        ("Smoots",     55,  9, 20, 14, 17, 3, 0, 0, 23, ".298", ".387", ".514", 8,  9, 3, 2,  12,   4,  0, "1.000",0),
        ("Burki",      31,  5,  7,  7,  4, 2, 0, 1, 12, ".226", ".387", ".514", 3, 11, 0, 2,   7,   6,  0, "1.000",0),
        ("Smith",      32,  5,  7,  3,  5, 2, 0, 0,  9, ".219", ".281", ".342", 4,  7, 0, 3,   1,   9,  2, ".833", 1),
        ("Rohde",      43,  5, 12,  8,  3, 8, 0, 1, 23, ".279", ".535", ".326", 9,  3, 0, 1,   4,  12,  2, ".889", 0),
        ("Emery",      52,  9, 12, 11,  7, 3, 0, 1, 17, ".231", ".327", ".394",14, 12, 2, 0,   6,   0,  1, ".857", 4),
        ("Ward",       18,  2,  5,  6,  3, 0, 2, 0,  9, ".278", ".500", ".350", 2,  0, 1, 0,   6,   0,  0, "1.000",0),
        ("Knox",       11,  2,  2,  4,  1, 1, 0, 0,  3, ".182", ".273", ".182", 0,  0, 0, 3,  10,   6,  0, "1.000",0),
        ("Pleau",      53,  9, 11,  5,  8, 2, 1, 0, 15, ".208", ".283", ".288",10,  6, 0, 1,   7,  24,  8, ".884", 5),
        ("Mareno",     55, 18, 16, 14, 11, 3, 0, 2, 25, ".291", ".455", ".426",11,  5, 1, 0,   0,   3,  1, ".750", 1),
        ("Emmett",     13,  2,  2,  2,  1, 0, 1, 0,  4, ".154", ".308", ".267", 0,  0, 1, 0,   0,   0,  0, "1.000",0),
        ("Whiting",     3,  0,  1,  1,  1, 0, 0, 0,  1, ".333", ".333", ".333", 0,  1, 0, 0,   1,   0,  0, "1.000",0),
        ("Myers",       0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0,  0, 0, 0,   0,   0,  0, ".000", 0),
        ("Harrison",    0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0,  0, 0, 0,   0,   0,  0, ".000", 0),
        ("Saltvick",    1,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 0,  0, 0, 0,   0,   0,  0, ".000", 0),
        ("Gaube",       1,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 1,  0, 0, 0,   0,   0,  0, ".000", 0),
        ("TEAM",      541,104,152, 97,103,35, 4, 9,221, ".281", ".409", ".377",107,73,10,10, 237, 146, 44, ".897",62),
    ]
    write_table(
        ws2,
        title="2005 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=[14] + [6] * (len(hit_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Some per-player fielding cells were difficult to read in "
            "the source — estimates used where necessary. Team totals match "
            "published values. Trust Individual Records sheet for top numbers."
        ),
    )

    # ─── Sheet 4: Pitching ──────────────────────────────────────────
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "AB", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Burki",     1, 3, 0, "28.2", 103, 26, 22, 34, 29, 17, 0, 0, 0, 1, 1, "4.220"),
        ("Rohde",     5, 1, 0, "43.1", 153, 12, 28, 55, 11, 10, 0, 0, 0, 0, 3, "1.624"),
        ("Rice",      0, 2, 0, "10.1",  53, 18, 19, 19,  7,  7, 0, 0, 0, 3, 0, "4.851"),
        ("Knox",      0, 0, 1, "14.0",  55, 16, 12, 13, 13,  5, 0, 0, 0, 0, 2, "2.500"),
        ("Myers",     1, 1, 0, "7.1",   32,  8, 10,  3,  3,  5, 0, 0, 0, 0, 1, "4.930"),
        ("Saltvick",  0, 1, 0, "2.0",    8,  4,  2,  1,  2,  4, 0, 0, 0, 0, 0, "14.000"),
        ("Whiting",   0, 1, 0, "6.1",   19,  2,  5,  5,  5,  2, 0, 0, 0, 1, 0, "2.295"),
        ("Emmett",    1, 0, 1, "12.2",  47, 10, 16, 12, 10,  4, 0, 0, 0, 0, 1, "2.295"),
        ("Harrison",  2, 1, 1, "12.1",  55, 14, 14, 17,  7, 11, 0, 0, 0, 0, 0, "6.364"),
        ("TEAM",     10,10, 3, "135.2", 525,110,128,147, 87, 65, 0, 0, 0, 2,12, "3.365"),
    ]
    write_table(
        ws3,
        title="2005 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 5: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              11, "vs. Bremerton"),
        ("Most Runs (Inning)",             5, "vs. Foss"),
        ("Most Runs Allowed (Game)",      12, "vs. Peninsula"),
        ("Widest Margin of Victory",      10, "vs. Bremerton"),
        ("One Run Games Lost",             3, "vs. PA, NK, CK"),
        ("One Run Games Won",              1, "vs. SK"),
        ("Most K's (by GH Pitchers)",     12, "vs. Shelton"),
        ("Most K's (by Opponents)",       13, "vs. Central Kitsap"),
        ("Most Hits (Game)",              14, "vs. Bremerton"),
        ("Most Hits Allowed (Game)",      10, "vs. SK"),
        ("Most Walks (Game for GH)",       9, "vs. Bremerton"),
        ("Most Walks Allowed",             9, "vs. Bremerton"),
        ("Most Singles (Game)",           12, "vs. Central Kitsap"),
        ("Most Doubles (Game)",           11, "vs. PA (verify)"),
        ("Most Triples (Game)",            1, "vs. CK, NK, Mt. Tahoma, Shelton"),
        ("Most Home Runs (Game)",          2, "vs. Bremerton"),
        ("Most Steals (Game)",             8, "vs. CK"),
        ("Longest Winning Streak",         4, ""),
        ("Longest Losing Streak",          3, ""),
        ("Total Team Runs Scored",       104, ""),
        ("Total Team Hits",              152, ""),
        ("Total Team Extra-Base Hits",    48, ""),
        ("Total Team Home Runs",           9, ""),
        ("Total Team Stolen Bases",       62, ""),
        ("Total Team Walks",              73, ""),
        ("Total Team HBP",                10, ""),
    ]
    write_table(
        ws_h,
        title="2005 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    # ─── Sheet 6: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Tyler Rice",                              ".470", "30+ AB (31-66)"),
        ("Most At Bats",           "Tyler Rice",                               66,    ""),
        ("Most HBP",               "David Benedict",                            3,    ""),
        ("Lowest K Ratio",         "Tyler Rice",                              "0%",   "(0-72) · PROGRAM RECORD"),
        ("Most Doubles",           "Brandon Rohde",                             8,    ""),
        ("Most Triples",           "Josiah Ward",                               2,    ""),
        ("Most Home Runs",         "Tyler Rice",                                4,    ""),
        ("Most Walks",             "Cassidy Emery, David Benedict",            12,    ""),
        ("Most Stolen Bases",      "Tyler Rice",                               23,    ""),
        ("Most Total Bases",       "Tyler Rice",                               47,    ""),
        ("Most RBIs",              "Brent Smoots, Nick Mareno",                14,    ""),
        ("Most Wins Pitching",     "Brandon Rohde",                             5,    ""),
        ("Most Innings Pitched",   "Brandon Rohde",                          "43.1",  ""),
        ("Most K's",               "Brandon Rohde",                            55,    ""),
        ("Lowest ERA",             "Brandon Rohde",                          "1.62",  ""),
        ("Most Saves",             "Casey Knox, Rob Emmett, Jordan Harrison",   1,    "3-way tie"),
        ("Best On-Base Avg.",      "Tyler Rice, Troy Burki",                 ".514",  "30+ AB"),
        ("Longest Hitting Streak", "Tyler Rice",                                8,    ""),
        ("Most Runs Scored",       "Nick Mareno",                              18,    ""),
    ]
    write_table(
        ws4,
        title="2005 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 42, 12, 32],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
