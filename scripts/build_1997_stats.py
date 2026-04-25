#!/usr/bin/env python3
"""Build 1997 Gig Harbor Varsity season stats xlsx from pages in
Historical/1997/. STATE CHAMPIONSHIP YEAR (AA). Tim Friedman set multiple
program Hall of Fame records this season."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1997" / "1997_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1997, [
        ("Individual Records",
         "STATE CHAMPIONSHIP YEAR. Multiple Hall of Fame records set: Tim Friedman's 47 hits, 13 2B, 7 HR (tied), 83 TB, 41 RBI, 45 R, 20-game hit streak; Matt Gardner's 10 W; Aaron Araujo's 32 BB. Verify all.",
         "NEEDS MANUAL UPDATE"),
        ("Team Highlights",
         "Clean typed source. Team stats 280 H (PROGRAM RECORD) and 78 SB (PROGRAM RECORD) are on the site's Championship Banner.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "No per-game schedule photo provided. Season record 23-2 per stats header.",
         "NOT PROVIDED"),
        ("Team Batting",
         "Overall Final Stats transcribed. Matt Gardner's IP (59.1 per Records) vs. 44.2 in stats table — records sheet is authoritative.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "Transcribed. Team: 23-2, 2.93 ERA, 198 K.",
         "NEEDS SPOT-CHECK"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Chad Ahrens",        12),
        (2,  "Aaron Araujo",       12),
        (3,  "Donald Averill",     12),
        (4,  "Sam Baurichter",     12),
        (5,  "Tim Friedman",       12),
        (6,  "Matt Gardner",       12),
        (7,  "Robert Iversen",     12),
        (8,  "Drake Hano",         11),
        (9,  "Anthony Gilich",     11),
        (10, "Adam Harris",        11),
        (11, "Willie Keith",       11),
        (12, "Justin Fagering",    10),
        (13, "Kris Keller",        10),
        (14, "Dustin Bienias",     10),
        (15, "Kurt Elliott",       10),
        (16, "Derek Phill",        10),
        (17, "Shane Yelish",       10),
    ]
    write_roster_sheet(
        ws_r,
        title="1997 Gig Harbor Varsity — Roster · STATE AA CHAMPIONS",
        rows=roster_rows,
        coaches=["Pete Jansen", "Mike Moeller"],
        manager="Mike Boyle",
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    bat_rows = [
        ("Friedman",    "166.0", 88, 45, 47, 13, 1, 7, 18, 41, 0,  3, 6, ".943", ".651", ".534"),
        ("Gilich",      "165.0", 72, 24, 21,  3, 0, 1, 18, 23, 2,  6,10, ".486", ".620", ".292"),
        ("Araujo",      "169.0", 64, 33, 29,  5, 0, 1, 32, 22, 0,  7, 1, ".469", ".625", ".453"),
        ("Gardner",     "174.0", 89, 33, 41, 10, 0, 2,  7, 23, 1, 12, 2, ".629", ".523", ".461"),
        ("Averill",     "130.0", 65, 18, 23,  6, 0, 2,  4, 13, 0,  3,14, ".538", ".543", ".354"),
        ("R. Iversen",  "169.0", 83, 24, 33, 10, 1, 1,  7, 15, 2, 16, 3, ".579", ".500", ".398"),
        ("Ahrens",      "113.0", 47, 15, 14,  1, 0, 0,  3, 12, 0, 10, 2, ".319", ".449", ".298"),
        ("Harris",      "167.0", 77, 29, 31,  4, 0, 0,  6,  7, 0,  2, 6, ".455", ".471", ".403"),
        ("Keith",         "8.0",  2,  0,  1,  0, 0, 0,  1,  0, 0,  0, 0, ".500",".500",  ".500"),
        ("Fagering",     "93.0", 34,  7,  8,  3, 0, 0,  2,  1, 2,  4, 1, ".441", ".444", ".235"),
        ("Baurichter",  "132.0", 56, 15, 22,  3, 2, 1,  6, 19, 0,  4, 2, ".571", ".488", ".393"),
        ("Yelish",       "37.0",  7,  2,  0,  0, 0, 0,  0,  0, 0,  7, 0, ".000", ".000", ".000"),
        ("Phill",        "13.0",  3,  3,  1,  0, 0, 0,  2,  0, 0,  2, 0, ".333", ".500", ".333"),
        ("Elliott",      "42.0",  4,  1,  1,  0, 0, 0,  1,  1, 0,  2, 0, ".250", ".400", ".250"),
        ("Cleary",        "1.0",  0,  0,  0,  0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Freeman",       "4.0",  0,  0,  0,  0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Snow",          "1.0",  0,  0,  0,  0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("TEAM",        "179.0",747,270,280, 58, 4,15,159,243,25,106,29, ".525", ".563", ".375"),
    ]
    write_table(
        ws2,
        title="1997 Gig Harbor Varsity — Team Batting (Overall Final Stats · 23-2)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Gilich",     "15.1",  79, 25, 12,  8,  7, 0, 0,  4, 3, 0, 1, ".368",  "3.65"),
        ("Gardner",    "59.1", 199, 41, 23, 13, 26, 0, 4, 79,10, 0, 2, ".242",  "2.12"),
        ("Araujo",     "44.2", 243, 43, 23, 18, 16, 0, 2, 29, 4, 1, 2, ".209",  "2.82"),
        ("Baurichter", "22.0", 104, 13,  6,  4,  6, 0, 2, 18, 2, 1, 0, ".194",  "1.27"),
        ("Keller",     "27.2", 130, 21, 19, 22,  4, 0, 4, 33, 1, 0, 0, ".241",  "5.57"),
        ("Bienias",     "1.0",   7,  3,  1,  1,  2, 0, 0,  1, 0, 0, 0, ".571",  "7.00"),
        ("Elliott",     "1.0",   5,  2,  0,  0,  0, 0, 0,  0, 0, 0, 0, ".400",  "0.00"),
        ("TEAM",      "179.0", 776,152, 91, 73, 82, 5,17,198,23, 2, 5, ".196",  "2.93"),
    ]
    write_table(
        ws3,
        title="1997 Gig Harbor Varsity — Team Pitching · STATE CHAMPIONS",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              27, "vs. F.P."),
        ("Most Runs (Inning)",            10, "vs. Washington / F.P."),
        ("Most Runs Allowed (Game)",       8, "vs. Enumclaw, East Catholic, Capital"),
        ("Widest Margin of Victory",      20, "vs. F.P."),
        ("One Run Games Lost",             2, "vs. Enumclaw, Tahoma"),
        ("One Run Games Won",              2, "vs. Yelm, Capital"),
        ("Most K's (by GH Pitchers)",     13, "vs. Burlington Edison, Fife"),
        ("Most K's (by Opponents)",        9, "vs. East Catholic"),
        ("Most Hits (Game)",              17, "vs. F.P., Fife, Washington"),
        ("Most Hits Allowed (Game)",      15, "vs. Enumclaw"),
        ("Most Walks (Game for GH)",      16, "vs. F.P."),
        ("Most Walks Allowed",            10, "vs. Yelm"),
        ("Most Singles (Game)",           15, "vs. Peninsula"),
        ("Most Doubles (Game)",            7, "vs. Washington"),
        ("Most Triples (Game)",            2, "vs. F.P., Capital"),
        ("Most Home Runs (Game)",          3, "vs. F.P."),
        ("Most Steals (Game)",             8, "vs. F.P."),
        ("Longest Winning Streak",         9, ""),
        ("Longest Losing Streak",          1, ""),
        ("Total Team Runs Scored",       270, ""),
        ("Total Team Runs Allowed",       91, ""),
        ("Total Team Hits",              280, "PROGRAM RECORD — most team hits all-time"),
        ("Total Team Extra-Base Hits",    78, ""),
        ("Total Team Home Runs",          14, ""),
        ("Total Team Stolen Bases",       78, "PROGRAM RECORD — most team steals all-time"),
        ("Total Team Walks",             159, ""),
        ("Total Team HBP",                29, ""),
    ]
    write_table(
        ws_h,
        title="1997 Gig Harbor Varsity — Team Highlights · STATE AA CHAMPIONS",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 60],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Tim Friedman",                 ".534", "30+ AB (47-88)"),
        ("Most Hits",              "Tim Friedman",                  47,    "PROGRAM RECORD"),
        ("Most At Bats",           "Matt Gardner",                  89,    ""),
        ("Most HBP",               "Anthony Gilich",                 6,    ""),
        ("Lowest K Ratio",         "Tim Friedman",                "2.7%",  "(3-111)"),
        ("Most Doubles",           "Tim Friedman",                  13,    "PROGRAM RECORD"),
        ("Most Triples",           "Matt Gardner, Sam Baurichter",   2,    ""),
        ("Most Home Runs",         "Tim Friedman",                   7,    "PROGRAM RECORD (tied)"),
        ("Most Walks",             "Aaron Araujo",                  32,    "PROGRAM RECORD"),
        ("Most Stolen Bases",      "Tim Friedman",                  19,    ""),
        ("Most Total Bases",       "Tim Friedman",                  83,    "PROGRAM RECORD"),
        ("Most RBIs",              "Tim Friedman",                  41,    "PROGRAM RECORD"),
        ("Most Wins Pitching",     "Matt Gardner",                  10,    "PROGRAM RECORD"),
        ("Most Innings Pitched",   "Matt Gardner",                "59.1",  ""),
        ("Most K's",               "Matt Gardner",                  79,    ""),
        ("Lowest ERA",             "Matt Gardner",                "2.12",  ""),
        ("Most Saves",             "Anthony Gilich, Aaron Araujo",   2,    ""),
        ("Best On-Base Avg.",      "Tim Friedman",                ".651",  "30+ AB"),
        ("Longest Hitting Streak", "Tim Friedman",                  20,    "PROGRAM RECORD"),
        ("Most Runs Scored",       "Tim Friedman",                  45,    "PROGRAM RECORD"),
    ]
    write_table(
        ws4,
        title="1997 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 36, 12, 40],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
