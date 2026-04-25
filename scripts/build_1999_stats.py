#!/usr/bin/env python3
"""Build 1999 Gig Harbor Varsity season stats xlsx from pages in
Historical/1999/."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1999" / "1999_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1999, [
        ("Individual Records", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No per-game schedule photo provided.", "NOT PROVIDED"),
        ("Team Batting", "1999 Overall Final Stats transcribed.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Transcribed. Team: 3.52 ERA.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Mat Cleary",         12),
        (2,  "Kurt Elliott",       12),
        (3,  "Justin Fagering",    12),
        (4,  "Kris Keller",        12),
        (5,  "Justin O'Brien",     12),
        (6,  "Kurt Wright",        12),
        (7,  "Shane Yelish",       12),
        (8,  "Kevin Freeman",      11),
        (9,  "Pete Jendro",        11),
        (10, "Trevor Lyle",        11),
        (11, "Ryan Snow",          11),
        (12, "Jeff Vanderbilt",    11),
        (13, "John Lugo",          10),
        (14, "Craig Manning",      10),
        (15, "Dan Okamoto",        10),
        (16, "Carl Olsen",         10),
    ]
    write_roster_sheet(
        ws_r,
        title="1999 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Brad Gobel"],
        manager="Mike Boyle",
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    bat_rows = [
        ("Elliott",    "118.0", 68, 20, 21, 6, 5, 2,  8, 14, 0, 13, 1, ".632", ".347", ".309"),
        ("Jendro",      "95.0", 41, 17, 14, 3, 0, 0,  7,  8, 0,  6, 3, ".415", ".434", ".341"),
        ("Freeman",    "118.0", 58, 16, 23, 2, 1, 0,  8, 21, 0,  4, 1, ".466", ".431", ".397"),
        ("Keller",     "118.0", 48, 15,  7, 2, 1, 0,  5, 12, 0,  5, 0, ".292", ".306", ".146"),
        ("Yelish",     "118.0", 49, 19, 12, 1, 0, 0, 19,  5, 0,  8, 1, ".265", ".507", ".245"),
        ("Snow",       "118.0", 44,  5,  7, 1, 1, 0,  4,  5, 0,  9, 2, ".273", ".345", ".159"),
        ("Cleary",     "117.0", 52,  7, 15, 4, 0, 1,  7,  9, 0,  6, 4, ".423", ".381", ".288"),
        ("Vanderbilt", "113.0", 46, 11, 13, 3, 0, 0,  7,  6, 4,  5, 2, ".348", ".345", ".283"),
        ("Lyle",        "57.0", 23,  4,  4, 3, 0, 0,  3,  8, 0,  6, 0, ".304", ".269", ".174"),
        ("O'Brien",     "10.0",  4,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Fagering",   "110.0", 50,  7, 14, 2, 0, 0,  1,  3, 0, 12, 0, ".320", ".320", ".280"),
        ("Lyle",        "47.0", 11,  2,  2, 0, 0, 0,  2,  1, 0,  2, 0, ".182", ".308", ".182"),
        ("Okamoto",      "4.0",  1,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Manning",      "9.0",  0,  0,  0, 0, 0, 0,  1,  0, 0,  0, 0, ".000", "1.000", ".000"),
        ("Shearer",      "4.0",  0,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Lugo",        "18.0",  0,  0,  0, 0, 0, 0,  1,  0, 0,  2, 0, ".000", "1.000", ".000"),
        ("Olsen",        "4.0",  0,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("TEAM",       "118.0",491,119,152,37, 8, 6,105,111,11,109,12, ".454", ".412", ".310"),
    ]
    write_table(
        ws2,
        title="1999 Gig Harbor Varsity — Team Batting (Overall Final Stats)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Elliott",    "5.1",   31,  6,  6,  5,  9, 0,  0,  8, 0, 1, 0, ".286",  "6.56"),
        ("Keller",    "25.0",  126, 23, 17, 10,  9, 0,  4, 27, 4, 1, 0, ".185",  "1.40"),
        ("Wright",    "30.2",  144, 29, 18, 16, 28, 0,  7, 22, 2, 2, 2, ".220",  "2.56"),
        ("Snow",       "0.1",    4,  2,  3,  2,  2, 0,  0,  1, 0, 0, 0, ".500", "42.00"),
        ("O'Brien",    "2.0",    6,  0,  0,  0,  1, 0,  0,  0, 0, 0, 0, ".000",  "0.00"),
        ("Fagering",   "7.1",   42, 12,  8,  7,  4, 0,  0,  3, 1, 0, 0, ".324",  "3.76"),
        ("Lyle",      "35.0",  161, 48, 26, 20, 12, 0,  3, 13, 4, 2, 1, ".320",  "2.86"),
        ("Okamoto",    "3.0",   16,  3,  3,  2,  4, 0,  1,  2, 0, 0, 0, ".214",  "4.67"),
        ("Lugo",       "2.0",    8,  1,  1,  1,  3, 0,  0,  2, 0, 0, 0, ".125",  "3.50"),
        ("Manning",    "5.2",   20,  1,  3,  2,  4, 0,  1,  3, 0, 0, 0, ".071",  "2.47"),
        ("Olsen",      "2.0",    0,  0,  0,  0,  0, 0,  0,  0, 0, 0, 0, ".000",  "0.00"),
        ("TEAM",     "118.0",  541,116, 95, 57, 89, 0, 13, 89, 12,7, 3, ".255",  "3.52"),
    ]
    write_table(
        ws3,
        title="1999 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              14, "vs. Mt. Tahoma"),
        ("Most Runs (Inning)",             7, "vs. Mt. Tahoma"),
        ("Most Runs Allowed (Game)",      15, "vs. S. Kitsap"),
        ("Widest Margin of Victory",      14, "vs. Mt. Tahoma"),
        ("One Run Games Lost",           "None", ""),
        ("One Run Games Won",              1, "vs. Timberline"),
        ("Most K's (by GH Pitchers)",      9, "vs. Stadium"),
        ("Most K's (by Opponents)",       11, "vs. S. Kitsap"),
        ("Most Hits (Game)",              16, "vs. River Ridge"),
        ("Most Hits Allowed (Game)",      13, "vs. S. Kitsap"),
        ("Most Walks (Game for GH)",       8, "vs. S. Kitsap"),
        ("Most Walks Allowed",             8, "vs. Bellarmine"),
        ("Most Singles (Game)",           11, "vs. River Ridge"),
        ("Most Doubles (Game)",            6, "vs. Foss"),
        ("Most Triples (Game)",            2, "vs. Mt. Tahoma"),
        ("Most Home Runs (Game)",          3, "vs. River Ridge"),
        ("Most Steals (Game)",             6, "vs. Lincoln"),
        ("Longest Winning Streak",         3, ""),
        ("Longest Losing Streak",          2, ""),
        ("Total Team Runs Scored",       119, ""),
        ("Total Team Runs Allowed",       95, ""),
        ("Total Team Hits",              152, ""),
        ("Total Team Extra-Base Hits",    51, ""),
        ("Total Team Home Runs",           6, ""),
        ("Total Team Stolen Bases",       23, ""),
        ("Total Team Walks",              85, ""),
        ("Total Team HBP",                12, ""),
    ]
    write_table(
        ws_h,
        title="1999 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Kevin Freeman",                     ".397", "30+ AB (23-58)"),
        ("Most Hits",              "Kevin Freeman",                      23,    ""),
        ("Most At Bats",           "Kurt Elliott",                       68,    ""),
        ("Most HBP",               "Kurt Wright",                         4,    ""),
        ("Lowest K Ratio",         "Kevin Freeman",                    "11.4%", "(8-70)"),
        ("Most Doubles",           "Kurt Elliott",                        6,    ""),
        ("Most Triples",           "Kurt Elliott",                        5,    ""),
        ("Most Home Runs",         "Kurt Elliott, Kurt Wright",           2,    ""),
        ("Most Walks",             "Shane Yelish",                       19,    ""),
        ("Most Stolen Bases",      "Shane Yelish, Justin Fagering",       4,    ""),
        ("Most Total Bases",       "Kurt Elliott",                       43,    ""),
        ("Most RBIs",              "Kevin Freeman",                      21,    ""),
        ("Most Wins Pitching",     "Trevor Lyle",                         4,    ""),
        ("Most Innings Pitched",   "Trevor Lyle",                        35,    ""),
        ("Most K's",               "Kris Keller",                        27,    ""),
        ("Lowest ERA",             "Kris Keller",                      "1.40",  ""),
        ("Most Saves",             "Kurt Wright",                         2,    ""),
        ("Best On-Base Avg.",      "Shane Yelish",                     ".507",  "30+ AB"),
        ("Longest Hitting Streak", "Kevin Freeman",                       8,    ""),
        ("Most Runs Scored",       "Kurt Elliott",                       20,    ""),
    ]
    write_table(
        ws4,
        title="1999 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 40, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
