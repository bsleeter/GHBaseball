"""Convert the 2026 GameChanger stats CSV into the canonical xlsx schema.

GameChanger exports one wide CSV with 174 columns covering batting,
pitching, and fielding plus three section labels. We map the columns
we care about into the same xlsx layout used by 2023/2024/2025
(Hitting & Fielding + Pitching sheets) so build_master_data.py picks
it up unchanged.

Run any time the CSV is refreshed; output overwrites
Historical/2026/2026_Season_Stats.xlsx.
"""
from __future__ import annotations

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "Historical" / "2026" / "Gig Harbor Varsity Tides Spring 2026 Stats.csv"
XLSX_OUT = ROOT / "Historical" / "2026" / "2026_Season_Stats.xlsx"

# Canonical sheet column layouts (mirrored from 2024/2025).
BAT_HEADER = ["Player", "GP", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR",
              "TB", "AVG", "SLG", "OBP", "OPS", "SO", "BB", "HBP", "SAC", "SF",
              "SB", "PO", "A", "E", "FLD%"]
PIT_HEADER = ["Player", "W", "L", "SV", "IP", "BF", "H", "R", "ER", "SO", "BB",
              "HBP", "HR", "ERA", "WHIP"]


def num(s):
    """Parse a stat-cell value to int/float when possible, else pass through.
    Empty strings / placeholders ('-', 'N/A') become None."""
    if s is None:
        return None
    t = s.strip() if isinstance(s, str) else s
    if t in ("", "-", "N/A", "—"):
        return None
    if isinstance(t, (int, float)):
        return t
    try:
        if "." in t:
            return float(t)
        return int(t)
    except (ValueError, AttributeError):
        return t  # leave as string (e.g., for AVG ".352" we keep the string)


def passthrough(s):
    """Keep ratio-style strings ('.352') as-is so the xlsx reads identically
    to a hand-typed sheet. Empty becomes None."""
    if s is None:
        return None
    t = s.strip() if isinstance(s, str) else s
    if t in ("", "-", "N/A", "—"):
        return None
    return t


def load_sections():
    """Return (batting_section_rows, pitching_section_rows, fielding_section_rows,
    headers_for_each, totals_row_per_section)."""
    with open(CSV_PATH, encoding="utf-8-sig") as f:
        all_rows = list(csv.reader(f))
    header = all_rows[1]
    ip_idx = header.index("IP")
    tc_idx = header.index("TC")

    bat_hdr = header[:ip_idx]
    pit_hdr = header[ip_idx:tc_idx]
    fld_hdr = header[tc_idx:]

    # Player rows: skip Totals + Glossary footer rows.
    player_rows = []
    totals_row = None
    for r in all_rows[2:]:
        if not r or not r[0]:
            continue
        if r[0].lower() == "totals":
            totals_row = r
            continue
        if r[0].lower() == "glossary":
            break
        # Skip rows that don't start with a jersey number (defensive)
        try:
            int(r[0])
        except (ValueError, TypeError):
            continue
        player_rows.append(r)

    return player_rows, totals_row, ip_idx, tc_idx, bat_hdr, pit_hdr, fld_hdr


def cell(row, hdr, key, *, offset=0):
    """Return the value at the column matching `key` in the section headed
    by `hdr`, accounting for global-row offset."""
    try:
        idx = hdr.index(key)
    except ValueError:
        return None
    return row[offset + idx] if offset + idx < len(row) else None


def build_batting_lines(player_rows, ip_idx, tc_idx, bat_hdr, fld_hdr):
    """Each output row matches BAT_HEADER. Pulls AB-block fields from the
    batting section and PO/A/E/FLD% from the fielding section."""
    out = []
    for r in player_rows:
        last = r[bat_hdr.index("Last")].strip()
        first = r[bat_hdr.index("First")].strip()
        full_name = f"{first} {last}".strip()
        line = [
            full_name,
            num(cell(r, bat_hdr, "GP")),
            num(cell(r, bat_hdr, "AB")),
            num(cell(r, bat_hdr, "R")),
            num(cell(r, bat_hdr, "H")),
            num(cell(r, bat_hdr, "RBI")),
            num(cell(r, bat_hdr, "1B")),
            num(cell(r, bat_hdr, "2B")),
            num(cell(r, bat_hdr, "3B")),
            num(cell(r, bat_hdr, "HR")),
            num(cell(r, bat_hdr, "TB")),
            passthrough(cell(r, bat_hdr, "AVG")),
            passthrough(cell(r, bat_hdr, "SLG")),
            passthrough(cell(r, bat_hdr, "OBP")),
            passthrough(cell(r, bat_hdr, "OPS")),
            num(cell(r, bat_hdr, "SO")),
            num(cell(r, bat_hdr, "BB")),
            num(cell(r, bat_hdr, "HBP")),
            num(cell(r, bat_hdr, "SAC")),
            num(cell(r, bat_hdr, "SF")),
            num(cell(r, bat_hdr, "SB")),
            num(cell(r, fld_hdr, "PO", offset=tc_idx)),
            num(cell(r, fld_hdr, "A", offset=tc_idx)),
            num(cell(r, fld_hdr, "E", offset=tc_idx)),
            passthrough(cell(r, fld_hdr, "FPCT", offset=tc_idx)),
        ]
        out.append(line)
    return out


def build_pitching_lines(player_rows, ip_idx, tc_idx, pit_hdr):
    """Each output row matches PIT_HEADER. Skip players with 0 IP."""
    out = []
    for r in player_rows:
        last = r[1].strip()
        first = r[2].strip()
        full_name = f"{first} {last}".strip()

        ip_val = passthrough(cell(r, pit_hdr, "IP", offset=ip_idx))
        # Skip pitchers with no innings recorded (avoids cluttering pitching
        # sheet with every batter).
        try:
            if ip_val is None or float(ip_val) == 0:
                continue
        except (ValueError, TypeError):
            continue

        line = [
            full_name,
            num(cell(r, pit_hdr, "W", offset=ip_idx)),
            num(cell(r, pit_hdr, "L", offset=ip_idx)),
            num(cell(r, pit_hdr, "SV", offset=ip_idx)),
            ip_val,
            num(cell(r, pit_hdr, "BF", offset=ip_idx)),
            num(cell(r, pit_hdr, "H", offset=ip_idx)),
            num(cell(r, pit_hdr, "R", offset=ip_idx)),
            num(cell(r, pit_hdr, "ER", offset=ip_idx)),
            num(cell(r, pit_hdr, "SO", offset=ip_idx)),
            num(cell(r, pit_hdr, "BB", offset=ip_idx)),
            num(cell(r, pit_hdr, "HBP", offset=ip_idx)),
            num(cell(r, pit_hdr, "HR", offset=ip_idx)),
            passthrough(cell(r, pit_hdr, "ERA", offset=ip_idx)),
            passthrough(cell(r, pit_hdr, "WHIP", offset=ip_idx)),
        ]
        out.append(line)
    return out


def build_team_totals(totals_row, ip_idx, tc_idx, bat_hdr, pit_hdr, fld_hdr):
    """Map the GameChanger Totals row to the TEAM lines for both sheets."""
    if totals_row is None:
        return None, None
    team_bat = [
        "TEAM",
        num(cell(totals_row, bat_hdr, "GP")),
        num(cell(totals_row, bat_hdr, "AB")),
        num(cell(totals_row, bat_hdr, "R")),
        num(cell(totals_row, bat_hdr, "H")),
        num(cell(totals_row, bat_hdr, "RBI")),
        num(cell(totals_row, bat_hdr, "1B")),
        num(cell(totals_row, bat_hdr, "2B")),
        num(cell(totals_row, bat_hdr, "3B")),
        num(cell(totals_row, bat_hdr, "HR")),
        num(cell(totals_row, bat_hdr, "TB")),
        passthrough(cell(totals_row, bat_hdr, "AVG")),
        passthrough(cell(totals_row, bat_hdr, "SLG")),
        passthrough(cell(totals_row, bat_hdr, "OBP")),
        passthrough(cell(totals_row, bat_hdr, "OPS")),
        num(cell(totals_row, bat_hdr, "SO")),
        num(cell(totals_row, bat_hdr, "BB")),
        num(cell(totals_row, bat_hdr, "HBP")),
        num(cell(totals_row, bat_hdr, "SAC")),
        num(cell(totals_row, bat_hdr, "SF")),
        num(cell(totals_row, bat_hdr, "SB")),
        num(cell(totals_row, fld_hdr, "PO", offset=tc_idx)),
        num(cell(totals_row, fld_hdr, "A", offset=tc_idx)),
        num(cell(totals_row, fld_hdr, "E", offset=tc_idx)),
        passthrough(cell(totals_row, fld_hdr, "FPCT", offset=tc_idx)),
    ]
    team_pit = [
        "TEAM",
        num(cell(totals_row, pit_hdr, "W", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "L", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "SV", offset=ip_idx)),
        passthrough(cell(totals_row, pit_hdr, "IP", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "BF", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "H", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "R", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "ER", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "SO", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "BB", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "HBP", offset=ip_idx)),
        num(cell(totals_row, pit_hdr, "HR", offset=ip_idx)),
        passthrough(cell(totals_row, pit_hdr, "ERA", offset=ip_idx)),
        passthrough(cell(totals_row, pit_hdr, "WHIP", offset=ip_idx)),
    ]
    return team_bat, team_pit


def _bold(c, fill=False):
    c.font = Font(bold=True)
    if fill:
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    c.alignment = Alignment(horizontal="center")


def write_workbook(batting, team_bat, pitching, team_pit, individual_records):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Hitting & Fielding ──
    ws = wb.create_sheet("Hitting & Fielding")
    ws.cell(row=1, column=1, value="2026 Gig Harbor Varsity — Hitting & Fielding").font = Font(bold=True, size=12)
    for ci, h in enumerate(BAT_HEADER, start=1):
        _bold(ws.cell(row=3, column=ci, value=h), fill=True)
    for ri, line in enumerate(batting, start=4):
        for ci, val in enumerate(line, start=1):
            ws.cell(row=ri, column=ci, value=val)
    if team_bat is not None:
        team_row = 4 + len(batting) + 1
        for ci, val in enumerate(team_bat, start=1):
            _bold(ws.cell(row=team_row, column=ci, value=val))

    # ── Pitching ──
    ws = wb.create_sheet("Pitching")
    ws.cell(row=1, column=1, value="2026 Gig Harbor Varsity — Pitching").font = Font(bold=True, size=12)
    for ci, h in enumerate(PIT_HEADER, start=1):
        _bold(ws.cell(row=3, column=ci, value=h), fill=True)
    for ri, line in enumerate(pitching, start=4):
        for ci, val in enumerate(line, start=1):
            ws.cell(row=ri, column=ci, value=val)
    if team_pit is not None:
        team_row = 4 + len(pitching) + 1
        for ci, val in enumerate(team_pit, start=1):
            _bold(ws.cell(row=team_row, column=ci, value=val))

    # ── Individual Records ──
    # Carry over what was in the previous xlsx (typed by hand). Re-reading
    # those records is out of scope for the CSV converter — preserve them.
    ws = wb.create_sheet("Individual Records")
    ws.cell(row=1, column=1, value="2026 Gig Harbor Varsity — Individual Records").font = Font(bold=True, size=12)
    for ci, h in enumerate(["Record", "Holder", "Value", "Qualifier / Note"], start=1):
        _bold(ws.cell(row=3, column=ci, value=h), fill=True)
    for ri, rec in enumerate(individual_records, start=4):
        for ci, val in enumerate(rec, start=1):
            ws.cell(row=ri, column=ci, value=val)

    wb.save(XLSX_OUT)


def read_existing_individual_records():
    """Pull existing typed records from the prior xlsx so we don't lose them.
    Returns list of (Record, Holder, Value, Qualifier) tuples."""
    from openpyxl import load_workbook
    if not XLSX_OUT.exists():
        return []
    wb = load_workbook(XLSX_OUT, data_only=True)
    if "Individual Records" not in wb.sheetnames:
        return []
    ws = wb["Individual Records"]
    out = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        out.append(tuple(row[:4]))
    return out


def main():
    rows, totals, ip_idx, tc_idx, bat_hdr, pit_hdr, fld_hdr = load_sections()
    batting = build_batting_lines(rows, ip_idx, tc_idx, bat_hdr, fld_hdr)
    pitching = build_pitching_lines(rows, ip_idx, tc_idx, pit_hdr)
    team_bat, team_pit = build_team_totals(totals, ip_idx, tc_idx, bat_hdr, pit_hdr, fld_hdr)
    # Don't preserve typed Individual Records — they were captured at an
    # earlier snapshot of the season. The year-page Hall of Fame is
    # auto-computed from the battingLines/pitchingLines below, so the new
    # CSV's per-player numbers drive the displayed records automatically.
    individual_records: list = []

    write_workbook(batting, team_bat, pitching, team_pit, individual_records)
    print(f"Wrote {XLSX_OUT}")
    print(f"  batting lines: {len(batting)} (+ TEAM: {team_bat is not None})")
    print(f"  pitching lines: {len(pitching)} (+ TEAM: {team_pit is not None})")
    print(f"  individual records: cleared (Hall of Fame auto-computes)")


if __name__ == "__main__":
    main()
