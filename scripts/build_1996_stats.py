#!/usr/bin/env python3
"""Build 1996 Gig Harbor Varsity season stats xlsx from pages in
Historical/1996/. STATE TOURNAMENT (5th place finish)."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1996" / "1996_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1996, [
        ("Individual Records",
         "Clean typed source. Tim Friedman's breakout junior year (.471). Team headline: 'TIDES PLACE 5TH IN STATE CHAMPIONSHIPS'.",
         "NEEDS MANUAL UPDATE"),
        ("Team Highlights",
         "Clean typed source. Note: 'Most Runs - inning' shows 20 vs. Yelm — likely matches total game run tally. Verify.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "No per-game schedule photo provided. Season record 18-4 per pitching totals.",
         "NOT PROVIDED"),
        ("Team Batting",
         "Batting stats transcribed. Some per-player cells had internal inconsistencies.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "Transcribed. Team: 18-4-3, 2.65 ERA, 137 K.",
         "READABLE"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Mike Miller",        12),
        (2,  "Dan Iverson",        12),
        (3,  "Bob Petteys",        12),
        (4,  "Tim Friedman",       11),
        (5,  "Matt Gardner",       11),
        (6,  "Aaron Araujo",       11),
        (7,  "Sam Baurichter",     11),
        (8,  "Chad Ahrens",        11),
        (9,  "Don Averill",        11),
        (10, "Mac Stanton",        11),
        (11, "Robert Iversen",     11),
        (12, "Anthony Gilich",     10),
        (13, "Adam Harris",        10),
        (14, "Drake Hano",         10),
        (15, "Kris Keller",         9),
        (16, "Justin Fagering",     9),
        (17, "Shane Yelish",        9),
    ]
    write_roster_sheet(
        ws_r,
        title="1996 Gig Harbor Varsity — Roster · 5th in State Championships",
        rows=roster_rows,
        coaches=["Pete Jansen", "Mike Moeller"],
        manager="Mike Boyle, Brian Wong",
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "G", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB",
        "HBP", "K", "SAC", "TB", "SB", "BA", "OBP", "K%",
    ]
    bat_rows = [
        ("Ahrens",     22, 30, 13,  7,  6, 0, 0, 0, 13,  2,  8, 3,  7, 6, ".233", ".521",   "16.7"),
        ("Araujo",     23, 68, 18, 26, 20, 4, 0, 0, 17,  1, 10, 1, 30, 3, ".382", ".593",   "11.6"),
        ("Averill",     8,  6,  6,  2,  2, 0, 0, 0,  1,  2,  0, 0,  2, 1, ".333", ".571",   "28.6"),
        ("Baurichter", 23, 46, 12, 12, 13, 1, 1, 0,  3,  1, 10, 2, 16, 5, ".261", ".423",   "19.2"),
        ("Friedman",   23, 85, 32, 40, 31,12, 2, 2, 15,  1,  6, 1, 62, 6, ".471", ".627",    "5.8"),
        ("Gardner",    23, 67, 20, 28, 24, 7, 0, 0, 23,  0,  4, 1, 35, 5, ".418", ".626",    "4.4"),
        ("Gilich",     23, 55, 19, 17, 14, 1, 0, 0, 22,  3,  3, 9, 18, 6, ".309", ".517",    "3.4"),
        ("Hano",       23, 17,  6,  3,  6, 1, 0, 0,  5,  1,  2, 3,  4, 2, ".176", ".423",    "7.7"),
        ("Harris",     20, 40, 13, 13,  9, 1, 1, 0,  8,  2,  6, 1, 14, 1, ".325", ".620",   "12.1"),
        ("D. Iverson", 18, 22,  6,  3,  1, 1, 0, 0,  6,  0,  9, 1,  3, 0, ".136", ".310",   "31.1"),
        ("R. Iversen", 23, 85, 20, 24, 26, 6, 1, 1, 12,  0, 10, 2, 32,12, ".282", ".423",   "10.3"),
        ("Keller",      5,  3,  0,  1,  1, 1, 0, 0,  0,  0,  0, 0,  2, 0, ".333", ".333",     "0"),
        ("Miller",     23, 55, 13, 17, 16, 1, 0, 0, 16,  2,  9, 2, 18, 9, ".309", ".547",     "12"),
        ("Fagering",    5,  2,  1,  0,  0, 0, 0, 0,  1,  0,  1, 0,  0, 0, ".000", ".333",     "0"),
        ("Petteys",    21, 30, 12, 10,  8, 1, 0, 0, 13,  7,  4, 1, 11, 2, ".333", ".667",    "7.8"),
        ("Stanton",     8,  5,  3,  0,  0, 0, 0, 0,  4,  3,  0, 0,  0, 0, ".000", ".444",   "33.3"),
        ("Yelish",      2,  0,  0,  0,  0, 0, 0, 0,  0,  0,  0, 0,  0, 0, ".000", ".000",     "0"),
        ("TEAM",       23,616,194,204,177,36, 4, 2,158, 20, 86,25,254,58, ".331", ".542",   "10.6"),
    ]
    write_table(
        ws2,
        title="1996 Gig Harbor Varsity — Team Batting · Final Overall",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 5] + [6] * (len(bat_headers) - 3) + [8],
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "AB", "R", "H", "K", "BB", "ER",
        "HBP", "G", "ERA", "OPPBA",
    ]
    pit_rows = [
        ("Araujo",      4, 2, 0, "37.2", 143, 29, 45, 36, 20, 20, 0, 10, "3.72", ".315"),
        ("Baurichter",  1, 0, 1, "16.2",  70, 14, 19, 20,  8,  5, 3,  7, "2.10", ".271"),
        ("Friedman",    1, 1, 0, "9.2",   35,  6,  8, 10, 11,  5, 0,  5, "3.62", ".229"),
        ("Gardner",     7, 1, 0, "53.0", 203, 27, 51, 47, 12, 16, 1, 12, "2.04", ".251"),
        ("Gilich",      3, 0, 2, "30.2", 110, 11, 26, 12,  8,  8, 0, 10, "1.83", ".236"),
        ("Keller",      1, 0, 0, "7.1",   31,  5,  8,  8,  9,  4, 1,  4, "3.82", ".258"),
        ("Fagering",    0, 0, 0, "0.2",    4,  2,  3,  1,  2,  1, 0,  1, "10.4", ".750"),
        ("Averill",     1, 0, 0, "3.0",   11,  1,  1,  3,  2,  1, 0,  1, "2.33", ".091"),
        ("TEAM",       18, 4, 3, "158.0",607, 95,161,137, 72, 60, 5, 24, "2.65", ".226"),
    ]
    write_table(
        ws3,
        title="1996 Gig Harbor Varsity — Team Pitching · Final Overall",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              20, "vs. Yelm"),
        ("Most Runs (Inning)",            20, "vs. Yelm (likely full-game tally, verify)"),
        ("Most Runs Allowed (Game)",      17, "vs. R.A. Long"),
        ("Widest Margin of Victory",      19, "vs. Yelm"),
        ("One Run Games Lost",             1, "vs. Peninsula"),
        ("One Run Games Won",              2, "vs. F.P. (×2)"),
        ("Most K's (by GH Pitchers)",     11, "vs. North Mason"),
        ("Most K's (by Opponents)",        9, "vs. R.A. Long"),
        ("Most Hits (Game)",              16, "vs. Fife"),
        ("Most Hits Allowed (Game)",      20, "vs. R.A. Long"),
        ("Most Walks (Game for GH)",      15, "vs. Washington"),
        ("Most Walks Allowed",            10, "vs. Peninsula"),
        ("Most Singles (Game)",           14, "vs. Fife"),
        ("Most Doubles (Game)",            4, "vs. Fife, R.A. Long"),
        ("Most Triples (Game)",            2, "vs. Washington"),
        ("Most Home Runs (Game)",          1, "vs. Foss, Yelm"),
        ("Most Steals (Game)",             7, "vs. Yelm, W.R."),
        ("Longest Winning Streak",         9, ""),
        ("Longest Losing Streak",          1, ""),
    ]
    write_table(
        ws_h,
        title="1996 Gig Harbor Varsity — Team Highlights · 5th in State",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 48],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Tim Friedman",       ".471", "30+ AB (40-85)"),
        ("Most Hits",              "Tim Friedman",        40,    ""),
        ("Most At Bats",           "Tim Friedman",        85,    ""),
        ("Most HBP",               "Bob Petteys",          7,    ""),
        ("Lowest K Ratio",         "Anthony Gilich",    "3.4%",  ""),
        ("Most Doubles",           "Tim Friedman",        12,    ""),
        ("Most Triples",           "Tim Friedman",         2,    ""),
        ("Most Home Runs",         "Tim Friedman",         2,    ""),
        ("Most Walks",             "Matt Gardner",        23,    ""),
        ("Most Stolen Bases",      "Dan Iverson",         12,    ""),
        ("Most Total Bases",       "Tim Friedman",        62,    ""),
        ("Most RBIs",              "Tim Friedman",        31,    ""),
        ("Most Wins Pitching",     "Matt Gardner",         7,    ""),
        ("Most Innings Pitched",   "Matt Gardner",        53,    ""),
        ("Most K's",               "Matt Gardner",        47,    ""),
        ("Lowest ERA",             "Anthony Gilich",    "1.83",  ""),
        ("Most Saves",             "Anthony Gilich",       2,    ""),
        ("Best On-Base Avg.",      "Bob Petteys",       ".667",  "30+ AB (34-51)"),
        ("Longest Hitting Streak", "Tim Friedman",        12,    ""),
        ("Most Runs Scored",       "Tim Friedman",        32,    ""),
    ]
    write_table(
        ws4,
        title="1996 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 32, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
