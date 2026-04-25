#!/usr/bin/env python3
"""Build 2003 Gig Harbor Varsity season stats xlsx from pages in
Historical/2003/. Kevin Owens (junior) had .464 AVG / .580 OBP — program
records contender."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2003" / "2003_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def _write_readme(ws, year, needs):
    ws["A1"] = f"{year} Season Stats — Manual Update Required"
    ws.merge_cells("A1:C1")
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 24
    for c, h in enumerate(("SHEET", "WHAT NEEDS ATTENTION", "STATUS"), start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 22
    for i, (sheet, issue, status) in enumerate(needs):
        for c, val in enumerate((sheet, issue, status), start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            cell.font = Font(name="Arial", size=10, color=NAVY,
                             bold=(c == 3 and status in ("NEEDS MANUAL UPDATE", "NOT PROVIDED")))
            cell.alignment = Alignment(
                horizontal="left" if c != 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = BORDER
            if c == 3 and status == "NEEDS MANUAL UPDATE":
                cell.fill = PatternFill("solid", start_color=FLAG)
            elif c == 3 and status == "NOT PROVIDED":
                cell.fill = PatternFill("solid", start_color="F8D7DA")
            elif i % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT)
        ws.row_dimensions[4 + i].height = 48
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 0: README ────────────────────────────────────────────
    ws_n = wb.active
    ws_n.title = "README"
    _write_readme(ws_n, 2003, [
        ("Individual Records",
         "Clean typed source. Verify against Team Batting/Pitching sheets before publishing.",
         "NEEDS MANUAL UPDATE"),
        ("Team Highlights",
         "Clean typed source. Values preserved as printed.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "No per-game schedule photo was provided for 2003. Season record (13-8) comes from the stats page header.",
         "NOT PROVIDED"),
        ("Team Batting",
         "Season-to-date stats transcribed (21 games). Some per-player cells had reading challenges — verify top numbers against Records sheet.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "Transcribed. Team totals: 13-8, 3.16 ERA, 111 K.",
         "READABLE"),
        ("Roster",
         "Clean transcription.",
         "READABLE"),
    ])

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Ryan Emmett",         12),
        (2,  "David Hunt",          12),
        (3,  "David Jackson",       12),
        (4,  "Alex Medeiros",       12),
        (5,  "Matt Stock",          12),
        (6,  "Kevin Owens",         11),
        (7,  "Kevin Bogue",         11),
        (8,  "Brandon Shurick",     11),
        (9,  "Jordan Weyhrauch",    11),
        (10, "Jeremy Ellison",      11),
        (11, "Matt Shearer",        10),
        (12, "Matt Pleau",          10),
        (13, "Graham Dorland",      10),
        (14, "Matt Schweitzer",     10),
        (15, "Tyler Rice",          10),
        (16, "Cassidy Emery",       10),
        (17, "Troy Burki",           9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2003 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    bat_rows = [
        ("Bogue",       "93.0",  18,  9,  6, 0, 0, 0,  2,  1, 2,  3, 1, ".333", ".400", ".333"),
        ("Burki",       "85.0",  15,  1,  5, 0, 0, 0,  7,  3, 3,  5, 0, ".333", ".444", ".333"),
        ("Dorland",     "77.0",  31,  7,  5, 2, 0, 0,  3,  6, 1,  5, 1, ".226", ".368", ".161"),
        ("Ellison",     "50.0",   7,  2,  3, 0, 0, 0,  1,  0, 0,  2, 0, ".429", ".500", ".429"),
        ("Emery",        "1.0",   1,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Emmett",     "139.0",  68, 13, 22, 5, 0, 0,  8, 11, 1,  7, 0, ".397", ".395", ".324"),
        ("Hunt",       "100.0",  32,  5, 11, 2, 0, 0, 11,  1, 1, 11, 2, ".406", ".523", ".344"),
        ("Jackson",    "141.0",  82, 24, 31, 5, 0, 0, 11, 12, 1, 15, 0, ".439", ".455", ".378"),
        ("Medeiros",   "141.0",  54, 10, 16, 2, 0, 0,  7, 10, 3, 12, 0, ".333", ".402", ".296"),
        ("Owens",      "141.0",  82, 27, 38,10, 2, 0, 10, 24, 1,  5, 7, ".720", ".580", ".464"),
        ("Pleau",        "6.0",   4,  1,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Rice",         "4.0",   2,  1,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Schweitzer", "108.0",  48, 10, 13, 4, 2, 2,  6,  7, 3,  8, 2, ".563", ".411", ".271"),
        ("Shearer",    "141.0",  59, 14, 20, 2, 0, 0, 11,  2, 1,  8, 1, ".373", ".451", ".339"),
        ("Shurick",     "69.0",  29,  3,  4, 1, 0, 0,  1,  1, 2,  6, 0, ".172", ".241", ".138"),
        ("Stock",      "140.0",  70, 14, 28, 5, 1, 0,  2, 13, 3, 14, 4, ".514", ".461", ".400"),
        ("Weyhrauch",   "22.0",   8,  2,  0, 0, 0, 0,  3,  0, 0,  4, 0, ".000", ".275", ".000"),
        ("TEAM",      "1399.0", 583,136,192,37, 5, 3, 81,112,24, 98,15, ".443", ".443", ".329"),
    ]
    write_table(
        ws2,
        title="2003 Gig Harbor Varsity — Team Batting (Season to Date, 21 games)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Bogue",      "20.1",  114, 32, 22, 16, 19, 0, 2, 18, 1, 1, 0, ".344",  "5.51"),
        ("Burki",       "5.1",   36, 20, 15,  8,  6, 0, 0,  7, 0, 0, 0, ".500",  "10.47"),
        ("Ellison",    "44.0",  183, 26, 17, 10, 16, 0, 7, 23, 4, 2, 1, ".175",  "1.59"),
        ("Emmett",     "21.2",  113, 33, 23, 13, 14, 0, 0, 24, 0, 2, 1, ".280",  "4.20"),
        ("Hunt",        "3.2",   21,  5,  5,  4,  3, 0, 2,  2, 0, 0, 1, ".250",  "7.64"),
        ("Jackson",    "40.0",  175, 35, 20, 11, 16, 0, 2, 35, 6, 2, 0, ".236",  "1.75"),
        ("Owens",       "0.2",    8,  3,  3,  5,  1, 0, 0,  1, 0, 0, 0, ".500", "52.50"),
        ("Shearer",     "1.0",    5,  3,  2,  2,  0, 0, 1,  0, 0, 0, 0, ".667",  "14.00"),
        ("Stock",       "1.0",    0,  0,  0,  0,  0, 0, 0,  0, 0, 0, 0, ".000",  "0.00"),
        ("Weyhrauch",   "2.0",    0,  0,  0,  0,  0, 0, 0,  0, 0, 0, 0, ".000",  "0.00"),
        ("TEAM",      "139.2",  655,147, 95, 63, 78, 0,15, 111,13, 8, 2, ".262",  "3.16"),
    ]
    write_table(
        ws3,
        title="2003 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              "?", "value not clearly visible"),
        ("Most Runs (Inning)",            15,  "vs. Central Kitsap"),
        ("Most Runs Allowed (Game)",      14,  "vs. Wilson"),
        ("Widest Margin of Victory",      14,  "vs. North Kitsap"),
        ("One Run Games Lost",             2,  "vs. Shelton (×2)"),
        ("One Run Games Won",              4,  "vs. Port Angeles (×2), Bremerton, C. Kitsap"),
        ("Most K's (by GH Pitchers)",     11,  "vs. Shelton"),
        ("Most K's (by Opponents)",       12,  "vs. South Kitsap"),
        ("Most Hits (Game)",              19,  "vs. Central Kitsap"),
        ("Most Hits Allowed (Game)",      14,  "vs. Foss"),
        ("Most Walks (Game for GH)",      12,  "vs. Eatonville"),
        ("Most Walks Allowed",             7,  "vs. South Kitsap"),
        ("Most Singles (Game)",           15,  "vs. Lincoln"),
        ("Most Doubles (Game)",            5,  "vs. Port Angeles"),
        ("Most Triples (Game)",            2,  "vs. Foss, NK, CK (×2)"),
        ("Most Home Runs (Game)",          1,  "vs. Shelton"),
        ("Most Steals (Game)",             5,  "vs. Shelton"),
        ("Longest Winning Streak",         4,  ""),
        ("Longest Losing Streak",          2,  ""),
        ("Total Team Runs Scored",       136,  ""),
        ("Total Team Hits",              192,  ""),
        ("Total Team Extra-Base Hits",    49,  ""),
        ("Total Team Home Runs",           5,  ""),
        ("Total Team Stolen Bases",       34,  ""),
        ("Total Team Walks",              81,  ""),
        ("Total Team HBP",                15,  ""),
    ]
    write_table(
        ws_h,
        title="2003 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Kevin Owens",                      ".464", "30+ AB (junior)"),
        ("Most At Bats",           "David Jackson",                     82,    ""),
        ("Most HBP",               "Matt Stock",                         4,    ""),
        ("Lowest K Ratio",         "Kevin Owens",                      "6.1%", "(5-82)"),
        ("Most Doubles",           "Kevin Owens",                       10,    ""),
        ("Most Triples",           "Kevin Owens, Matt Schweitzer",       2,    ""),
        ("Most Home Runs",         "Matt Schweitzer",                    2,    ""),
        ("Most Walks",             "David Hunt",                        11,    ""),
        ("Most Stolen Bases",      "David Jackson",                     12,    ""),
        ("Most Total Bases",       "Kevin Owens",                       59,    ""),
        ("Most RBIs",              "Kevin Owens",                       24,    ""),
        ("Most Wins Pitching",     "David Jackson",                      6,    ""),
        ("Most Innings Pitched",   "Jeremy Ellison",                    44,    ""),
        ("Most K's",               "David Jackson",                     35,    ""),
        ("Lowest ERA",             "Jeremy Ellison",                  "1.59",  ""),
        ("Most Saves",             "David Hunt, Jeremy Ellison",         1,    ""),
        ("Best On-Base Avg.",      "Kevin Owens",                     ".580",  "30+ AB"),
        ("Longest Hitting Streak", "Kevin Owens",                       11,    ""),
        ("Most Runs Scored",       "David Jackson",                     24,    ""),
    ]
    write_table(
        ws4,
        title="2003 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 38, 12, 32],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
