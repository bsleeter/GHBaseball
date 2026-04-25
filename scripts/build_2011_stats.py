#!/usr/bin/env python3
"""Build 2011 Gig Harbor Varsity season stats xlsx from pages in
Historical/2011/. Per-player batting has some illegible cells — team
totals and Individual Records are authoritative."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2011" / "2011_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "Roster"
    roster_rows = [
        (1,  "Parker Guinn",         12),
        (2,  "Spencer Manjarrez",    12),
        (3,  "Michael McCall",       12),
        (4,  "Justin Bonnell",       12),
        (5,  "Eric Johnson",         12),
        (6,  "Colin Walters",        12),
        (7,  "Jordan Pearson",       12),
        (8,  "Eric Flint",           12),
        (9,  "Lucas Marshall",       12),
        (10, "Ryan Olson",           12),
        (11, "Daniel Hendrickson",   12),
        (12, "Ryan Nelson",          12),
        (13, "Ryan Smith",           12),
        (14, "Steven Daily",         11),
        (15, "Jake Ayers",           11),
        (16, "Kyle O'Leary",         11),
        (17, "Austin Eibel",         11),
        (18, "Garrett Gallinger",    10),
        (19, "Nick Gagliardi",       10),
        (20, "Zack Fick",            10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2011 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Scott Hatteberg", "Jim Peschek"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "BB", "R", "H", "1B", "2B", "3B", "HR", "RBI",
        "AVG", "OBP", "SLG", "G", "SO", "HBP", "SAC-B", "SAC-F", "TB",
    ]
    # Per-player rows refined from a careful re-read of the source photo.
    # For the top hitters, breakdowns are reconciled with AVG, SLG, and TB.
    # Team totals and Individual Records remain authoritative.
    bat_rows = [
        ("Guinn",        77, 21, 33, 34, 20, 9, 1, 4, 28, ".442", ".566", ".740", 24,  2, 4, 3, 0, 57),
        ("Bonnell",      84, 11, 26, 33, 22, 8, 3, 0, 15, ".393", ".465", ".560", 24,  5, 0, 1, 1, 47),
        ("Manjarrez",    80, 16, 26, 27, 13, 5, 1, 5, 21, ".338", ".465", ".575", 24, 11, 2, 0, 1, 46),
        ("Nelson",       80,  2, 13, 21, 19, 2, 0, 0,  8, ".263", ".293", ".313", 24, 19, 3, 1, 1, 25),
        ("Gallinger",    81, 14, 23, 27, 19, 8, 0, 0, 12, ".333", ".433", ".432", 22, 19, 2, 0, 1, 35),
        ("Eibel",        59,  4, 14, 15, 10, 4, 1, 0,  9, ".254", ".314", ".390", 20, 12, 1, 2, 0, 23),
        ("Ayers",        55, 18, 18, 18, 13, 3, 1, 1, 11, ".327", ".474", ".473", 23,  7, 2, 0, 0, 26),
        ("Marshall",     50, 18, 12, 15,  9, 2, 4, 0, 17, ".300", ".388", ".500", 24,  5, 4, 0, 0, 25),
        ("Walters",      38, 12, 10,  6,  6, 0, 0, 0,  0, ".158", ".283", ".158", 16, 12, 0, 0, 0,  6),
        ("Pearson",      22,  6,  5,  8,  8, 0, 0, 0,  6, ".364", ".464", ".364", 19,  5, 3, 0, 0,  8),
        ("Hendrickson",   5,  3,  1,  1,  1, 0, 0, 0,  0, ".200", ".500", ".200", 11,  0, 1, 0, 0,  1),
        ("Smith",        55,  7, 14, 20, 16, 4, 0, 0,  7, ".364", ".432", ".436", 22, 10, 0, 0, 0, 24),
        ("Johnson",       5,  2,  2,  2,  2, 0, 0, 0,  1, ".400", ".400", ".400", 22,  0, 0, 0, 0,  2),
        ("McCall",        2,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 17,  2, 0, 0, 0,  0),
        ("Flint",         1,  0,  1,  0,  0, 0, 0, 0,  0, ".000","1.000", ".000", 12,  0, 0, 0, 0,  0),
        ("O'Leary",       2,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 11,  0, 0, 0, 0,  0),
        ("Daily",         0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  3,  0, 0, 0, 0,  0),
        ("Gagliardi",     0,  0,  1,  0,  0, 0, 0, 0,  0, ".000","1.000", ".000",  9,  0, 0, 0, 0,  0),
        ("Olson",         0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000", 14,  1, 0, 0, 0,  0),
        ("Alexander",     1,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  1,  0, 0, 0, 0,  0),
        ("Scanlan",       1,  3,  3,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  3,  0, 0, 0, 0,  0),
        ("Fick",          3,  3,  3,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  3,  0, 0, 0, 0,  0),
        ("TEAM",        696,150,208,237,165,51, 8,13,177, ".341", ".461", ".493", 24, 129, 14, 5, 10, 343),
    ]
    write_table(
        ws2,
        title="2011 Gig Harbor Varsity — Team Batting",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Re-examined from the source photo. Top hitters' breakdowns "
            "(Guinn, Bonnell, Manjarrez, Gallinger) reconciled with their AVG, "
            "SLG, and TB. Caveat: the team batting page reflects an in-season "
            "snapshot — Guinn's records-page AB count is 106 vs. the 77 shown "
            "here, suggesting end-of-season totals were higher for some players."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "#BF", "RS", "ER", "BB", "H", "HR", "SACF", "SACB",
        "HBP", "W", "L", "SV", "BAA", "ERA",
    ]
    pit_rows = [
        ("Johnson",     "4.00",  21,  4,  4,  2,  5, 0, 0, 0, 0, 0, 0, 0, ".263",  "7.00"),
        ("McCall",     "54.00", 230, 19, 13, 27, 25, 0, 0, 0, 0, 7, 2, 0, ".125",  "1.69"),
        ("Flint",      "10.33",  39,  8,  6, 15,  9, 0, 0, 0, 0, 1, 2, 0, ".323",  "5.67"),
        ("O'Leary",    "21.00", 105, 23, 17, 11, 30, 0, 0, 0, 0, 2, 3, 0, ".254",  "6.86"),
        ("Daily",      "16.33",  92, 22, 16, 18, 20, 0, 0, 0, 0, 0, 3, 0, ".213",  "6.86"),
        ("Gagliardi",  "22.67", 117, 25, 17, 12, 22, 0, 0, 0, 0, 1, 1, 0, ".214",  "5.25"),
        ("Olson",      "22.67", 108, 14,  6, 10, 27, 0, 0, 0, 0, 2, 2, 2, ".250",  "1.85"),
        ("Alexander",   "1.00",  12,  7,  6,  2,  7, 0, 0, 0, 0, 0, 0, 0, ".375",  "42.00"),
        ("Scanlan",     "3.00",  22,  6,  2,  5,  6, 0, 0, 0, 0, 0, 0, 0, ".125",  "4.67"),
        ("Fick",        "1.00",  12,  7,  7,  4,  7, 0, 0, 0, 0, 0, 0, 0, ".500",  "49.00"),
        ("TEAM",      "156.00", 788,135, 94,106,150, 0, 0, 1,14, 16, 8, 4, ".225",  "4.22"),
    ]
    write_table(
        ws3,
        title="2011 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Parker Guinn",       ".442", "30+ AB (34-77)"),
        ("Most Plate Appearances", "Parker Guinn",        106,   "originally labeled 'Most At Bats' but is PA"),
        ("Most HBP",               "Lucas Marshall",        4,   ""),
        ("Lowest K Ratio",         "Parker Guinn",       "1.9%", "(2 K / 106 PA)"),
        ("Most Doubles",           "Parker Guinn",          9,   ""),
        ("Most Triples",           "Justin Bonnell",        3,   ""),
        ("Most Home Runs",         "Spencer Manjarrez",     5,   ""),
        ("Most Walks",             "Parker Guinn",         21,   ""),
        ("Most Stolen Bases",      "Spencer Manjarrez",    16,   ""),
        ("Most Total Bases",       "Parker Guinn",         57,   ""),
        ("Most RBIs",              "Parker Guinn",         28,   ""),
        ("Best On-Base Avg.",      "Parker Guinn",       ".566", "30+ AB"),
        ("Longest Hitting Streak", "Ryan Nelson",          11,   ""),
        ("Most Runs Scored",       "Parker Guinn",         33,   ""),
        ("Most Wins Pitching",     "Michael McCall",        7,   ""),
        ("Most Innings Pitched",   "Michael McCall",       54,   ""),
        ("Most K's",               "Michael McCall",       73,   ""),
        ("Lowest ERA",             "Michael McCall",     "1.69", ""),
        ("Most Saves",             "Ryan Olson",            2,   ""),
    ]
    write_table(
        ws4,
        title="2011 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 28, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
