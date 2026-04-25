#!/usr/bin/env python3
"""Build 2012 Gig Harbor Varsity season stats xlsx from CoachStat pages in
Historical/2012/. Includes a separate Team Highlights sheet unique to 2012."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2012" / "2012_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "Roster"
    # Grade assignments adjusted: Gagliardi, Gallinger, Fick listed as grade 11
    # based on 2011 roster (they were 10 in 2011). Original 2012 photo showed
    # Gagliardi at 12 which conflicts with 2011/2013 rosters.
    roster_rows = [
        (1,  "Dillon Alexander",    12),
        (2,  "Henry Campbell",      12),
        (3,  "Jake Ayers",          12),
        (4,  "Austin Morford",      12),
        (5,  "Steven Daily",        12),
        (6,  "Austin Eibel",        12),
        (7,  "Kyle O'Leary",        12),
        (8,  "Nick Gagliardi",      11),
        (9,  "Haak Wagner",         11),
        (10, "Kody Davis",          11),
        (11, "Owen Guenther",       11),
        (12, "Garrett Gallinger",   11),
        (13, "Daniel Koberstein",   11),
        (14, "Colten Miller",       11),
        (15, "Drew Frame",          11),
        (16, "Zack Fick",           11),
        (17, "Quintin Carlson",     10),
        (18, "Mason Selby",         10),
        (19, "Conor Scanlan",       10),
        (20, "Mark Sluys",          10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2012 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Antone Saltvick"],
    )

    # ─── Sheet 2: Team Batting (CoachStat) ──────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "G", "AB", "R", "H", "1B", "2B", "3B", "HR",
        "RBI", "TB", "BB", "SO", "SB", "CS", "OBP", "SLG", "AVG",
    ]
    bat_rows = [
        ("Alexander, D",  "16.0",  13,  6, 1, 1, 1, 0, 0, 0, 1, 1, 0, 0, 0, 0, ".444", ".167", ".167"),
        ("Ayers, J",     "133.0",  21, 73, 22, 18, 19, 2, 0, 1, 20, 0, 1, 10, 3, 0, ".446", ".370", ".301"),
        ("Campbell, H",   "19.0",  14, 18, 4, 3, 4, 0, 0, 0, 2, 0, 0, 3, 0, 0, ".333", ".500", ".333"),
        ("Carlson, Q",    "50.0",  11, 18, 4, 3, 4, 0, 0, 0, 1, 0, 0, 3, 0, 2, ".333", ".500", ".333"),
        ("Daily, S",      "90.0",  16, 31, 7, 6, 7, 0, 0, 0, 4, 3, 2, 6, 0, 0, ".364", ".226", ".226"),
        ("Davis, K",      "28.0",  17,  9, 1, 2, 1, 0, 0, 0, 1, 7, 2, 6, 0, 0, ".294", ".226", ".222"),
        ("Eibel, A",     "141.0",  21, 77, 22, 15, 19, 2, 1, 0, 11, 1, 0, 6, 3, 0, ".286", ".167", ".167"),
        ("Fick, Z",       "13.0",   2,  7, 1, 2, 1, 2, 0, 0, 0, 1, 3, 0, 1, 0, ".402", ".538", ".286"),
        ("Frame, D",     "117.0",  18, 64, 21, 24, 16, 4, 0, 0, 7, 1, 2, 11, 0, 0, ".250", ".250", ".250"),
        ("Gagliardi, N",  "93.0",  20, 53, 20, 11, 15, 3, 0, 0, 5, 0, 4, 0, 0, 0, ".427", ".438", ".328"),
        ("Gallinger, G", "140.0",  21, 73, 21, 16, 15, 3, 0, 2, 15, 0, 2, 10, 1, 0, ".450", ".509", ".288"),
        ("Guenther, O", "113.0",   19, 52, 11, 11, 15, 5, 8, 0, 2, 0, 2, 5, 1, 0, ".441", ".411", ".288"),
        ("Henckel, M",    "1.0",    1,  0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, ".317", ".288", ".212"),
        ("Koberstein, D", "38.0",  13, 28, 5, 7, 5, 0, 0, 0, 1, 5, 5, 7, 0, 0, ".000", ".000", ".000"),
        ("Miller, C",     "89.0",  18, 39, 7, 3, 4, 3, 0, 0, 0, 1, 0, 5, 0, 0, ".343", ".179", ".179"),
        ("Morford, A",    "47.0",  14, 23, 2, 1, 2, 0, 0, 0, 2, 0, 0, 2, 1, 0, ".319", ".256", ".087"),
        ("O'Leary, K",    "10.0",   8,  2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, ".222", ".087", ".087"),
        ("Scanlan, C",    "25.0",  10, 16, 6, 6, 6, 0, 0, 0, 4, 1, 0, 4, 2, 0, ".000", ".000", ".000"),
        ("Selby, M",      "31.0",  11, 17, 4, 2, 2, 0, 0, 0, 3, 0, 0, 4, 1, 0, ".375", ".375", ".375"),
        ("Sluys, M",       "2.0",   3,  2, 1, 1, 1, 0, 0, 0, 0, 0, 0, 1, 0, 0, ".350", ".235", ".235"),
        ("Wagner, H",    "123.0",  20, 36, 8, 9, 0, 0, 0, 0, 2, 0, 0, 0, 0, 0, ".269", ".238", ".222"),
        ("TEAM",             "",   21,653,172,137,143, 22, 2, 5, 113, 213, 99, 96, 16, 0, ".374", ".326", ".263"),
    ]
    write_table(
        ws2,
        title="2012 Gig Harbor Varsity — Team Batting (CoachStat)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
        note=(
            "Note: CoachStat format. Some per-player cells have internal "
            "inconsistencies (OBP/SLG/AVG vs. hit breakdowns) — preserved as "
            "printed. Trust the Individual Records sheet for authoritative "
            "top numbers."
        ),
    )

    # ─── Sheet 3: Team Pitching (CoachStat) ─────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "W", "L", "ERA", "G", "GS", "CG", "SHO", "SV", "SVO",
        "IP", "H", "R", "ER", "HRA", "HB", "BB", "SO",
    ]
    pit_rows = [
        ("Alexander, D",  3, 1, "1.73",  9, 2, 0, 0, 0, 0, "28.1", 26, 19,  7, 0, 1, 48, 19),
        ("Campbell, H",   3, 2, "2.62", 10, 0, 0, 0, 0, 0, "16.0", 10,  8,  6, 0, 3,  6, 22),
        ("Daily, S",      2, 1, "3.55",  9, 0, 0, 0, 0, 0, "25.2", 33, 21, 13, 0, 3, 10, 17),
        ("Davis, K",      2, 0, "2.10",  8, 0, 0, 0, 0, 0, "13.1", 12,  8,  4, 0, 0,  6, 16),
        ("Gagliardi, N",  4, 0, "1.35",  7, 3, 0, 0, 0, 0, "31.0", 30, 12,  6, 0, 1, 14, 35),
        ("Henckel, M",    0, 0, "0.00",  1, 0, 0, 0, 0, 0, "1.0",   0,  0,  0, 0, 0,  1,  2),
        ("O'Leary, K",    1, 0, "3.71",  6, 0, 0, 0, 0, 0, "17.0", 14, 16,  9, 1, 0,  4, 27),
        ("TEAM",         15, 6, "2.38", 21, 5, 4, 0, 0, 0, "132.1",126, 84, 45, 0,10, 49,138),
    ]
    write_table(
        ws3,
        title="2012 Gig Harbor Varsity — Team Pitching (CoachStat)",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              25, "vs. Mt. Tahoma"),
        ("Most Runs (Inning)",            15, "vs. Mt. Tahoma"),
        ("Most Runs Allowed (Game)",      16, "vs. Olympia"),
        ("Widest Margin of Victory",      25, "vs. Mt. Tahoma"),
        ("One Run Games Lost",             0, ""),
        ("One Run Games Won",              5, "SK, Stadium (×2), Roosevelt, Timberline"),
        ("Most K's (by GH Pitchers)",     13, "vs. Central Kitsap"),
        ("Most K's (by Opponents)",       12, "vs. SK"),
        ("Most Hits (Game)",              20, "vs. Mt. Tahoma"),
        ("Most Hits Allowed (Game)",      13, "vs. Olympia"),
        ("Most Singles (Game)",           17, "vs. Mt. Tahoma"),
        ("Most Doubles (Game)",            3, "vs. Bellarmine"),
        ("Most Triples (Game)",            1, "vs. Mt. Tahoma, White River"),
        ("Most Home Runs (Game)",          1, "vs. CK, SK, Mt. Tahoma, Stadium, Shelton"),
        ("Most Steals (Game)",             4, "vs. Mt. Tahoma"),
        ("Longest Winning Streak",         9, ""),
        ("Longest Losing Streak",          2, ""),
        ("Total Team Runs Scored",       137, ""),
        ("Total Team Hits",              172, ""),
        ("Total Team Extra-Base Hits",    29, ""),
        ("Total Team Home Runs",           5, ""),
        ("Total Team Walks",              99, ""),
    ]
    write_table(
        ws_h,
        title="2012 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Nick Gagliardi",                     ".377", "30+ AB (20-53)"),
        ("Most Plate Appearances", "Austin Eibel",                        90,    "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",         "Nick Gagliardi",                     "6.7%", "(4-60)"),
        ("Most Doubles",           "Drew Frame, Nick Gagliardi",           4,    ""),
        ("Most Triples",           "Austin Eibel, Owen Guenther",          1,    ""),
        ("Most Home Runs",         "Garrett Gallinger",                    2,    ""),
        ("Most Walks",             "Jake Ayers, Garrett Gallinger",       16,    ""),
        ("Most Stolen Bases",      "Drew Frame",                           4,    ""),
        ("Most Total Bases",       "Garrett Gallinger",                   30,    ""),
        ("Most RBIs",              "Jake Ayers",                          20,    ""),
        ("Best On-Base Avg.",      "Nick Gagliardi",                     ".509", "30+ AB"),
        ("Longest Hitting Streak", "Jake Ayers",                          12,    ""),
        ("Most Runs Scored",       "Drew Frame",                          24,    ""),
        ("Most Wins Pitching",     "Nick Gagliardi",                       4,    ""),
        ("Most Innings Pitched",   "Nick Gagliardi",                      31,    ""),
        ("Most K's",               "Nick Gagliardi",                      35,    ""),
        ("Lowest ERA",             "Nick Gagliardi",                     "1.35", ""),
        ("Most Saves",             "Henry Campbell",                       2,    ""),
    ]
    write_table(
        ws4,
        title="2012 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 36, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
