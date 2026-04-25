#!/usr/bin/env python3
"""Build 1991 Gig Harbor Varsity season stats xlsx."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1991" / "1991_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1991, [
        ("Individual Records", "Clean typed source labeled '1991 Records'.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source labeled 'Highlights of 1991 Team'.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No schedule photo provided.", "NOT PROVIDED"),
        ("Team Batting", "Per-player Overall Stats transcribed. Some TB cells in source don't sum cleanly to 1B/2B/3B/HR breakdown — preserved as printed.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Per-player Overall Pitching Stats transcribed.", "READABLE"),
        ("Roster", "Clean transcription (some entries had hand-drawn circles next to numbers — ignored).", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Greg Bleistein",      11),
        (2,  "Travis Brown",        12),
        (3,  "Andy Cherbas",         9),
        (4,  "Scott Crawford",      11),
        (5,  "Jason Dupuis",        12),
        (6,  "Scott Harter",        11),
        (7,  "Chad Hoskins",        10),
        (8,  "Billy Kirk",          11),
        (9,  "Lonnie Ledbetter",    12),
        (10, "Jim Peschek",         11),
        (11, "Sig Siegmund",        11),
        (12, "Ryan Sturn",          11),
        (13, "Colin Wark",          12),
        (14, "Ed Wolfe",            11),
        (15, "Isaac Wong",          11),
    ]
    write_roster_sheet(ws_r, "1991 Gig Harbor Varsity — Roster", roster_rows,
                       coaches=["Peter Jansen", "Terry Teeple"],
                       manager="Carl Anderson (Batboy: Jon Erickson)")

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB", "HBP",
        "K", "SB", "SAC", "TB", "E", "OBP", "AVG", "GP",
    ]
    bat_rows = [
        ("Bleistein",  22,  2,  8,  6, 1, 0, 0,  6, 0, 27, 1, 0,  8, 1, ".364", ".364", 10),
        ("Brown",      35,  8,  8,  7, 1, 0, 0,  7, 1,  7,16, 3,  1,18, ".372", ".229", 17),
        ("Cherbas",    37,  7, 11,  6, 3, 0, 0,  5, 0,  8,19, 6,  1,17, ".381", ".297", 16),
        ("Crawford",   49,  7, 17,  7, 6, 0, 0,  4, 2,  4, 1, 0, 25, 2, ".396", ".347", 17),
        ("Dupuis",     58, 15, 17,  6, 1, 0, 0,  7, 4,  8,12, 6,  0,27, ".369", ".293", 17),
        ("Harter",     24,  3,  5,  2, 0, 0, 0,  4, 1,  7,24, 0, 12, 3, ".345", ".208", 18),
        ("Hoskins",    31,  6, 14,  8, 2, 1, 0, 10, 2,  5,12, 3,  0,31,  ".605", ".452", 16),
        ("Kirk",       51, 18, 19,  9, 4, 1, 0,  8, 2, 11,17, 3,  3,34,  ".475", ".373", 18),
        ("Ledbetter",  45, 14, 12,  7, 1, 2, 0,  8, 0,  5, 9, 1,  2,25,  ".364", ".267", 17),
        ("Peschek",    51, 13, 17, 17, 3, 1, 1,  9, 2, 12,20, 1,  0,36,  ".452", ".333", 17),
        ("Siegmund",   61, 12, 19,  7, 5, 0, 0,  4, 2,  8,12, 2,  1,32,  ".373", ".311", 17),
        ("Sturn",      33,  7, 11,  5, 1, 1, 0,  6, 0, 18, 4, 0,  0,14,  ".353", ".333", 14),
        ("Wark",        3,  0,  0,  0, 0, 0, 0,  1, 0,  0,33, 1,  0, 0,  ".000", ".000",  6),
        ("Wong",       19,  2,  5,  7, 2, 0, 0,  2, 0,  2, 9, 1,  1, 9,  ".333", ".263", 12),
        ("Wolfe",       2,  1,  1,  0, 1, 0, 0,  0, 0,  0, 0, 0,  0, 3,  ".500", ".500",  1),
    ]
    write_table(
        ws2,
        title="1991 Gig Harbor Varsity — Overall Batting Statistics",
        headers=bat_headers,
        rows=bat_rows,
        col_widths=[14] + [6] * (len(bat_headers) - 2) + [6],
        left_align_cols={1},
        note=(
            "Note: Some TB column values in the source don't sum cleanly to "
            "the visible 1B/2B/3B/HR breakdown — preserved as printed. "
            "Some narrow columns (HBP, SB, SAC, GP) had reading challenges "
            "for a few players; values are best-effort."
        ),
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = ["Player", "W", "L", "IP", "R", "ER", "H", "BB", "K", "HBP", "ERA", "CG", "G"]
    pit_rows = [
        ("Crawford",  2, 4, "37.0",  27, 18, 32, 25, 42, 3, "3.41", 3, 10),
        ("Cherbas",   1, 2, "12.1",  17, 22, 17,  8, 13, 0, "6.24", 0,  7),
        ("Peschek",   0, 0, "1.1",    1,  1,  3,  2,  0, 0, "5.25", 0,  1),
        ("Siegmund",  6, 2, "59.2",  41, 19, 69, 13, 27, 2, "2.23", 6,  5),
        ("Wark",      0, 1, "6.1",   12,  6,  8,  6,  3, 1, "6.63", 0,  1),
        ("Wolfe",     0, 0, "2.0",    0,  0,  3,  1,  4, 0, "0.00", 0,  1),
        ("Wong",      0, 0, "1.1",    5,  3,  4,  2,  1, 0, "15.75",0,  1),
    ]
    write_table(
        ws3,
        title="1991 Gig Harbor Varsity — Overall Pitching Statistics",
        headers=pit_headers,
        rows=pit_rows,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    write_table(
        ws_h,
        title="Highlights of 1991 Team",
        headers=["Highlight", "Value", "Context"],
        rows=[
            ("Most Runs (Game)",              14, "vs. F.P."),
            ("Most Runs (Inning)",             6, "vs. Steilacoom, Enumclaw"),
            ("Most Runs Allowed (Game)",      19, "vs. Enumclaw"),
            ("Widest Margin of Victory",      11, "vs. F.P."),
            ("Widest Margin of Defeat",       12, "vs. White River"),
            ("One Run Games Won",              1, "vs. Fife"),
            ("One Run Games Lost",             1, "vs. Peninsula"),
            ("Most K's (by GH Pitchers)",     11, "vs. Tahoma"),
            ("Most K's (by Opponents)",       11, "vs. Peninsula"),
            ("Most Hits (Game)",              17, "vs. F.P."),
            ("Most Hits Allowed (Game)",      19, "vs. Enumclaw"),
            ("Most Walks (Game for GH)",      11, "vs. Enumclaw"),
            ("Most Walks Allowed",             7, "vs. Enumclaw"),
            ("Most Singles (Game)",           12, "vs. Enumclaw, F.P."),
            ("Most Doubles (Game)",            4, "vs. White River, F.P."),
            ("Most Triples (Game)",            2, "vs. Yelm"),
            ("Most Home Runs (Game)",          1, "vs. F.P."),
            ("Most Steals (Game)",             7, "vs. Steilacoom"),
            ("Most Errors (Game)",             6, "vs. Enumclaw, White River"),
        ],
        col_widths=[34, 10, 48], left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    write_table(ws4, "1991 Records",
        ["Record", "Holder", "Value", "Qualifier / Note"],
        [
            ("Highest Average",        "Chad Hoskins",                 ".452", ""),
            ("Most Hits",              "Sig Siegmund, Billy Kirk",      19,    ""),
            ("Most At Bats",           "Sig Siegmund",                  61,    ""),
            ("Most Runs Scored",       "Billy Kirk",                    18,    ""),
            ("Most HBP",               "Jason Dupuis",                   4,    ""),
            ("Most Doubles",           "Scott Crawford",                 6,    ""),
            ("Most Triples",           "Lonnie Ledbetter",               2,    ""),
            ("Most Home Runs",         "Jim Peschek",                    1,    ""),
            ("Most Walks",             "Chad Hoskins",                  10,    ""),
            ("Most Stolen Bases",      "Andy Cherbas, Jason Dupuis",     6,    ""),
            ("Most Total Bases",       "Jim Peschek",                   36,    ""),
            ("Most RBIs",              "Jim Peschek",                   17,    ""),
            ("Most Wins Pitching",     "Sig Siegmund",                   6,    ""),
            ("Most Innings Pitched",   "Sig Siegmund",                "59.2",  ""),
            ("Most K's",               "Scott Crawford",                42,    ""),
            ("Lowest ERA",             "Sig Siegmund",                "2.23",  ""),
            ("Best On-Base Average",   "Chad Hoskins",                ".605",  ""),
            ("Lowest K Ratio (Bat)",   "Scott Crawford",              ".037",  "(1 K / 27 AB)"),
        ],
        col_widths=[26, 36, 12, 22], left_align_cols={2, 4})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
