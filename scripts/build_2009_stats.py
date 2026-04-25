#!/usr/bin/env python3
"""Build 2009 Gig Harbor Varsity season stats xlsx from pages in
Historical/2009/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2009" / "2009_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "Roster"
    roster_rows = [
        (1,  "Chet Thompson",        12),
        (2,  "Mike Barnett",         12),
        (3,  "Cameron Holcomb",      12),
        (4,  "Nico Youngren",        12),
        (5,  "Scott Benedict",       12),
        (6,  "Mike Jones",           12),
        (7,  "Brian Bullatt",        12),
        (8,  "David Bigelow",        11),
        (9,  "Scott Schultz",        11),
        (10, "Carl Beck",            11),
        (11, "Michael McCall",       10),
        (12, "Parker Guinn",         10),
        (13, "Colin Walters",        10),
        (14, "Jordan Pearson",       10),
        (15, "Spencer Manjarrez",    10),
        (16, "Lucas Marshall",       10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2009 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "BB", "R", "H", "1B", "2B", "3B", "HR", "RBI",
        "AVG", "SLG", "OBP", "HBP", "SAC-B", "SAC-F", "SO", "TB",
    ]
    bat_rows = [
        ("Youngren",   53, 21, 25, 17, 13, 1, 3, 0, 12, ".321", ".453", ".514",  0, 0, 0, 12, 24),
        ("Barnett",    59, 15, 18, 23, 17, 2, 1, 0, 15, ".390", ".525", ".520",  1, 0, 0,  7, 31),
        ("Thompson",   56, 12, 17, 23, 17, 2, 3, 0, 16, ".411", ".607", ".514",  1, 0, 0,  7, 34),
        ("Schultz",    52, 12, 15, 17, 10, 6, 0, 1, 14, ".327", ".500", ".478",  3, 0, 0,  7, 26),
        ("Bigelow",    63,  7, 15, 17,  9, 5, 1, 2, 17, ".270", ".476", ".338",  1, 0, 0,  8, 30),
        ("Benedict",   51,  9, 12, 17, 13, 4, 0, 0,  9, ".333", ".451", ".462",  4, 0, 0,  3, 23),
        ("Guinn",      21,  2,  4,  7,  6, 1, 0, 0,  4, ".333", ".381", ".400",  1, 0, 1,  2,  8),
        ("Beck",       10,  3,  4,  1,  1, 0, 0, 0,  0, ".100", ".100", ".300",  1, 0, 0,  2,  1),
        ("Walters",    12,  0,  6,  5,  4, 1, 0, 0,  0, ".417", ".500", ".417",  0, 0, 0,  2,  6),
        ("Bullatt",     5,  0,  2,  2,  1, 0, 1, 0,  2, ".400", ".800", ".400",  0, 0, 1,  1,  4),
        ("Holcomb",    52,  3,  8, 10,  8, 0, 0, 1,  2, ".192", ".231", ".250",  1, 0, 0, 14, 12),
        ("McCall",      8,  2,  3,  4,  2, 1, 0, 1,  3, ".500","1.000", ".600",  0, 0, 0,  2,  8),
        ("Manjarrez",  46,  7, 14, 13,  9, 3, 0, 1, 11, ".283", ".413", ".414",  4, 0, 0, 13, 19),
        ("Pearson",    14,  3,  7,  5,  4, 1, 0, 0,  3, ".357", ".429", ".500",  1, 0, 0,  2,  6),
        ("Flint",       0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0, 0, 0,  0,  0),
        ("Jones",      30,  3, 10,  7,  4, 2, 0, 1,  3, ".233", ".433", ".333",  3, 0, 0, 11, 13),
        ("Marshall",    1,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0, 0, 0,  0,  0),
        ("TEAM",      533, 99,150,168,119, 34, 7, 8,137, ".315", ".450", ".438", 22, 0, 6, 92,240),
    ]
    write_table(
        ws2,
        title="2009 Gig Harbor Varsity — Team Batting",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Per-player batting transcribed from CoachStat report. Some "
            "hit breakdowns don't sum cleanly to published AVG/TB — preserved "
            "as printed. Trust Individual Records sheet for authoritative top "
            "numbers."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "#BF", "RS", "ER", "BB", "H", "HBP",
        "W", "L", "SV", "BAA", "ERA", "SO",
    ]
    pit_rows = [
        ("Thompson",   "1.0",    6,  0,  0,  2,  0,  0, 0, 0, 0, ".250",  "0.00",  1),
        ("Schultz",   "47.2",  196, 21, 11, 12, 27,  9, 4, 5, 0, ".154",  "1.62", 56),
        ("Bigelow",   "32.1",  156, 28, 17, 11, 23,  6, 2, 1, 0, ".165",  "3.68", 30),
        ("Benedict",   "1.0",    6,  3,  2,  2,  1,  1, 0, 0, 0, ".333", "14.00",  0),
        ("Bullatt",   "10.1",   58, 13, 13,  8, 11,  1, 0, 1, 1, ".224",  "8.81", 14),
        ("McCall",    "25.2",  114, 17,  7, 14, 20,  2, 4, 1, 0, ".204",  "1.91", 17),
        ("Flint",      "0.1",    7,  4,  4,  1,  4,  1, 0, 1, 0, ".800", "84.00",  0),
        ("Jones",      "9.0",   38,  1,  1,  5,  4,  1, 1, 0, 0, ".125",  "0.78",  7),
        ("TEAM",     "127.1",  581, 87, 55, 55, 91, 21,12, 8, 1, ".180",  "3.02",125),
    ]
    write_table(
        ws3,
        title="2009 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              32, "vs. Mt. Tahoma"),
        ("Most Runs (Inning)",             8, "vs. Mt. Tahoma"),
        ("Most Runs Allowed (Game)",      14, "vs. Olympia"),
        ("Widest Margin of Victory",      30, "vs. Mt. Tahoma"),
        ("One Run Games Lost",             3, "vs. CK, SK, Bellarmine"),
        ("One Run Games Won",              3, "vs. Wilson, Bellarmine, CK"),
        ("Most K's (by GH Pitchers)",     13, "vs. Shelton"),
        ("Most K's (by Opponents)",       15, "vs. CK (9 innings)"),
        ("Most Hits (Game)",              24, "vs. Mt. Tahoma"),
        ("Most Hits Allowed (Game)",      13, "vs. CK"),
        ("Most Walks (Game for GH)",      11, "vs. Stadium"),
        ("Most Walks Allowed",            12, "vs. Shelton"),
        ("Most Singles (Game)",           14, "vs. Mt. Tahoma"),
        ("Most Doubles (Game)",            8, "vs. Mt. Tahoma"),
        ("Most Triples (Game)",            2, "vs. Mt. Tahoma"),
        ("Most Home Runs (Game)",          2, "vs. Foss"),
        ("Most Steals (Game)",             8, "vs. Foss"),
        ("Longest Winning Streak",        10, ""),
        ("Longest Losing Streak",          3, ""),
        ("Total Team Runs Scored",       160, ""),
        ("Total Team Hits",              168, ""),
        ("Total Team Extra-Base Hits",    49, ""),
        ("Total Team Home Runs",           8, ""),
        ("Total Team Stolen Bases",       71, ""),
        ("Total Team Walks",              99, ""),
        ("Total Team HBP",                22, ""),
    ]
    write_table(
        ws_h,
        title="2009 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Chet Thompson",                        ".411", "30+ AB (23-56)"),
        ("Most Plate Appearances", "Mike Barnett",                          75,    "originally labeled 'Most At Bats' but is PA"),
        ("Most HBP",               "Scott Benedict, Spencer Manjarrez",      4,    ""),
        ("Lowest K Ratio",         "Scott Benedict",                       "4.7%", "(3 K / 64 PA)"),
        ("Most Doubles",           "Scott Schultz",                          6,    ""),
        ("Most Triples",           "Nico Youngren, Chet Thompson",           3,    ""),
        ("Most Home Runs",         "David Bigelow",                          2,    ""),
        ("Most Walks",             "Nico Youngren",                         21,    ""),
        ("Most Stolen Bases",      "Nico Youngren",                         19,    ""),
        ("Most Total Bases",       "Chet Thompson",                         34,    ""),
        ("Most RBIs",              "Chet Thompson",                         23,    ""),
        ("Best On-Base Avg.",      "Mike Barnett",                         ".520", "30+ AB"),
        ("Longest Hitting Streak", "Chet Thompson",                          7,    ""),
        ("Most Runs Scored",       "Nico Youngren",                         25,    ""),
        ("Most Wins Pitching",     "Scott Schultz, Michael McCall",          4,    ""),
        ("Most Innings Pitched",   "Scott Schultz",                        "47.2", ""),
        ("Most K's",               "Scott Schultz",                         56,    ""),
        ("Lowest ERA",             "Scott Schultz",                       "1.62",  ""),
        ("Most Saves",             "Brian Bullatt",                          1,    ""),
    ]
    write_table(
        ws4,
        title="2009 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 38, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
