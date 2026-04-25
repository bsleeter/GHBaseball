#!/usr/bin/env python3
"""Build 1994 Gig Harbor Varsity season stats xlsx."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1994" / "1994_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1994, [
        ("Individual Records", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No per-game schedule photo provided. Season record 13-7-2 per pitching W-L.", "NOT PROVIDED"),
        ("Team Batting", "Re-read from clear source photo. Per-player H/AB/AVG values reconciled. Some TB cells in source don't sum perfectly to breakdown — preserved as printed.", "READABLE"),
        ("Team Pitching", "Re-read from clear source photo. Team: 13-7, 2.69 ERA, 114 K.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Joel Miller",        11),
        (2,  "Andy Cherbas",       12),
        (3,  "Ryan Sawyers",       12),
        (4,  "Tom Friedman",       11),
        (5,  "Jason Lippert",      11),
        (6,  "Paul Baurichter",    11),
        (7,  "Derek Vitcovich",    11),
        (8,  "Kevin Graybill",     11),
        (9,  "Christian Lindmark", 12),
        (10, "Kevin Feltus",       11),
        (11, "Rich Langford",      11),
        (12, "Brandon Ratliff",    12),
        (13, "Aaron Ford",         11),
        (14, "Matt Gardner",        9),
        (15, "Aaron Araujo",        9),
    ]
    write_roster_sheet(ws_r, "1994 Gig Harbor Varsity — Roster", roster_rows,
                       coaches=["Peter Jansen", "Bob Maguinez", "Tony Anderson"])

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "G", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB",
        "HBP", "K", "K%", "SAC", "ERR", "TB", "SB", "AVG", "OBP",
    ]
    bat_rows = [
        ("J. Miller",   21, 72, 23, 32, 20, 5, 3, 0, 11, 4,  5,  "6.9", 0, 4, 43, 15, ".444", ".540"),
        ("Cherbas",     20, 67, 27, 26, 19, 3, 1, 4, 11, 4, 11, "14.1", 1, 7, 40,  2, ".388", ".459"),
        ("Lippert",     20, 55, 14, 17, 11,  2, 0, 0, 11, 0,  9, "16.9", 0, 1, 19,  3, ".309", ".405"),
        ("Graybill",    18, 40,  7, 14, 14, 2, 1, 0,  5, 1, 10, "10.4", 0, 1, 19,  3, ".350", ".447"),
        ("Sawyers",     20, 58, 22, 16, 14, 3, 3, 0, 12, 1, 16, "16.9", 0, 0, 22,  0, ".276", ".408"),
        ("Baurichter",  20, 56, 11, 13,  7, 0, 1, 0, 12, 0,  6,  "8.7", 0, 0, 15,  4, ".232", ".388"),
        ("Friedman",    20, 62, 10, 16, 19, 4, 1, 1,  3, 0,  5,  "7.2", 0, 0, 25,  2, ".258", ".364"),
        ("Vitcovich",   19, 44, 14, 14,  7, 3, 0, 0,  3, 0,  2,  "3.7", 0, 1, 22,  9, ".318", ".364"),
        ("Lindmark",    17, 34,  6,  7,  7, 1, 0, 0,  4, 0,  6, "15.4", 0, 1,  8,  0, ".206", ".286"),
        ("Feltus",      15, 25,  9,  8,  5, 2, 0, 0,  4, 0,  7, "20.6", 1, 2, 10,  0, ".320", ".485"),
        ("Langford",    13, 10,  6,  3,  5, 0, 0, 0,  2, 0,  2, "13.3", 1, 2,  4,  0, ".300", ".550"),
        ("Ford",         7,  4,  2,  1,  0, 0, 0, 0,  2, 0,  0,  "0.0", 0, 0,  1,  0, ".250", ".500"),
        ("Ratliff",      7,  8,  2,  2,  0, 0, 0, 0,  2, 0,  2, "22.2", 0, 0,  2,  0, ".250", ".333"),
        ("Gardner",      4,  4,  0,  1,  0, 0, 0, 0,  0, 0,  1,    "—", 0, 0,  1,  0, ".250", ".250"),
        ("Araujo",       2,  5,  1,  1,  0, 0, 0, 0,  0, 0,  1, "20.0", 0, 0,  1,  0, ".200", ".200"),
        ("TEAM",        22,544,154,171,128,25, 9, 5, 81, 6, 81, "14.9", 2,19,232, 38, ".314", ".410"),
    ]
    write_table(
        ws2,
        title="1994 Gig Harbor Varsity — Overall Batting Statistics",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 5] + [6] * (len(bat_headers) - 3) + [6, 7],
        left_align_cols={1},
        note=(
            "Note: Re-read from clear source photo. Per-player H values were "
            "reconciled with AVG and AB. Some TB cells don't sum cleanly to "
            "the visible 1B/2B/3B/HR breakdown — preserved as printed. Team "
            "total row computed from per-player sums (may differ slightly from "
            "source's printed totals)."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF", "R", "H", "K", "BB", "ER",
        "HBP", "G", "ERA", "OPPBA",
    ]
    pit_rows = [
        ("Cherbas",    1, 0, 0,  "9.2",  42, 13, 14,  3,  7,  7, 0, 3,  "5.07", ".333"),
        ("Sawyers",    4, 2, 0, "39.2", 145, 12, 28, 34,  9, 10, 2, 9,  "1.76", ".193"),
        ("Lippert",    1, 5, 0, "26.1",  97, 26, 24, 25, 25, 17, 2, 7,  "4.52", ".247"),
        ("Lindmark",   7, 0, 0, "46.1", 183, 22, 45, 39, 12, 14, 5, 9,  "2.12", ".246"),
        ("Ratliff",    0, 0, 0,  "0.0",   0,  0,  0,  0,  0,  0, 0, 1,  "0.00", ".000"),
        ("Gardner",    0, 0, 0,  "3.1",   9,  3,  2,  0,  1,  0, 0, 2,  "6.31", ".222"),
        ("J. Miller",  0, 0, 1,  "1.0",   4,  0,  0,  1,  0,  0, 0, 1,  "0.00",  "—"),
        ("Araujo",     0, 0, 0,  "3.1",  13,  2,  4,  1,  5,  1, 0, 2,  "2.10", ".308"),
        ("Ford",       0, 0, 0,  "3.1",  15,  6,  5,  0,  6,  5, 0, 1, "10.51", ".333"),
        ("TEAM",      13, 7, 3, "134.0", 511, 84,123,114, 66, 57, 8,20,  "2.69", ".241"),
    ]
    write_table(
        ws3,
        title="1994 Gig Harbor Varsity — Overall Pitching Statistics",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    write_table(ws_h, "1994 Gig Harbor Varsity — Team Highlights",
        ["Highlight", "Value", "Context"],
        [
            ("Most Runs (Game)",              19, "vs. Fife"),
            ("Most Runs (Inning)",             8, "vs. Tahoma"),
            ("Most Runs Allowed (Game)",      12, "vs. Capital"),
            ("Widest Margin of Victory",      17, "vs. F.P."),
            ("One Run Games Won",              1, "vs. Peninsula"),
            ("One Run Games Lost",             2, "vs. Enumclaw, Tahoma"),
            ("Most K's (by GH Pitchers)",     11, "vs. F.P."),
            ("Most K's (by Opponents)",        9, "vs. Enumclaw"),
            ("Most Hits (Game)",              18, "vs. Fife"),
            ("Most Hits Allowed (Game)",      14, "vs. Enumclaw"),
            ("Most Walks (Game for GH)",      15, "vs. F.P."),
            ("Most Walks Allowed",            13, "vs. Capital"),
            ("Most Singles (Game)",           13, "vs. W.R."),
            ("Most Doubles (Game)",            5, "vs. Peninsula"),
            ("Most Triples (Game)",            3, "vs. Fife"),
            ("Most Home Runs (Game)",          4, "vs. Fife, Curtis, Enumclaw, W.R."),
            ("Most Steals (Game)",             6, "vs. Fife, N.M."),
            ("Longest Winning Streak",         4, ""),
            ("Longest Losing Streak",          3, ""),
        ],
        col_widths=[34, 10, 48], left_align_cols={1, 3})

    ws4 = wb.create_sheet("Individual Records")
    write_table(ws4, "1994 Gig Harbor Varsity — Individual Records",
        ["Record", "Holder", "Value", "Qualifier / Note"],
        [
            ("Highest Average",        "Joel Miller",                   ".444", "30+ AB (32-72)"),
            ("Most Hits",              "Joel Miller",                    32,    ""),
            ("Most At Bats",           "Joel Miller",                    72,    ""),
            ("Most Runs Scored",       "Joel Miller",                    23,    ""),
            ("Most HBP",               "Andy Cherbas, Joel Miller",       4,    "tied"),
            ("Lowest K Ratio",         "Derek Vitcovich",              "3.7%",  "(2 K / 44 AB)"),
            ("Most Doubles",           "Joel Miller",                     5,    ""),
            ("Most Triples",           "Joel Miller, Ryan Sawyers",       3,    "tied"),
            ("Most Home Runs",         "Andy Cherbas",                    4,    ""),
            ("Most Walks",             "Paul Baurichter",                12,    ""),
            ("Most Stolen Bases",      "Joel Miller",                    15,    ""),
            ("Most Total Bases",       "Joel Miller",                    43,    ""),
            ("Most RBIs",              "Joel Miller",                    20,    ""),
            ("Most Wins Pitching",     "Christian Lindmark",              7,    "7-0 record"),
            ("Most Innings Pitched",   "Christian Lindmark",           "46.1",  ""),
            ("Most K's",               "Christian Lindmark",             39,    ""),
            ("Lowest ERA",             "Ryan Sawyers",                 "1.76",  ""),
            ("Best On-Base Avg.",      "Joel Miller",                  ".500",  "30+ AB · per records page"),
            ("Longest Hitting Streak", "Joel Miller",                     9,    ""),
        ],
        col_widths=[26, 36, 12, 22], left_align_cols={2, 4})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
