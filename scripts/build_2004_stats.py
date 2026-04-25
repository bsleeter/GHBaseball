#!/usr/bin/env python3
"""Build 2004 Gig Harbor Varsity season stats xlsx from pages in
Historical/2004/. Notable: Nick Mareno's freshman debut (.500 AVG in 54 AB,
15-game hit streak). Kevin Owens / Kevin Bogue senior year."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2004" / "2004_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches, manager=None):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    if manager:
        r = last + 1 + len(coaches) + 1
        ws.cell(row=r, column=1, value=f"MANAGER: {manager}").font = Font(
            name="Arial", size=10, italic=True, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def _write_readme(ws, year, needs):
    ws["A1"] = f"{year} Season Stats — Manual Update Required"
    ws.merge_cells("A1:C1")
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 24
    for c, h in enumerate(("SHEET", "WHAT NEEDS ATTENTION", "STATUS"), start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 22
    for i, (sheet, issue, status) in enumerate(needs):
        for c, val in enumerate((sheet, issue, status), start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            cell.font = Font(name="Arial", size=10, color=NAVY,
                             bold=(c == 3 and status in ("NEEDS MANUAL UPDATE", "NOT PROVIDED")))
            cell.alignment = Alignment(
                horizontal="left" if c != 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = BORDER
            if c == 3 and status == "NEEDS MANUAL UPDATE":
                cell.fill = PatternFill("solid", start_color=FLAG)
            elif c == 3 and status == "NOT PROVIDED":
                cell.fill = PatternFill("solid", start_color="F8D7DA")
            elif i % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT)
        ws.row_dimensions[4 + i].height = 48
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 0: README ────────────────────────────────────────────
    ws_n = wb.active
    ws_n.title = "README"
    _write_readme(ws_n, 2004, [
        ("Individual Records",
         "Re-derived directly from the 2004 Team Batting and Team Pitching sheets. The original records page was hand-annotated with 2005 updates and unreliable. CAVEAT: the batting page is labeled 'Season to Date thru 22 Games' — TB/AB values may have grown by season's end. Header says 12-9 but pitching W-L sums to 11-9.",
         "NEEDS SPOT-CHECK"),
        ("Team Highlights",
         "Image was worn; some context (opponents, streaks) unclear. Values transcribed at best-effort.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "No per-game schedule photo was provided for 2004. Season record (12-9) comes from stats page header.",
         "NOT PROVIDED"),
        ("Team Batting",
         "Season-to-date stats transcribed. Source had some internal inconsistencies (OBP/SLG values don't always match breakdowns) — preserved as printed.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "Transcribed. Team totals 11-9 / 3.58 ERA (header states 12 wins — verify).",
         "NEEDS SPOT-CHECK"),
        ("Roster",
         "Clean transcription. Kevin Owens (senior 2004 player) later became a coach in 2015.",
         "READABLE"),
    ])

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Kevin Owens",        12),
        (2,  "Jay Hemley",          12),
        (3,  "Jeremy Ellison",      12),
        (4,  "Jordan Weyhrauch",    12),
        (5,  "Kevin Bogue",         12),
        (6,  "Graham Dorland",      11),
        (7,  "Matt Shearer",        11),
        (8,  "Matt Pleau",          11),
        (9,  "Tyler Rice",          11),
        (10, "Cassidy Emery",       11),
        (11, "Seth Whiting",        11),
        (12, "Troy Burki",          10),
        (13, "Rob Emmett",          10),
        (14, "David Benedict",      10),
        (15, "Josiah Ward",         10),
        (16, "Casey Knox",          10),
        (17, "Nick Mareno",          9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2004 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
        manager="David Colman",
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    bat_rows = [
        ("Benedict",   "127.0",  38,  5,  8, 0, 0, 0,  7,  6, 1,  9, 1, ".211", ".340", ".211"),
        ("Bogue",       "91.0",  19,  3,  4, 1, 0, 0,  1,  1, 0,  3, 0, ".263", ".375", ".211"),
        ("Burki",       "96.0",  32,  8, 10, 2, 0, 0, 10,  7, 0, 10, 3, ".375", ".511", ".313"),
        ("Dorland",     "80.0",  30, 11,  8, 3, 0, 1, 12, 11, 0,  1, 0, ".467", ".273", ".267"),
        ("Ellison",     "29.0",  11,  2,  3, 0, 0, 0,  0,  0, 0,  5, 0, ".273", ".273", ".273"),
        ("Emery",      "139.0",  75, 21, 35, 4, 0, 3,  6, 24, 3,  9, 0, ".640", ".494", ".467"),
        ("Emmett",      "70.0",  26,  4,  3, 1, 0, 0,  2,  0, 2,  5, 0, ".154", ".261", ".115"),
        ("Hemley",      "43.0",  21,  3,  3, 1, 0, 1,  3,  5, 0,  6, 0, ".333", ".333", ".143"),
        ("Knox",         "4.0",   0,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("MacDonald",    "4.0",   2,  0,  0, 0, 0, 0,  0,  0, 0,  2, 0, ".000", ".000", ".000"),
        ("Mareno",     "108.0",  54, 15, 27, 5, 0, 1,  5, 12, 5,  1, 0, ".648", ".508", ".500"),
        ("Morris",       "2.0",   0,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Owens",      "149.0",  72, 25, 28, 6, 0, 2, 14, 25, 2,  2, 4, ".556", ".511", ".389"),
        ("Pleau",      "136.0",  63, 11, 23, 6, 0, 1, 10, 18, 2,  9, 4, ".508", ".456", ".365"),
        ("Rice",       "116.0",  40, 12, 11, 1, 0, 1,  8,  4, 0,  7, 3, ".375", ".431", ".275"),
        ("Shearer",    "136.0",  65, 17, 23, 7, 0, 2, 12, 20, 0,  9, 1, ".554", ".462", ".354"),
        ("Shurick",     "39.0",  19,  7,  5, 2, 1, 0,  3,  2, 0,  4, 0, ".474", ".364", ".263"),
        ("Ward",        "40.0",  11,  6,  0, 0, 0, 0,  3,  0, 0,  3, 1, ".000", ".267", ".000"),
        ("Weyhrauch",   "67.0",  26, 10,  6, 2, 0, 0,  7,  4, 2,  5, 3, ".308", ".486", ".231"),
        ("Whiting",     "14.0",   4,  0,  1, 0, 0, 0,  0,  1, 0,  1, 2, ".250", ".429", ".250"),
        ("TEAM",      "1490.0", 608,160,198,41, 1,12,103,140,18, 98,20, ".456", ".436", ".326"),
    ]
    write_table(
        ws2,
        title="2004 Gig Harbor Varsity — Team Batting (Season to Date, 22 games)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
        note=(
            "Note: 'Shurick' appears in the stats page but is not on the typed "
            "roster — likely a mid-season addition or call-up. Player is kept "
            "as-printed."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Bogue",     "64.2",  262, 60, 20, 12, 10, 0,  3, 78, 8, 1, 1, ".241",  "1.30"),
        ("Burki",     "26.1",  148, 38, 34, 22, 29, 0,  2, 23, 1, 3, 0, ".325",  "5.85"),
        ("Dorland",    "1.1",   11,  6,  5,  3,  2, 0,  0,  0, 0, 0, 0, ".667", "15.75"),
        ("Ellison",   "12.1",   76, 27, 26, 10,  7, 0,  3,  9, 0, 3, 0, ".409",  "5.68"),
        ("Emmett",     "2.1",   18,  6,  3,  2,  2, 0,  2,  1, 0, 0, 0, ".400",  "6.00"),
        ("Knox",       "4.0",   19,  5,  4,  1,  0, 0,  0,  5, 0, 0, 1, ".263",  "1.75"),
        ("Morris",     "0.2",    2,  1,  1,  1,  0, 0,  0,  0, 0, 1, 0, ".500", "10.50"),
        ("Rice",      "14.0",   63, 16, 15, 10,  4, 0,  1, 14, 1, 0, 0, ".276",  "5.00"),
        ("Whiting",    "9.1",   48, 13,  9,  8,  5, 0,  1,  3, 1, 1, 1, ".310",  "6.00"),
        ("TEAM",     "135.0",  647,172,117, 69, 58, 0, 12,133,11, 9, 3, ".298",  "3.58"),
    ]
    write_table(
        ws3,
        title="2004 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              20, "vs. Bremerton"),
        ("Most Runs (Inning)",             "?", "vs. Shelton"),
        ("Most Runs Allowed (Game)",      14, "vs. Port Angeles"),
        ("Widest Margin of Victory",      20, "vs. Bremerton"),
        ("One Run Games Lost",             2, "vs. Port Angeles, South Kitsap"),
        ("One Run Games Won",              1, ""),
        ("Most K's (by GH Pitchers)",     13, "vs. Olympia"),
        ("Most K's (by Opponents)",        9, "vs. Shelton"),
        ("Most Hits (Game)",              17, "vs. Central Kitsap"),
        ("Most Hits Allowed (Game)",      17, "vs. Wilson"),
        ("Most Walks (Game for GH)",       8, "vs. Eatonville"),
        ("Most Walks Allowed",            10, "vs. South Kitsap"),
        ("Most Singles (Game)",           12, "vs. Central Kitsap"),
        ("Most Doubles (Game)",            5, "vs. Wilson"),
        ("Most Triples (Game)",            2, "vs. Olympic, Bellarmine"),
        ("Most Home Runs (Game)",          3, "vs. Bremerton"),
        ("Most Steals (Game)",             7, ""),
        ("Longest Winning Streak",         2, "verify"),
        ("Longest Losing Streak",         "?", "illegible"),
        ("Total Team Runs Scored",       160, ""),
        ("Total Team Hits",              198, ""),
        ("Total Team Extra-Base Hits",    54, ""),
        ("Total Team Home Runs",          12, ""),
        ("Total Team Stolen Bases",       42, ""),
        ("Total Team Walks",             103, ""),
        ("Total Team HBP",                20, ""),
    ]
    write_table(
        ws_h,
        title="2004 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
        note=(
            "Note: Team Highlights photo was worn. Some context cells "
            "(streaks, opponents) unclear — marked with '?' or 'illegible'."
        ),
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    # Records re-derived directly from the 2004 Team Batting and Team
    # Pitching sheets (header: 'Season to Date Statistics thru 22 Games').
    # The original records page was hand-annotated with 2005 updates and
    # had higher values for some categories — likely captured at a later
    # snapshot. Where the batting sheet lacked a column (SB, hit streak),
    # the records-page value is retained with a note.
    rec_rows = [
        ("Highest Average",        "Nick Mareno",                          ".500", "30+ AB (27-54) · freshman"),
        ("Most At Bats",           "Cassidy Emery",                         75,    ""),
        ("Most HBP",               "Kevin Owens, Matt Pleau",                4,    "tied"),
        ("Lowest K Ratio",         "Nick Mareno",                         "1.85%", "(1-54) · 30+ AB"),
        ("Most Doubles",           "Matt Shearer",                           7,    ""),
        ("Most Triples",           "Brandon Shurick",                        1,    "only player with triples"),
        ("Most Home Runs",         "Cassidy Emery",                          3,    ""),
        ("Most Walks",             "Kevin Owens",                           14,    ""),
        ("Most Stolen Bases",      "Cassidy Emery",                          9,    "from records page (SB not in batting sheet)"),
        ("Most Total Bases",       "Cassidy Emery",                         48,    ""),
        ("Most RBIs",              "Kevin Owens",                           25,    ""),
        ("Best On-Base Avg.",      "Kevin Owens, Troy Burki",            ".511",   "30+ AB · tied"),
        ("Longest Hitting Streak", "Nick Mareno",                           15,    "from records page (not derivable from cumulative stats)"),
        ("Most Runs Scored",       "Kevin Owens",                           25,    ""),
        ("Most Wins Pitching",     "Kevin Bogue",                            8,    "8-1 record"),
        ("Most Innings Pitched",   "Kevin Bogue",                         "64.2",  ""),
        ("Most K's",               "Kevin Bogue",                           78,    ""),
        ("Lowest ERA",             "Kevin Bogue",                         "1.30",  ""),
        ("Most Saves",             "Kevin Bogue, Casey Knox, Seth Whiting",  1,    "3-way tie"),
    ]
    write_table(
        ws4,
        title="2004 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 40, 12, 56],
        left_align_cols={2, 4},
        note=(
            "Records re-derived from the 2004 Team Batting and Team Pitching "
            "sheets ('Season to Date Statistics thru 22 Games'). The original "
            "records page was hand-annotated with 2005 updates. CAVEAT: the "
            "batting page may be a near-final snapshot — end-of-season Most "
            "Total Bases and AB values may have been higher. SB and hit streak "
            "are not derivable from cumulative stats."
        ),
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
