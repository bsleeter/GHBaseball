#!/usr/bin/env python3
"""Build 2010 Gig Harbor Varsity season stats xlsx from pages in
Historical/2010/. This is the Spencer Manjarrez record-setting year —
.538 AVG, 7 HR, .677 OBP, 26 SB (program Hall of Fame records)."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2010" / "2010_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "Roster"
    roster_rows = [
        (1,  "Carl Beck",            12),
        (2,  "David Bigelow",        12),
        (3,  "Barrett Schmidtke",    12),
        (4,  "Scott Schultz",        12),
        (5,  "Parker Guinn",         11),
        (6,  "Spencer Manjarrez",    11),
        (7,  "Michael McCall",       11),
        (8,  "Justin Bonnell",       11),
        (9,  "Eric Johnson",         11),
        (10, "Colin Walters",        11),
        (11, "Jordan Pearson",       11),
        (12, "Eric Flint",           11),
        (13, "Lucas Marshall",       11),
        (14, "Ryan Olson",           11),
        (15, "Steven Daily",         10),
        (16, "Jake Ayers",           10),
        (17, "Garrett Gallinger",     9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2010 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Scott Hatteberg", "Jim Peschek"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "R", "H", "1B", "2B", "3B", "HR", "RBI", "BB",
        "AVG", "SLG", "OBP", "HBP", "SAC-B", "SAC-F", "SO", "TB",
    ]
    # Re-transcribed from clean batting-report photo (BC91544D-...).
    # Per-player rows verified — all 1B/2B/3B/HR breakdowns sum to H, and
    # TB / SLG / OBP cross-check cleanly. The earlier transcription had
    # Guinn's AB inflated (used 97 from records page, which is actually PA).
    bat_rows = [
        ("Guinn",        81, 34, 32, 23, 6, 1, 2, 18, 14, ".395", ".568",  ".495", 2, 0, 0,  6, 46),
        ("Bigelow",      64, 23, 27, 12, 8, 0, 7, 39, 12, ".422", ".875",  ".519", 2, 0, 1, 11, 56),
        ("Schultz",      42,  9, 12,  6, 5, 0, 1, 13,  9, ".286", ".476",  ".412", 0, 0, 0, 11, 20),
        ("Manjarrez",    65, 36, 35, 18, 7, 3, 7, 28, 23, ".538","1.062",  ".677", 7, 0, 1, 11, 69),
        ("Marshall",     48, 13, 12,  7, 3, 1, 1,  9, 16, ".250", ".417",  ".464", 4, 0, 1, 13, 20),
        ("Beck",         70, 20, 24, 18, 5, 0, 1, 18,  9, ".343", ".457",  ".402", 0, 0, 3,  7, 32),
        ("Walters",      62,  9, 16, 12, 4, 0, 0, 14,  4, ".258", ".323",  ".294", 0, 0, 2, 14, 20),
        ("Pearson",      14,  8,  5,  4, 1, 0, 0,  4,  2, ".357", ".429",  ".438", 0, 1, 0,  8,  6),
        ("Daily",        39,  8, 13, 12, 1, 0, 0,  6,  4, ".333", ".359",  ".395", 0, 0, 0, 11, 14),
        ("Bonnell",      55, 17, 22, 22, 0, 0, 0, 19,  6, ".400", ".400",  ".453", 1, 0, 2,  4, 22),
        ("Johnson",      14,  1,  2,  2, 0, 0, 0,  1,  1, ".143", ".143",  ".250", 1, 0, 0,  7,  2),
        ("McCall",        5,  2,  3,  2, 1, 0, 0,  3,  0, ".600", ".800",  ".667", 1, 0, 0,  2,  4),
        ("Flint",         4,  5,  2,  2, 0, 0, 0,  1,  4, ".500", ".500",  ".750", 0, 0, 0,  1,  2),
        ("Schmidtke",     5,  0,  0,  0, 0, 0, 0,  1,  2, ".000", ".000",  ".286", 0, 0, 0,  1,  0),
        ("Ayers",        29, 13,  5,  5, 0, 0, 0,  7,  3, ".172", ".172",  ".242", 0, 0, 1,  6,  5),
        ("Gallinger",    43, 11, 14,  9, 4, 0, 1, 13,  2, ".326", ".488",  ".383", 2, 0, 0,  5, 21),
        ("TEAM",        640,209,224,154,45, 5,20,194,111, ".350", ".530",  ".430",20, 1,11,118,339),
    ]
    write_table(
        ws2,
        title="2010 Gig Harbor Varsity — Team Batting",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Re-transcribed from clean batting-report photo. All "
            "per-player 1B/2B/3B/HR breakdowns sum to H, and TB matches "
            "SLG×AB. The records page lists 'Most At Bats: Guinn 97' — "
            "that 97 is actually PA (81 AB + 14 BB + 2 HBP + 0 SAC); see "
            "Individual Records sheet."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "#BF", "RS", "ER", "BB", "H", "HR", "HBP",
        "W", "L", "SV", "BAA", "ERA", "SO",
    ]
    pit_rows = [
        ("Bigelow",    "34.1",  155, 19, 11, 12, 26, 0, 1, 4, 2, 2, ".186", "2.24", 52),
        ("Schultz",    "37.2",  149, 14,  8, 14, 20, 1, 3, 4, 3, 0, ".152", "1.49", 45),
        ("Beck",        "1.0",    5,  0,  0,  1,  1, 0, 0, 0, 0, 0, ".250", "0.00",  3),
        ("Daily",      "11.0",   48,  7,  6, 14,  3, 0, 0, 2, 0, 0, ".088", "3.82",  8),
        ("McCall",     "39.0",  154, 11,  8,  3, 21, 0, 2, 4, 0, 0, ".111", "1.44", 20),
        ("Flint",       "3.1",   21,  5,  2,  3,  2, 0, 0, 0, 0, 0, ".111", "4.20",  6),
        ("Schmidtke",  "16.0",   68,  5,  4,  7, 10, 0, 0, 2, 1, 0, ".169", "1.75", 19),
        ("TEAM",      "142.1",  600, 57, 39, 54, 83, 1, 8,18, 6, 2, ".155", "1.92",183),
    ]
    write_table(
        ws3,
        title="2010 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              41, "vs. Lincoln"),
        ("Most Runs (Inning)",            20, "vs. Lincoln"),
        ("Most Runs Allowed (Game)",      11, "vs. Federal Way"),
        ("Widest Margin of Victory",      41, "vs. Lincoln"),
        ("One Run Games Lost",             3, "vs. Puyallup, CK, Redmond"),
        ("One Run Games Won",              3, "vs. Wilson, Olympia, Timberline"),
        ("Most K's (by GH Pitchers)",     14, "vs. Shelton"),
        ("Most K's (by Opponents)",       17, "vs. Puyallup (11 innings)"),
        ("Most Hits (Game)",              31, "vs. Lincoln"),
        ("Most Hits Allowed (Game)",       9, "vs. Federal Way"),
        ("Most Walks (Game for GH)",      11, "vs. Lincoln"),
        ("Most Walks Allowed",            12, "vs. Foss"),
        ("Most Singles (Game)",           21, "vs. Lincoln"),
        ("Most Doubles (Game)",            7, "vs. Lincoln"),
        ("Most Triples (Game)",            2, "vs. Bellarmine"),
        ("Most Home Runs (Game)",          3, "vs. Lincoln, Shelton"),
        ("Most Steals (Game)",             7, "vs. Foss"),
        ("Longest Winning Streak",         9, ""),
        ("Longest Losing Streak",          2, ""),
        ("Total Team Runs Scored",       209, ""),
        ("Total Team Hits",              224, ""),
        ("Total Team Extra-Base Hits",    70, ""),
        ("Total Team Home Runs",          20, "PROGRAM RECORD — program bests team HR mark"),
        ("Total Team Stolen Bases",       72, ""),
        ("Total Team Walks",             111, ""),
        ("Total Team HBP",                20, ""),
    ]
    write_table(
        ws_h,
        title="2010 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 48],
        left_align_cols={1, 3},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Spencer Manjarrez",                                ".538", "30+ AB (35-65) · PROGRAM RECORD"),
        ("Most Plate Appearances", "Parker Guinn",                                      97,    "PA = AB + BB + HBP + SAC (records page labels this 'At Bats')"),
        ("Most HBP",               "Spencer Manjarrez",                                  7,    ""),
        ("Lowest K Ratio",         "Parker Guinn",                                     "6.2%", "6 K / 97 PA"),
        ("Most Doubles",           "David Bigelow",                                      8,    ""),
        ("Most Triples",           "Spencer Manjarrez",                                  3,    ""),
        ("Most Home Runs",         "David Bigelow, Spencer Manjarrez",                   7,    "PROGRAM RECORD (tied)"),
        ("Most Walks",             "Spencer Manjarrez",                                 23,    ""),
        ("Most Stolen Bases",      "Spencer Manjarrez",                                 26,    "PROGRAM RECORD"),
        ("Most Total Bases",       "Spencer Manjarrez",                                 69,    ""),
        ("Most RBIs",              "David Bigelow",                                     39,    ""),
        ("Best On-Base Avg.",      "Spencer Manjarrez",                                ".677", "30+ AB · PROGRAM RECORD"),
        ("Longest Hitting Streak", "Parker Guinn",                                      13,    ""),
        ("Most Runs Scored",       "Spencer Manjarrez",                                 36,    ""),
        ("Most Wins Pitching",     "Scott Schultz, David Bigelow, Michael McCall",       4,    ""),
        ("Most Innings Pitched",   "Michael McCall",                                    39,    ""),
        ("Most K's",               "David Bigelow",                                     52,    ""),
        ("Lowest ERA",             "Michael McCall",                                  "1.44",  ""),
        ("Most Saves",             "David Bigelow",                                      2,    ""),
    ]
    write_table(
        ws4,
        title="2010 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 46, 12, 36],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
