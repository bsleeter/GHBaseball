#!/usr/bin/env python3
"""Build 2014 Gig Harbor Varsity season stats xlsx from scanned records
in Historical/2014/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2014" / "2014_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None, col_widths=None,
                left_align_cols=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    r = 4
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Schedule"
    schedule_rows = [
        ("Timberline",   "0-7"),
        ("Curtis",       "2-0"),
        ("Port Angeles", "2-0"),
        ("Stadium",      "23-0"),
        ("Wilson",       "5-3"),
        ("Yelm",         "13-0"),
        ("Yelm",         "0-2"),
        ("Bellarmine",   "6-4"),
        ("Olympia",      "5-4"),
        ("Stadium",      "6-3"),
        ("S. Kitsap",    "4-2"),
        ("S. Kitsap",    "3-2"),
        ("Shelton",      "1-5"),
        ("N. Thurston",  "1-5"),
        ("Bellarmine",   "4-5"),
        ("Olympia",      "11-0"),
        ("C. Kitsap",    "3-4"),
        ("Peninsula",    "5-4"),
        ("Olympia",      "1-5"),
        ("S. Kitsap",    "1-3"),
        ("Puyallup",     "2-3"),
        ("Skyview",      "2-3"),
        ("Union",        "8-5"),
        ("Kentwood",     "4-22"),
        ("Cascade",      "4-8"),
        ("Season Record", "16-10"),
    ]
    write_table(
        ws,
        title="2014 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        col_widths=[22, 18],
        left_align_cols={1},
    )

    # ─── Sheet 2: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Drew Barnett",        12),
        (2,  "Sterling Brown",      12),
        (3,  "Quintin Carlson",     12),
        (4,  "Matt Jones",          12),
        (5,  "Chad Glover",         12),
        (6,  "Conor Scanlan",       12),
        (7,  "Mason Selby",         12),
        (8,  "Mark Sluys",          12),
        (9,  "Joe Bundick",         12),
        (10, "Dean Hassan",         12),
        (11, "Matthew Henckel",     11),
        (12, "Casey Gearhart",      11),
        (13, "Grant Sutton",        11),
        (14, "Hunter Werner",       11),
        (15, "Drew Gallinger",      10),
        (16, "Michael Toglia",      10),
        (17, "Jon Burghardt",       10),
        (18, "Alex Morford",        10),
        (19, "Patrick Fletcher",    10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2014 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Larry Roehr", "Dale Payne"],
    )

    # ─── Sheet 3: Hitting & Fielding ────────────────────────────────
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Barnett",   41,  8, 13,  5, 11, 0, 0, 2, 21, ".317", ".350", ".406",  9,  3, 0, 2,  33,  0, 0, "1.000", 0),
        ("Brown",     84, 21, 23,  9, 20, 5, 0, 0, 28, ".274", ".333", ".385",  9,  8, 2, 1,  21, 31, 9,  ".852", 5),
        ("Bundick",   10,  2,  2,  1,  2, 0, 0, 0,  2, ".200", ".200", ".273",  2,  1, 0, 0,   2,  1, 0, "1.000", 0),
        ("Burghardt", 23,  2,  5,  4,  4, 0, 0, 1,  9, ".217", ".391", ".308",  8,  2, 1, 0,  10,  0, 0, "1.000", 0),
        ("Carlson",   52, 19,  8,  8,  6, 2, 0, 0, 10, ".154", ".192", ".271", 11,  4, 0, 0,  25,  3, 1,  ".964", 2),
        ("Fletcher",   2,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  2,  0, 0, 0,   0,  0, 0,  ".000", 0),
        ("Gallinger",  0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   3,  4, 2,  ".778", 0),
        ("Gearhart",   3,  0,  2,  0,  2, 0, 0, 0,  2, ".667", ".667", ".667",  0,  0, 0, 0,   0,  1, 0, "1.000", 0),
        ("Glover",    23,  4,  3,  1,  3, 0, 0, 0,  3, ".130", ".130", ".200",  6,  2, 1, 2,   2,  5, 2,  ".778", 0),
        ("Hassan",    46, 10, 11,  6, 11, 0, 0, 0, 11, ".239", ".239", ".327",  7,  6, 2, 0,   3,  4, 2,  ".778", 0),
        ("Henckel",    4,  0,  1,  1,  1, 0, 0, 0,  1, ".250", ".250", ".250",  1,  0, 0, 0,   1,  2, 0, "1.000", 0),
        ("Jones",     39,  7,  5,  1,  4, 1, 0, 0,  6, ".128", ".154", ".200",  5,  4, 0, 0,   0,  0, 0,  ".000", 0),
        ("Morford",    0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   2,  3, 0, "1.000", 0),
        ("Scanlan",   83, 11, 21, 12, 18, 2, 0, 1, 27, ".253", ".325", ".344", 12,  5, 1, 2,  14, 21, 1,  ".972", 0),
        ("Selby",     56,  8, 14,  7, 10, 2, 1, 1, 26, ".253", ".313", ".271", 22,  1, 1, 1,  35, 10, 3,  ".938", 3),
        ("Sluys",     70,  2, 23,  9, 19, 2, 2, 0, 33, ".329", ".471", ".434",  7, 12, 0, 1,  41, 59,13,  ".885", 4),
        ("Sutton",    75, 17, 25, 24, 18, 4, 1, 1, 34, ".333", ".453", ".419", 19,  7, 0, 1,  14, 21, 1,  ".972", 1),
        ("Toglia",    74, 15, 23, 14, 14, 2, 5, 0, 32, ".311", ".432", ".419",  5,  6, 1, 0, 159, 11, 1, "1.000", 4),
        ("Werner",    24,  2,  6,  4,  6, 0, 0, 0,  6, ".250", ".250", ".308",  7,  2, 0, 0,   3,  3, 0, "1.000", 0),
        ("TEAM",     692,130,187,112,150, 27, 6, 4,238, ".270", ".344", ".356",135, 72,20,18, 410,230,43,  ".937",30),
    ]
    write_table(
        ws2,
        title="2014 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=[14] + [7] * (len(hit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Pitching ──────────────────────────────────────────
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF/AB", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Carlson",    0, 0, 0, "3.0",   16,  2,  5,  1,  0,  1,   5,  0, 0, 0, 0, "2.333"),
        ("Gallinger",  2, 2, 1, "35.1", 152, 22, 34, 25,  7, 17,  25,  7, 1, 1, 1, "3.390"),
        ("Gearhart",   4, 1, 1, "25.0", 112, 14, 28, 14, 14, 13,  21,  6, 1, 2, 2, "3.640"),
        ("Hassan",     2, 1, 0, "21.0",  86,  9, 18, 16, 11,  8,  10,  7, 1, 0, 1, "2.667"),
        ("Henckel",    1, 1, 0, "10.2",  53,  8,  8, 16, 12,  8,   6,  2, 0, 0, 1, "5.490"),
        ("Morford",    0, 0, 0, "1.1",   13,  7,  7,  1,  4,  7,   1,  1, 0, 2, 0, "44.55"),
        ("Scanlan",    2, 3, 0, "28.1", 129, 14, 32, 16, 12, 13,  22,  8, 0, 2, 3, "3.238"),
        ("Selby",      0, 0, 0, "1.2",   11,  4,  5,  2,  0,  3,   2,  2, 0, 1, 0, "17.50"),
        ("Sutton",     0, 1, 0, "4.0",   18,  0,  2,  4,  4,  0,   2,  0, 0, 0, 0, "0.000"),
        ("Toglia",     4, 1, 2, "34.1", 133,  4, 27, 23,  4,  3,  21,  6, 0, 0, 2, "0.616"),
        ("Werner",     1, 1, 0, "12.1",  58, 11, 15,  7,  7, 10,  10,  5, 0, 0, 0, "5.785"),
        ("TEAM",      16,10, 5, "177.1", 781, 95,179,125, 77, 83, 125, 44, 2, 7,10, "3.281"),
    ]
    write_table(
        ws3,
        title="2014 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Grant Sutton",                                        ".333", "30+ AB (25-75)"),
        ("Most Plate Appearances", "Sterling Brown",                                       97,    "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",         "Mark Sluys",                                          "7.9%", "(7-89)"),
        ("Most Doubles",           "Sterling Brown, Michael Toglia",                        5,    ""),
        ("Most Triples",           "Mark Sluys, Michael Toglia",                            2,    ""),
        ("Most Home Runs",         "Conor Scanlan, Mark Sluys, Grant Sutton, Jon Burghardt", 1,   ""),
        ("Most Walks",             "Mark Sluys",                                           12,    ""),
        ("Most Stolen Bases",      "Sterling Brown",                                        5,    ""),
        ("Most Total Bases",       "Grant Sutton",                                         34,    ""),
        ("Most RBIs",              "Grant Sutton",                                         24,    ""),
        ("Best On-Base Avg.",      "Grant Sutton",                                        ".419", "30+ AB"),
        ("Longest Hitting Streak", "Grant Sutton",                                          8,    ""),
        ("Most Runs Scored",       "Sterling Brown",                                       21,    ""),
        ("Most Wins Pitching",     "Michael Toglia, Casey Gearhart",                        4,    ""),
        ("Most Innings Pitched",   "Drew Gallinger",                                      "35.1", ""),
        ("Most K's",               "Drew Gallinger",                                       25,    ""),
        ("Lowest ERA",             "Michael Toglia",                                      "0.62", "14+ IP (0.616)"),
        ("Most Saves",             "Michael Toglia",                                        2,    ""),
    ]
    write_table(
        ws4,
        title="2014 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 44, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
