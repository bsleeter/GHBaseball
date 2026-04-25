#!/usr/bin/env python3
"""Build 1992 Gig Harbor Varsity season stats xlsx."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import write_table, write_readme, write_roster_sheet
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "1992" / "1992_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 1992, [
        ("Individual Records", "Clean typed source labeled '1992 Records'.", "READABLE"),
        ("Team Highlights", "Typed source labeled 'Highlights of 1992 Team'.", "READABLE"),
        ("Schedule", "No schedule photo provided.", "NOT PROVIDED"),
        ("Team Batting", "Hand-written stats sheet (Pierce County League — Season Records). Per-player TB cells don't sum cleanly to the visible breakdown — preserved as printed.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Hand-written stats sheet (bottom of stats page). Five pitchers logged: Cherbas, Lindmark, Lippert, Sawyers, Siegmund.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Greg Bleistein",      12),
        (2,  "Andy Cherbas",        10),
        (3,  "Dale Cox",            11),
        (4,  "Tom Friedman",         9),
        (5,  "Scott Harter",        12),
        (6,  "Chad Heilesen",       11),
        (7,  "Billy Jack Kirk",     12),
        (8,  "Christian Lindmark",  10),
        (9,  "Jason Lippert",        9),
        (10, "Joel Miller",          9),
        (11, "Jim Peschek",         12),
        (12, "Ryan Sawyers",        10),
        (13, "Sig Siegmund",        12),
        (14, "Derek Vitcovich",      9),
        (15, "Isaac Wong",          12),
    ]
    write_roster_sheet(ws_r, "1992 Gig Harbor Varsity — Roster", roster_rows,
                       coaches=["Peter Jansen", "Kevin Miller"],
                       manager="(Scorekeeper: Jon Erickson)")

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "G", "AB", "R", "H", "RBI", "2B", "3B", "HR", "BB",
        "HBP", "K", "SAC", "TB", "SB", "AVG", "OBP",
    ]
    bat_rows = [
        ("Bleistein", 14, 22,  2,  2,  1, 0, 0, 0,  3, 1,  9, 1,  6,  0, ".090", ".231"),
        ("Cherbas",   15, 47, 10, 13,  6, 2, 0, 0,  9, 2,  7, 1, 27,  6, ".277", ".414"),
        ("Cox",        9,  5,  3,  0,  0, 0, 0, 0,  2, 0,  2, 0,  2,  1, ".000", ".286"),
        ("Friedman",  16, 44,  6, 14, 13, 3, 1, 1,  5, 3,  8, 0, 27,  3, ".318", ".423"),
        ("Harter",    16, 39,  7, 10,  8, 1, 0, 0,  4, 5, 11, 1, 20,  4, ".256", ".396"),
        ("Heilesen",   9, 11,  1,  1,  0, 0, 0, 0,  0, 1,  4, 1,  2,  1, ".090", ".167"),
        ("Kirk",      16, 50, 13, 12,  6, 0, 0, 0, 12, 0,  8, 3, 24,  8, ".240", ".387"),
        ("Lindmark",   7,  5,  1,  1,  1, 0, 0, 0,  0, 0,  3, 0,  1,  0, ".200", ".200"),
        ("Lippert",   14, 21,  4,  2,  3, 1, 0, 0,  2, 1,  5, 1,  5,  0, ".095", ".208"),
        ("Miller",    16, 32,  9,  5,  1, 0, 0, 0,  7, 0,  7, 2, 12, 10, ".156", ".308"),
        ("Peschek",   16, 51, 10, 20, 15, 5, 1, 0,  8, 2,  6, 1, 37,  2, ".392", ".492"),
        ("Sawyers",   14, 30,  5, 12,  8, 1, 1, 0, 11, 0,  2, 0, 27,  1, ".400", ".561"),
        ("Siegmund",  16, 52,  4, 14,  3, 2, 0, 0,  3, 0,  2, 2, 21,  1, ".269", ".333"),
        ("Vitcovich",  7,  4,  0,  0,  0, 0, 0, 0,  2, 0,  3, 0,  2,  0, ".000",  "—"),
        ("Wong",      12, 25,  2,  7,  1, 1, 0, 0,  2, 0,  5, 0, 10,  0, ".280", ".333"),
        ("TEAM",     208,438, 77,113, 66,16, 3, 1, 70,15, 82,13,223, 37, ".258", ".368"),
    ]
    write_table(
        ws2,
        title="1992 Gig Harbor Varsity — Pierce County League Season Records (Batting)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 5] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
        note=(
            "Note: Hand-written stats sheet (Pierce County League — Season Records). "
            "Per-player TB cells don't sum cleanly to the printed 1B/2B/3B/HR "
            "breakdown — preserved as printed. The records page (Highest Avg "
            "Sawyers .400, Most Hits Peschek 20, Most TB Peschek 37, etc.) is "
            "the authoritative source for top numbers."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "G", "IP", "BF", "R", "H", "K", "BB", "ER",
        "HBP", "ERA",
    ]
    pit_rows = [
        ("Cherbas",   0, 0, 0, 4,  "8.2",  34,  3,  7,  5,  2,  1, 1,  "0.81"),
        ("Lindmark",  1, 1, 0, 5, "19.0",  85, 16, 20, 16,  7,  5, 2,  "1.81"),
        ("Lippert",   2, 0, 0, 6, "22.0",  84, 14, 20, 14, 15, 11, 1,  "3.50"),
        ("Sawyers",   3, 2, 0, 8, "27.0", 106, 18, 25, 31, 11,  4, 1,  "1.04"),
        ("Siegmund",  5, 2, 0, 9, "39.0", 159, 20, 39, 30,  8,  9, 2,  "1.62"),
        ("TEAM",     11, 5, 0,32,"115.2", 468, 71,111, 96, 43, 30, 7,  "1.82"),
    ]
    write_table(
        ws3,
        title="1992 Gig Harbor Varsity — Pierce County League Season Records (Pitching)",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [6] * (len(pit_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Five pitchers logged. Records page confirms: Most Wins Siegmund 5, "
            "Most IP Siegmund 39, Most K's Sawyers 31, Lowest ERA Sawyers 1.04 "
            "(min. 15 IP). Cherbas's IP read as 8.2 to match printed 0.81 ERA "
            "with 1 ER."
        ),
    )

    ws_h = wb.create_sheet("Team Highlights")
    write_table(ws_h, "Highlights of 1992 Team",
        ["Highlight", "Value", "Context"],
        [
            ("Most Runs (Game)",              9, "vs. Enumclaw, F.P."),
            ("Most Runs (Inning)",            4, "vs. F.P."),
            ("Most Runs Allowed (Game)",     10, "vs. Peninsula"),
            ("Widest Margin of Victory",      6, "vs. Tahoma, F.P."),
            ("One Run Games Won",             4, "vs. F.P., Tahoma, Fife, Peninsula"),
            ("One Run Games Lost",            1, "vs. White River"),
            ("Most K's (by GH Pitchers)",    12, "vs. Yelm"),
            ("Most K's (by Opponents)",      14, "vs. Enumclaw"),
            ("Most Hits (Game)",             11, "vs. Tahoma"),
            ("Most Hits Allowed (Game)",     12, "vs. Enumclaw, Fife"),
            ("Most Walks (Game for GH)",      7, "vs. Yelm"),
            ("Most Walks Allowed",            6, "vs. Enumclaw, F.P."),
            ("Most Singles (Game)",           9, "vs. Washington"),
            ("Most Doubles (Game)",           3, "vs. Tahoma"),
            ("Most Triples (Game)",           1, "vs. Washington, F.P., White River, Fife"),
            ("Most Home Runs (Game)",         1, "vs. Tahoma"),
            ("Most Steals (Game)",            5, "vs. Yelm, Washington"),
            ("Most Errors (Game)",            9, "vs. Peninsula"),
            ("Longest Winning Streak",        3, "(twice)"),
            ("Longest Losing Streak",         1, "(five times)"),
        ],
        col_widths=[34, 10, 56], left_align_cols={1, 3})

    ws4 = wb.create_sheet("Individual Records")
    write_table(ws4, "1992 Records",
        ["Record", "Holder", "Value", "Qualifier / Note"],
        [
            ("Highest Average",        "Ryan Sawyers",                               ".400", ""),
            ("Most Hits",              "Jim Peschek",                                 20,    ""),
            ("Most At Bats",           "Sig Siegmund",                                52,    ""),
            ("Most Runs Scored",       "Billy Jack Kirk",                             13,    ""),
            ("Most HBP",               "Scott Harter",                                 5,    ""),
            ("Most Doubles",           "Jim Peschek",                                  5,    ""),
            ("Most Triples",           "Tom Friedman, Jim Peschek, Ryan Sawyers",      1,    ""),
            ("Most Home Runs",         "Tom Friedman",                                 1,    ""),
            ("Most Walks",             "Billy Jack Kirk",                             12,    ""),
            ("Most Stolen Bases",      "Joel Miller",                                 10,    ""),
            ("Most Total Bases",       "Jim Peschek",                                 37,    ""),
            ("Most RBIs",              "Jim Peschek",                                 15,    ""),
            ("Most Wins Pitching",     "Sig Siegmund",                                 5,    ""),
            ("Most Innings Pitched",   "Sig Siegmund",                                39,    ""),
            ("Most K's",               "Ryan Sawyers",                                31,    ""),
            ("Lowest ERA",             "Ryan Sawyers",                              "1.04",  "Min. 15 IP"),
            ("Best On-Base Average",   "Ryan Sawyers",                              ".561",  ""),
            ("Lowest K Ratio (Bat)",   "Sig Siegmund",                              ".035",  "(2 K / 57 AB)"),
            ("Longest Hitting Streak", "Andy Cherbas, Jim Peschek, Billy Jack Kirk",   6,    ""),
        ],
        col_widths=[26, 46, 12, 22], left_align_cols={2, 4})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
