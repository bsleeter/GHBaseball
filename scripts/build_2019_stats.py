#!/usr/bin/env python3
"""Build the 2019 Gig Harbor Varsity season stats xlsx from the scanned
records in Historical/2019/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2019" / "2019_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_section_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE, italic=True)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1))
            if c == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None, section_rows=None, col_widths=None):
    """section_rows: dict of {row_index_in_rows: label} to insert a spanning header row above."""
    section_rows = section_rows or {}
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    # Headers on row 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20

    # Data rows
    r = 4
    for i, row in enumerate(rows):
        # Section header insertion
        if i in section_rows:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            sc = ws.cell(row=r, column=1, value=section_rows[i])
            style_section_header(sc)
            r += 1

        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1))
        r += 1

    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule ───
    ws = wb.active
    ws.title = "Schedule"
    schedule_rows = [
        ("Bainbridge", "0-4"),
        ("Bellarmine", "4-10"),
        ("Wilson", "9-1"),
        ("Bellevue", "5-0"),
        ("Timberline", "4-2"),
        ("Timberline", "6-4"),
        ("Yelm", "5-0"),
        ("Yelm", ""),
        ("N. Thurston", "0-5"),
        ("N. Thurston", "2-5"),
        ("Rogers", "14-1"),
        ("Capital", "13-0"),
        ("Capital", "5-1"),
        ("N. Kitsap", "3-6"),
        ("Shelton", "6-0"),
        ("Shelton", "12-1"),
        ("Peninsula", "0-2"),
        ("Peninsula", "9-8"),
        ("Capital", "4-8"),
        ("C. Kitsap", "12-1"),
        ("N. Kitsap", "3-1"),
        ("Capital", "3-0"),
        ("Mt. View", "9-3"),
        ("Lk. Wash.", "0-7"),
        ("Capital", "11-4"),
        # State tournament
        ("Blanchet", "2-1"),
        ("Ingraham", "15-2"),
        # Season totals
        ("Season Record", "20-7"),
    ]
    write_table(
        ws,
        title="2019 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        section_rows={25: "WA 3A State Tournament"},
        col_widths=[22, 18],
    )

    # ─── Sheet 2: Roster ───
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Luke Deschenes",    11),
        (2,  "Cade Dessert",      11),
        (3,  "Cree Evenson",      12),
        (4,  "Luke Finnigan",     11),
        (5,  "Grady Glover",      12),
        (6,  "Cage Hardy",        12),
        (7,  "Grant Hassan",      12),
        (8,  "Peter Losh",        11),
        (9,  "Tyler Peterson",    10),
        (10, "Cole Pringle",      12),
        (11, "Zane Skansi",       11),
        (12, "Chad Sorenson",     12),
        (13, "Max Sparrow",       12),
        (14, "Sebastian Toglia",  12),
        (15, "Zach Toglia",       10),
        (16, "Eric Underwood",    12),
        (17, "Ryan Voves",        10),
        (18, "Owen Wild",         12),
    ]
    _write_roster_sheet(
        ws_r,
        title="2019 Gig Harbor Tides — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Larry Roehr", "Oscar Ortiz", "Todd Davis"],
    )

    # ─── Sheet 3: Hitting & Fielding ───
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Deschenes", 40, 2, 10, 6, 9, 1, 0, 0, 11, ".250", ".275", ".388", 8, 5, 0, 0, 4, 72, 5, "1.000", 0),
        ("Dessert",   81, 18, 23, 18, 16, 6, 1, 0, 31, ".284", ".383", ".426", 15, 17, 3, 2, 45, 4, 0, "1.000", 0),
        ("Evenson",   19, 8, 3, 3, 3, 0, 0, 0, 3, ".158", ".158", ".333", 2, 5, 0, 2, 1, 3, 0, "1.000", 0),
        ("Finnigan",  3, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 0, 0, 0, 0, 0, 0, 0, ".000", 0),
        ("Glover",    79, 14, 19, 12, 17, 2, 0, 0, 21, ".241", ".266", ".341", 11, 7, 5, 1, 27, 50, 4, ".951", 0),
        ("Hardy",     69, 3, 20, 10, 19, 1, 0, 0, 21, ".290", ".290", ".479", 11, 11, 14, 2, 4, 0, 0, ".881", 0),
        ("Hassan",    2, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 0, 0, 0, 0, 0, 3, 2, ".600", 0),
        ("Losh",      7, 8, 3, 0, 1, 2, 1, 0, 0, ".429", ".571", ".600", 2, 1, 2, 0, 1, 4, 0, ".833", 1),
        ("Skansi",    83, 23, 31, 26, 22, 7, 0, 2, 44, ".373", ".530", ".469", 7, 13, 2, 2, 35, 47, 12, ".872", 2),
        ("Sorenson",  30, 11, 10, 5, 10, 0, 0, 0, 10, ".333", ".333", ".545", 11, 12, 2, 0, 2, 3, 0, ".750", 1),
        ("Sparrow",   0, 0, 0, 0, 0, 0, 0, 0, 0, ".000", ".000", ".000", 0, 0, 0, 0, 3, 1, 0, "1.000", 0),
        ("Toglia, S", 83, 17, 26, 10, 20, 6, 0, 0, 32, ".313", ".386", ".447", 14, 11, 6, 1, 29, 3, 1, ".970", 3),
        ("Toglia, Z", 73, 22, 29, 28, 21, 6, 2, 0, 39, ".397", ".534", ".532", 8, 19, 2, 1, 15, 39, 6, ".900", 0),
        ("Voves",     70, 11, 17, 9, 9, 6, 2, 0, 27, ".243", ".386", ".329", 13, 5, 4, 6, 19, 3, 0, ".864", 4),
        ("Wild",      82, 28, 29, 17, 22, 6, 0, 1, 38, ".354", ".463", ".470", 5, 14, 4, 5, 14, 100, 12, ".991", 2),
        ("TEAM",      716, 170, 220, 145, 170, 42, 5, 3, 281, ".307", ".392", ".441", 104, 125, 46, 19, 361, 200, 39, ".935", 25),
    ]
    hit_widths = [12] + [7] * (len(hit_headers) - 1)
    write_table(
        ws2,
        title="2019 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=hit_widths,
    )

    # ─── Sheet 3: Pitching ───
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Dessert",   1, 0, 0, 3.0, 16, 2, 4, 3, 2, 1, 3, 1, 0, 0, 1, 2.333),
        ("Finnigan",  5, 1, 0, 30.2, 128, 15, 20, 23, 13, 11, 16, 3, 0, 1, 8, 2.510),
        ("Hassan",    1, 0, 1, 18.2, 86, 9, 23, 10, 6, 5, 21, 1, 0, 1, 3, 1.874),
        ("Pringle",   0, 1, 0, 10.1, 53, 12, 18, 4, 2, 9, 16, 1, 1, 0, 2, 6.098),
        ("Sparrow",   2, 2, 0, 27.0, 124, 22, 27, 29, 14, 15, 21, 5, 0, 1, 2, 3.889),
        ("Underwood", 2, 0, 1, 12.1, 48, 5, 11, 10, 6, 4, 8, 2, 0, 0, 0, 2.270),
        ("Wild",      9, 3, 0, 78.0, 304, 15, 46, 112, 22, 14, 41, 5, 0, 0, 4, 1.256),
        ("TEAM",      20, 7, 2, 179.0, 759, 80, 149, 191, 65, 59, 126, 18, 1, 4, 20, 2.307),
    ]
    pit_widths = [12] + [7] * (len(pit_headers) - 1)
    write_table(
        ws3,
        title="2019 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=pit_widths,
    )

    # ─── Sheet 4: Individual Records ───
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",       "Zach Toglia",             ".397",  "30+ AB (29-73)"),
        ("Most Plate Appearances","Sebastian Toglia",         104,     "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",        "Owen Wild",               "5.0%",  "(5-101)"),
        ("Most Doubles",          "Zane Skansi",              7,       ""),
        ("Most Triples",          "Zach Toglia, Ryan Voves",  2,       ""),
        ("Most Home Runs",        "Zane Skansi",              2,       ""),
        ("Most Walks",            "Zach Toglia",              19,      ""),
        ("Most Stolen Bases",     "Zane Skansi",              7,       ""),
        ("Most Total Bases",      "Zach Toglia",              44,      ""),
        ("Most RBIs",             "Zach Toglia",              28,      ""),
        ("Best On-Base Avg.",     "Chad Sorenson",           ".545",  "30+ AB"),
        ("Longest Hitting Streak","Owen Wild",                6,       ""),
        ("Most Runs Scored",      "Owen Wild",                28,      ""),
        ("Most Wins Pitching",    "Owen Wild",                9,       ""),
        ("Most Innings Pitched",  "Owen Wild",                78,      ""),
        ("Most K's",              "Owen Wild",                112,     ""),
        ("Lowest ERA",            "Owen Wild",               "1.27",  "21+ IP"),
        ("Most Saves",            "Grant Hassan, Eric Underwood", 1,  ""),
    ]
    write_table(
        ws4,
        title="2019 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 30, 12, 22],
    )
    # Left-align holder and qualifier columns for readability
    for row in ws4.iter_rows(min_row=4, max_row=3 + len(rec_rows), min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in ws4.iter_rows(min_row=4, max_row=3 + len(rec_rows), min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
