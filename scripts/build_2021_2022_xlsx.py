"""Build 2021 and 2022 xlsx files from the source PDFs.

The 2021 and 2022 seasons came as PDFs (no spreadsheets exist on disk),
so this script transcribes the data into the canonical xlsx format used
by build_master_data.py — same shape as Historical/2023/2024 (which have
sheets: Hitting & Fielding, Pitching, optionally Individual Records).

Run once to (re)generate. Schedule data lives in scripts/data/team_seasons.json
since the build pipeline expects it there for years without xlsx schedule sheets.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
HISTORICAL = ROOT / "Historical"


# ─── Stat-line transcriptions (from the source PDFs) ─────────────────────────

BAT_HEADER = ["Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
              "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "SB"]
PIT_HEADER = ["Player", "W", "L", "SV", "IP", "BF", "R", "H", "SO", "BB", "ER",
              "1B", "2B", "3B", "HR", "HBP", "ERA"]


def n(v):
    """Coerce empty-string/None to 0 (PDF blanks meant 'no count')."""
    return 0 if v in (None, "") else v


# 2021 batting (page 1 of 2021 PDF). Players excluded if 0 AB and no
# meaningful contribution beyond fielding-only.
BATTING_2021 = [
    # Player,         AB, R, H, RBI, 1B, 2B, 3B, HR, TB, AVG,    SLG,    OBP,    SO, BB, HBP, SAC, SB
    ["Colello",       17, 5, 4, 3,   3,  1,  0,  0,  6,  .235,   .353,   .381,   4,  4,  0,   0,   0],
    ["Cosmos",        1,  0, 0, 0,   0,  0,  0,  0,  0,  .000,   .000,   .000,   0,  0,  0,   0,   0],
    ["Flaherty",      4,  0, 0, 0,   0,  0,  0,  0,  0,  .000,   .000,   .200,   2,  1,  0,   0,   0],
    ["Hento",         2,  2, 1, 0,   1,  0,  0,  0,  1,  .500,   .500,   .500,   1,  0,  0,   0,   2],
    ["Hodges",        17, 5, 6, 4,   4,  2,  0,  0,  8,  .353,   .471,   .577,   1,  5,  4,   1,   0],
    ["Holum",         9,  1, 2, 0,   2,  0,  0,  0,  2,  .222,   .222,   .300,   3,  1,  0,   0,   0],
    ["Jones",         31, 11, 10, 6, 9,  1,  0,  0,  11, .323,   .355,   .417,   8,  3,  2,   0,   8],
    ["Knowles",       19, 2, 6, 6,   5,  1,  0,  0,  7,  .316,   .368,   .435,   4,  4,  1,   0,   0],
    ["Lundberg",      25, 9, 4, 1,   3,  1,  0,  0,  5,  .160,   .200,   .400,   8,  9,  1,   0,   5],
    ["McLellan",      11, 1, 2, 2,   2,  0,  0,  0,  2,  .182,   .182,   .250,   2,  1,  0,   0,   1],
    ["Miller",        8,  0, 0, 0,   0,  0,  0,  0,  0,  .000,   .000,   .200,   7,  2,  0,   0,   0],
    ["Peterson",      33, 5, 8, 11,  6,  2,  0,  0,  10, .242,   .303,   .286,   3,  2,  1,   0,   2],
    ["Reed",          25, 4, 4, 3,   3,  1,  0,  0,  5,  .160,   .200,   .222,   8,  1,  1,   1,   0],
    ["Toglia",        30, 10, 14, 13, 9, 4,  0,  1,  21, .467,   .700,   .556,   4,  5,  1,   0,   0],
    ["Voves",         26, 4, 8, 5,   7,  1,  0,  0,  10, .308,   .385,   .308,   5,  0,  0,   2,   0],
]
# 2021 team batting totals from PDF
TEAM_BAT_2021 = [
    "TEAM", 258, 59, 69, 53, 54, 12, 2, 1, 88, .267, .341, .382, 61, 39, 9, 7, 18,
]

BATTING_2022 = [
    # Player,             AB, R, H, RBI, 1B, 2B, 3B, HR, TB, AVG,    SLG,    OBP,    SO, BB, HBP, SAC, SB
    ["Alex Vela",          7,  1, 1, 1,   1,  0,  0,  0,  1,  .143,   .143,   .250,   1,  0,  1,   0,   0],
    ["Brady Altman",       79, 33, 31, 24, 22, 5, 0,  4,  44, .392,   .557,   .505,   13, 17, 1,   2,   13],
    ["Sam Barber",         11, 1, 2, 3,   2,  0,  0,  0,  2,  .182,   .182,   .357,   2,  3,  0,   0,   1],
    ["Justin Holum",       62, 24, 18, 10, 14, 1, 0,  0,  22, .290,   .355,   .436,   11, 10, 6,   1,   17],
    ["JD Dunham",          66, 20, 21, 12, 15, 5, 0,  1,  28, .318,   .424,   .400,   14, 6,  3,   0,   3],
    ["Sam Haddon",         18, 5, 3, 2,   1,  2,  0,  0,  5,  .167,   .278,   .375,   9,  4,  2,   3,   0],
    ["Ryland Heckman",     44, 10, 17, 8,  9,  5,  1,  1,  26, .386,   .591,   .471,   7,  6,  1,   0,   2],
    ["Trevor Hellwich",    49, 6, 13, 10, 12, 1,  0,  0,  14, .265,   .286,   .357,   7,  5,  2,   1,   5],
    ["Nolan Howard",       9,  1, 1, 1,   1,  0,  0,  0,  1,  .111,   .111,   .385,   4,  3,  1,   0,   0],
    ["Lopes",              13, 3, 2, 3,   1,  1,  0,  0,  3,  .154,   .231,   .267,   0,  2,  0,   1,   1],
    ["Kaden Marler",       45, 10, 10, 9,  9,  0,  0,  0,  9,  .222,   .200,   .386,   17, 11, 1,   0,   6],
    ["Jack McLellan",      65, 11, 19, 19, 14, 4, 1,  0,  29, .292,   .446,   .370,   14, 7,  1,   1,   0],
    ["Luke Miller",        15, 7, 1, 0,   1,  0,  0,  0,  1,  .067,   .067,   .364,   5,  7,  0,   0,   0],
    ["Will Payne",         23, 3, 4, 4,   2,  2,  0,  0,  6,  .174,   .261,   .240,   6,  2,  0,   0,   0],
    ["Cutter Peterson",    30, 5, 6, 5,   4,  2,  0,  0,  8,  .200,   .267,   .385,   6,  7,  2,   2,   3],
    ["Chase Pringle",      46, 13, 11, 8,  9,  1,  0,  0,  11, .239,   .239,   .386,   15, 11, 0,   1,   4],
    ["Wriley Schreiner",   55, 13, 19, 18, 15, 3, 0,  1,  22, .345,   .400,   .455,   10, 6,  5,   3,   2],
    ["Krilich",            2,  0, 1, 0,   0,  1,  0,  0,  2,  .500,   .500,   .500,   1,  0,  0,   0,   0],
]
TEAM_BAT_2022 = [
    "TEAM", 646, 166, 178, 138, 133, 36, 7, 2, 224, .276, .347, .400, 146, 107, 27, 13, 57,
]

# 2021 pitching from PDF page 2
PITCHING_2021 = [
    # Player,    W, L, SV, IP,  BF, R, H, SO, BB, ER, 1B, 2B, 3B, HR, HBP, ERA
    ["Colello",  1, 0, 0, 5.0,  23, 4, 7, 4,  4,  4,  7,  0,  0,  0,  0,   5.600],
    ["Cosmos",   1, 0, 0, 3.0,  14, 4, 4, 5,  1,  3,  2,  2,  0,  0,  0,   7.000],
    ["Flaherty", 3, 0, 0, 19.1, 76, 5, 11, 10, 5, 5,  8,  1,  1,  0,  1,   1.810],
    ["Hento",    1, 1, 0, 4.0,  19, 4, 4, 4,  2,  3,  2,  0,  0,  0,  2,   5.250],
    ["Hodges",   0, 0, 0, 1.0,  3,  0, 0, 0,  0,  0,  0,  0,  0,  0,  0,   0.000],
    ["Holum",    0, 0, 3, 6.2,  27, 2, 4, 4,  2,  2,  3,  1,  0,  0,  2,   2.100],
    ["Lundberg", 0, 0, 0, 1.0,  4,  1, 0, 0,  1,  0,  0,  0,  0,  0,  0,   0.000],
    ["McLellan", 0, 0, 0, 1.0,  4,  0, 0, 1,  0,  0,  0,  0,  0,  0,  0,   0.000],
    ["Miller",   1, 1, 1, 13.1, 63, 8, 17, 10, 4, 7,  13, 3,  1,  0,  3,   3.675],
    ["Vorpahl",  2, 0, 0, 10.2, 43, 3, 10, 6,  3, 3,  10, 0,  0,  0,  0,   1.968],
    ["Voves",    0, 0, 0, 1.0,  7,  1, 1, 3,  2,  1,  1,  0,  0,  0,  1,   7.000],
]
# 2021 team pitching totals
TEAM_PIT_2021 = ["TEAM", 8, 2, 5, 66.0, 283, 31, 59, 46, 24, 28, 47, 7, 2, 3, 6, 2.970]

# 2022 pitching from PDF page 2
PITCHING_2022 = [
    # Player,             W, L, SV, IP,   BF,  R,  H,  SO, BB, ER, 1B, 2B, 3B, HR, HBP, ERA
    ["Luke Miller",       1, 1, 0, 6.1,  32,  5,  7,  7,  3,  3,  7,  0,  0,  0,  0,   3.443],
    ["Jack McLellan",     0, 1, 0, 0.1,  3,   2,  1,  1,  1,  0,  1,  0,  0,  0,  0,   0.000],
    ["Sam Barber",        4, 0, 0, 23.0, 113, 12, 23, 24, 13, 12, 17, 6,  0,  0,  2,   3.652],
    ["Garrick Cosmos",    4, 1, 0, 44.2, 202, 30, 51, 46, 11, 23, 39, 10, 1,  1,  4,   3.602],
    ["Ryland Heckman",    2, 2, 0, 24.2, 116, 16, 12, 19, 16, 10, 7,  4,  1,  0,  8,   2.834],
    ["Drew Hento",        0, 1, 0, 22.1, 121, 21, 26, 13, 8,  9,  20, 6,  1,  0,  5,   2.825],
    ["Justin Holum",      0, 0, 1, 1.1,  6,   0, 1,  1,  1,  0,  1,  0,  0,  0,  0,   0.000],
    ["Brendan Masini",    2, 2, 0, 21.2, 106, 19, 24, 23, 11, 18, 16, 8,  0,  0,  4,   5.806],
    ["Cutter Peterson",   0, 0, 0, 1.0,  5,   1, 2,  0,  0,  1,  2,  0,  0,  0,  1,   7.000],
    ["Chase Pringle",     0, 0, 0, 7.2,  36,  2, 8,  5,  3,  1,  7,  1,  0,  0,  0,   0.909],
    ["Ryan Vorpahl",      0, 5, 1, 15.1, 83,  31, 25, 17, 17, 23, 21, 1,  0,  0,  12,  10.523],
    ["Westfall",          0, 0, 0, 1.1,  12,  4, 5,  2,  4,  4,  2,  1,  1,  0,  1,   21.053],
]
TEAM_PIT_2022 = ["TEAM", 13, 13, 2, 170.2, 835, 143, 179, 158, 88, 104, 140, 41, 4, 1, 36, 4.265]


# ─── Sheet writers ───────────────────────────────────────────────────────────

def _bold(cell, fill=False):
    cell.font = Font(bold=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    cell.alignment = Alignment(horizontal="center")


def write_batting_sheet(ws, year, batting_lines, team_total):
    ws.cell(row=1, column=1, value=f"{year} Gig Harbor Varsity — Hitting & Fielding").font = Font(bold=True, size=12)
    # Row 2 reserved for note (left blank)
    for ci, h in enumerate(BAT_HEADER, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        _bold(c, fill=True)
    for ri, line in enumerate(batting_lines, start=4):
        for ci, val in enumerate(line, start=1):
            ws.cell(row=ri, column=ci, value=val)
    team_row = 4 + len(batting_lines) + 1
    for ci, val in enumerate(team_total, start=1):
        c = ws.cell(row=team_row, column=ci, value=val)
        _bold(c)


def write_pitching_sheet(ws, year, pitching_lines, team_total):
    ws.cell(row=1, column=1, value=f"{year} Gig Harbor Varsity — Pitching").font = Font(bold=True, size=12)
    for ci, h in enumerate(PIT_HEADER, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        _bold(c, fill=True)
    for ri, line in enumerate(pitching_lines, start=4):
        for ci, val in enumerate(line, start=1):
            ws.cell(row=ri, column=ci, value=val)
    team_row = 4 + len(pitching_lines) + 1
    for ci, val in enumerate(team_total, start=1):
        c = ws.cell(row=team_row, column=ci, value=val)
        _bold(c)


def build_workbook(year, batting, team_bat, pitching, team_pit):
    wb = Workbook()
    # Drop default sheet
    default = wb.active
    wb.remove(default)
    bat_ws = wb.create_sheet("Hitting & Fielding")
    write_batting_sheet(bat_ws, year, batting, team_bat)
    pit_ws = wb.create_sheet("Pitching")
    write_pitching_sheet(pit_ws, year, pitching, team_pit)
    return wb


def main():
    # 2021
    out_2021 = HISTORICAL / "2021" / "2021_Season_Stats.xlsx"
    wb = build_workbook(2021, BATTING_2021, TEAM_BAT_2021, PITCHING_2021, TEAM_PIT_2021)
    wb.save(out_2021)
    print(f"Wrote {out_2021}")

    # 2022
    out_2022 = HISTORICAL / "2022" / "2022_Season_Stats.xlsx"
    wb = build_workbook(2022, BATTING_2022, TEAM_BAT_2022, PITCHING_2022, TEAM_PIT_2022)
    wb.save(out_2022)
    print(f"Wrote {out_2022}")


if __name__ == "__main__":
    main()
