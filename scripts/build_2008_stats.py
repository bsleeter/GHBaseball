#!/usr/bin/env python3
"""Build 2008 Gig Harbor Varsity season stats xlsx from pages in
Historical/2008/. The Individual Records page is hand-annotated —
original 2008 records were later crossed out and updated with 2009
records. Original 2008 values extracted from typed entries."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2008" / "2008_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None,
                col_widths=None, left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 0: README / Manual Update Required ───────────────────
    ws_n = wb.active
    ws_n.title = "README"
    ws_n["A1"] = "2008 Season Stats — Manual Update Required"
    ws_n.merge_cells("A1:C1")
    style_title(ws_n["A1"])
    ws_n.row_dimensions[1].height = 24

    readme_lines = [
        ("SHEET", "WHAT NEEDS ATTENTION", "STATUS"),
        ("Individual Records",
         "Re-derived directly from the 2008 Team Batting and Team Pitching sheets. The original records page was hand-annotated mid-2009 and unreliable. CAVEAT: the batting page appears to be a season-to-date snapshot, so end-of-season AB/TB/RBI/OBP and Most Triples values may have been higher than shown.",
         "NEEDS SPOT-CHECK"),
        ("Team Highlights",
         "Source page had the same cross-out/update treatment. Original typed 2008 values preserved here. A few context fields (e.g., 'Most Singles vs. NK') may need confirmation.",
         "NEEDS MANUAL UPDATE"),
        ("Schedule",
         "No schedule photo was provided for 2008 — no per-game results are in this workbook. Season record (13-7) comes from the pitching report's W-L totals.",
         "NOT PROVIDED"),
        ("Team Batting",
         "CoachStat report transcribed. Some per-player hit breakdowns (1B/2B/3B/HR) do not sum cleanly to published AVG/TB in the source sheet — values preserved as printed.",
         "NEEDS SPOT-CHECK"),
        ("Team Pitching",
         "CoachStat report transcribed. Team totals match source.",
         "READABLE"),
        ("Roster",
         "Clean transcription from the printed roster page.",
         "READABLE"),
    ]

    for c, h in enumerate(readme_lines[0], start=1):
        cell = ws_n.cell(row=3, column=c, value=h)
        style_header(cell)
    ws_n.row_dimensions[3].height = 22

    for i, row in enumerate(readme_lines[1:]):
        for c, val in enumerate(row, start=1):
            cell = ws_n.cell(row=4 + i, column=c, value=val)
            cell.font = Font(name="Arial", size=10, color=NAVY,
                             bold=(c == 3 and val in ("NEEDS MANUAL UPDATE", "NOT PROVIDED")))
            cell.alignment = Alignment(
                horizontal="left" if c != 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = BORDER
            if c == 3 and val == "NEEDS MANUAL UPDATE":
                cell.fill = PatternFill("solid", start_color=FLAG)
            elif c == 3 and val == "NOT PROVIDED":
                cell.fill = PatternFill("solid", start_color="F8D7DA")
            elif i % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT)
        ws_n.row_dimensions[4 + i].height = 48

    ws_n.column_dimensions["A"].width = 22
    ws_n.column_dimensions["B"].width = 72
    ws_n.column_dimensions["C"].width = 22
    ws_n.freeze_panes = "A4"

    # ─── Sheet 1: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Zach Pearson",       12),
        (2,  "Anthony Nikula",     12),
        (3,  "Vince Purchase",     12),
        (4,  "Brandon Rohde",      12),
        (5,  "Bubba Brown",        12),
        (6,  "Drew Young",         12),
        (7,  "Kykle Mauren",       12),
        (8,  "Chet Thompson",      11),
        (9,  "Mike Barnett",       11),
        (10, "Cameron Holcomb",    11),
        (11, "Nico Youngren",      11),
        (12, "Scott Benedict",     11),
        (13, "Mike Jones",         11),
        (14, "David Bigelow",      10),
        (15, "Scott Schultz",      10),
        (16, "Spencer Manjarrez",   9),
    ]
    _write_roster_sheet(
        ws_r,
        title="2008 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Ed T'sas"],
    )

    # ─── Sheet 2: Team Batting ──────────────────────────────────────
    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "AB", "BB", "R", "H", "1B", "2B", "3B", "HR", "RBI",
        "AVG", "SLG", "OBP", "HBP", "SAC-B", "SAC-F", "SO", "TB",
    ]
    bat_rows = [
        ("Barnett",    60, 16, 23, 21, 13, 7, 0, 1, 16, ".350", ".517", ".494",  0, 0, 1,  8, 31),
        ("Thompson",   60, 16, 24, 20, 13, 5, 0, 2, 18, ".333", ".517", ".468",  0, 0, 1,  6, 31),
        ("Holcomb",    58,  9, 14,  9,  6, 3, 0, 0,  9, ".155", ".207", ".338",  0, 0, 0,  4, 12),
        ("Rohde",      54,  3,  8, 15, 12, 2, 0, 1, 10, ".278", ".370", ".328",  0, 0, 0,  3, 20),
        ("Bigelow",    58,  7, 17, 22, 18, 1, 0, 2, 13, ".379", ".569", ".449",  0, 2, 2,  8, 30),
        ("Youngren",   57, 14, 16, 22, 16, 2, 1, 0,  5, ".386", ".561", ".529",  1, 0, 1,  5, 32),
        ("Benedict",   28,  7,  8,  7,  5, 2, 0, 0,  7, ".250", ".357", ".457",  3, 0, 0,  2, 10),
        ("Pearson",    21, 10,  4,  3,  3, 0, 0, 0,  0, ".143", ".143", ".286",  1, 0, 0,  4,  3),
        ("Young",      36, 11, 15, 11,  9, 1, 0, 0,  7, ".306", ".333", ".440",  6, 0, 1,  9, 12),
        ("Mauren",      6,  0,  5,  2,  2, 0, 0, 0,  0, ".333", ".333", ".333",  0, 0, 0,  1,  2),
        ("Brown",      32,  3,  8,  8,  5, 1, 0, 2, 10, ".250", ".469", ".350",  2, 1, 0,  5, 15),
        ("Jones",      20,  1,  6,  4,  4, 0, 0, 0,  3, ".200", ".200", ".238",  0, 0, 0,  3,  4),
        ("Purchase",   20,  5,  4,  4,  4, 0, 0, 0,  1, ".200", ".200", ".370",  1, 0, 0,  2,  4),
        ("Nikula",     22,  4,  4,  5,  4, 0, 0, 1,  5, ".227", ".409", ".423",  2, 0, 0,  4,  9),
        ("Schultz",    11,  2,  2,  1,  1, 0, 0, 0,  0, ".091", ".091", ".250",  2, 0, 0,  2,  1),
        ("Manjarrez",  11,  1,  4,  4,  4, 0, 0, 0,  2, ".364", ".364", ".462",  1, 0, 0,  2,  4),
        ("TEAM",      554,109,162,158,119,24, 1,  9,106, ".285", ".401", ".411", 19, 3, 6, 68,220),
    ]
    write_table(
        ws2,
        title="2008 Gig Harbor Varsity — Team Batting",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14] + [6] * (len(bat_headers) - 1),
        left_align_cols={1},
        note=(
            "Note: Per-player values transcribed from CoachStat report. Some "
            "hit breakdowns don't sum cleanly to published AVG/TB — preserved "
            "as printed. Team totals approximate (source image had cell-level "
            "reading challenges). Trust Individual Records sheet for "
            "authoritative top numbers."
        ),
    )

    # ─── Sheet 3: Team Pitching ─────────────────────────────────────
    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "#BF", "RS", "ER", "BB", "H", "HBP",
        "W", "L", "SV", "BAA", "ERA", "SO",
    ]
    pit_rows = [
        ("Thompson",   "4.1",    18,  5,  1,  4,  2, 0, 0, 0, 0, ".143",  "1.62",  5),
        ("Holcomb",    "1.2",     9,  2,  2,  0,  3, 0, 0, 0, 0, ".333",  "8.40",  1),
        ("Rohde",     "44.0",   188, 20, 18, 19, 36, 4, 5, 2, 1, ".213",  "2.86", 54),
        ("Bigelow",   "22.0",   116, 23,  9, 14, 20, 4, 3, 0, 1, ".204",  "2.86", 21),
        ("Benedict",   "0.2",     5,  2,  2,  3,  0, 0, 0, 0, 0, ".000", "21.00",  0),
        ("Mauren",    "15.2",    68, 10,  9,  9, 18, 2, 3, 1, 2, ".234",  "4.02", 21),
        ("Brown",     "27.1",   124, 27, 25, 18, 37, 2, 1, 4, 0, ".356",  "6.40", 26),
        ("Jones",      "2.0",    11,  5,  5,  4,  2, 1, 0, 0, 0, ".200", "21.00",  2),
        ("Schultz",    "5.0",    20,  1,  1,  2,  3, 1, 1, 0, 0, ".176",  "1.40",  7),
        ("TEAM",     "122.1",   583, 95, 72, 73,120,14,13, 7, 4, ".242",  "4.12",135),
    ]
    write_table(
        ws3,
        title="2008 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Team Highlights ───────────────────────────────────
    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              24, "vs. Lincoln"),
        ("Most Runs (Inning)",            14, "vs. Lincoln"),
        ("Most Runs Allowed (Game)",      21, "vs. CK"),
        ("Widest Margin of Victory",      23, "vs. Lincoln"),
        ("One Run Games Lost",             3, "vs. SK, Bellarmine, CK"),
        ("One Run Games Won",              2, "vs. Stadium, SK"),
        ("Most K's (by GH Pitchers)",     10, "vs. Lincoln"),
        ("Most K's (by Opponents)",       13, "vs. Olympia"),
        ("Most Hits (Game)",              17, "vs. Lincoln"),
        ("Most Hits Allowed (Game)",      13, "vs. CK"),
        ("Most Walks (Game for GH)",      11, "vs. Bellarmine"),
        ("Most Walks Allowed",            12, "vs. CK"),
        ("Most Singles (Game)",           10, "vs. NK"),
        ("Most Doubles (Game)",            8, "vs. Lincoln"),
        ("Most Triples (Game)",            2, "vs. Lincoln"),
        ("Most Home Runs (Game)",          2, "vs. Lincoln"),
        ("Most Steals (Game)",             8, "vs. Foss"),
        ("Longest Winning Streak",         7, ""),
        ("Longest Losing Streak",          3, ""),
        ("Total Team Runs Scored",       145, ""),
        ("Total Team Hits",              167, ""),
        ("Total Team Extra-Base Hits",    62, ""),
        ("Total Team Home Runs",          11, ""),
        ("Total Team Stolen Bases",       44, ""),
        ("Total Team Walks",              94, ""),
        ("Total Team HBP",                21, ""),
    ]
    write_table(
        ws_h,
        title="2008 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
        note=(
            "Note: The Team Highlights page was later annotated with 2009 "
            "updates (crossed-out values and handwritten additions). Values "
            "shown here are the original typed 2008 entries."
        ),
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    # Records re-derived directly from the 2008 Team Batting and Team
    # Pitching sheets. Note: the 2008 batting page appears to be a season-
    # to-date snapshot rather than end-of-season totals — several values
    # likely climbed higher by season's end. The hand-annotated records
    # page had end-of-season values for some records but its 2009 cross-
    # outs make it unreliable.
    rec_rows = [
        ("Highest Average",        "Nico Youngren",                                ".386", "30+ AB (22-57)"),
        ("Most Plate Appearances", "Mike Barnett",                                  75,    "from records page; originally labeled 'Most At Bats' but is PA (≈AB+BB+HBP)"),
        ("Most HBP",               "Drew Young",                                     6,    ""),
        ("Lowest K Ratio",         "Brandon Rohde",                               "5.6%",  "(3 K / 54 AB) · 30+ AB"),
        ("Most Doubles",           "Mike Barnett",                                   7,    ""),
        ("Most Triples",           "Nico Youngren",                                  1,    "only player with triples"),
        ("Most Home Runs",         "Thompson, Bigelow, Brown",                       2,    "3-way tie"),
        ("Most Walks",             "Mike Barnett, Chet Thompson",                   16,    "tied"),
        ("Most Stolen Bases",      "Chet Thompson",                                 16,    "from records page (SB not in batting sheet)"),
        ("Most Total Bases",       "Nico Youngren",                                 32,    ""),
        ("Most RBIs",              "David Bigelow",                                 18,    ""),
        ("Best On-Base Avg.",      "Nico Youngren",                              ".529",   "30+ AB"),
        ("Longest Hitting Streak", "Mike Barnett",                                   8,    "from records page (not derivable from cumulative stats)"),
        ("Most Runs Scored",       "Chet Thompson",                                 24,    ""),
        ("Most Wins Pitching",     "Brandon Rohde",                                  5,    "5-2 record"),
        ("Most Innings Pitched",   "Brandon Rohde",                              "44.0",   ""),
        ("Most K's",               "Brandon Rohde",                                 54,    ""),
        ("Lowest ERA",             "Brandon Rohde, David Bigelow",               "2.86",   "21+ IP · tied"),
        ("Most Saves",             "Kyle Mauren",                                    2,    ""),
    ]
    write_table(
        ws4,
        title="2008 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 36, 12, 56],
        left_align_cols={2, 4},
        note=(
            "Records re-derived from the 2008 Team Batting and Team Pitching "
            "sheets. Note: the batting sheet appears to be a season-to-date "
            "snapshot — actual end-of-season values for AB, RBI, TB, OBP, "
            "and triples may have been higher. SB and hit streaks are not "
            "derivable from cumulative stats."
        ),
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
