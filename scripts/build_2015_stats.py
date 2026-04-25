#!/usr/bin/env python3
"""Build 2015 Gig Harbor Varsity season stats xlsx from scanned records
in Historical/2015/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2015" / "2015_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None, col_widths=None,
                left_align_cols=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    r = 4
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Schedule"
    schedule_rows = [
        ("Olympic",      "3-2"),
        ("Shelton",      "17-3"),
        ("G. Kapowsin",  "5-9"),
        ("Curtis",       "2-1"),
        ("S. Kitsap",    "4-2"),
        ("Bellarmine",   "1-11"),
        ("Bellarmine",   "3-8"),
        ("Yelm",         "3-1"),
        ("S. Kitsap",    "1-4"),
        ("Timberline",   "2-1"),
        ("Timberline",   "3-1"),
        ("Peninsula",    "7-1"),
        ("Capital",      "10-3"),
        ("P. Angeles",   "4-5"),
        ("Olympia",      "11-1"),
        ("Olympia",      "2-1"),
        ("Stadium",      "4-3"),
        ("Stadium",      "13-2"),
        ("Wilson",       "12-1"),
        ("Curtis",       "5-3"),
        ("Rogers",       "7-1"),
        ("Tahoma",       "8-1"),
        ("Decatur",      "16-2"),
        ("University",   "3-0"),
        ("Kentwood",     "1-0"),
        ("Newport",      "1-4"),
        ("Decatur",      "3-1"),
        ("Season Record", "21-7"),
    ]
    write_table(
        ws,
        title="2015 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        col_widths=[22, 18],
        left_align_cols={1},
    )

    # ─── Sheet 2: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Wyatt Elledge",         12),
        (2,  "Casey Gearhart",        12),
        (3,  "Matthew Henckel",       12),
        (4,  "Garrett Lean",          12),
        (5,  "Grant Sutton",          12),
        (6,  "Hunter Werner",         12),
        (7,  "Colton Robinson",       11),
        (8,  "Mike Toglia",           11),
        (9,  "Jon Burghardt",         11),
        (10, "Neal Hassan",           11),
        (11, "Jeremy Schnurman",      11),
        (12, "Patrick Fletcher",      11),
        (13, "Andrew Parker",         11),
        (14, "Alex Morford",          11),
        (15, "Drew Gallinger",        11),
        (16, "RJ Green",              10),
        (17, "Chad Stevens",          10),
        (18, "Cameron MacIntosh",     10),
        (19, "Alex Harrison",         10),
        (20, "Avery Jones",           10),
        (21, "Patrick Fredrickson",   10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2015 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Larry Roehr", "Kevin Owens"],
    )

    # ─── Sheet 3: Hitting & Fielding ────────────────────────────────
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Burghardt",   78,  8, 28, 25, 15, 7, 2, 4, 51, ".359", ".654", ".479", 11, 17, 1, 3,  25,  51,  1,  ".987", 0),
        ("Elledge",     44, 14,  5,  4,  4, 1, 0, 0,  6, ".114", ".136", ".204", 13,  4, 1, 0,   8,  13,  2,  ".913", 4),
        ("Fletcher",    22,  8,  5,  1,  1, 2, 0, 0, 10, ".227", ".455", ".348",  6,  3, 1, 0,   3,   5,  3,  ".750", 0),
        ("Fredrickson",  1,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   1,   0,  0, "1.000", 0),
        ("Gearhart",     3,  1,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  2,  0, 0, 0,   0,   0,  0,  ".800", 0),
        ("Green",       51, 14, 14,  6, 12, 2, 0, 0, 16, ".275", ".314", ".339",  2,  4, 1, 4,  25,   1,  2,  ".931", 4),
        ("Harrison",    15,  1,  2,  1,  2, 0, 0, 0,  2, ".133", ".133", ".316",  9,  3, 1, 0,  20,   4,  7,  ".889", 0),
        ("Hassan",      30,  8,  6,  1,  2, 1, 4, 0, 22, ".200", ".333", ".314", 10,  1, 2, 0,   2,   9,  7,  ".611", 0),
        ("Henckel",      1,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   1,   2, 13,  ".188", 0),
        ("Jones",        0,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   1,   2,  0, "1.000", 0),
        ("Lean",        24,  5,  8,  5,  5, 2, 1, 0, 13, ".333", ".542", ".385",  9,  2, 0, 0,  27,   4,  0, "1.000", 0),
        ("MacIntosh",   47,  6,  9,  7,  6, 3, 0, 0, 12, ".191", ".255", ".321",  8,  9, 0, 4,  27,  38,  3,  ".956", 2),
        ("Morford",      5,  0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  2,  1, 0, 0,   3,   0,  0, "1.000", 0),
        ("Parker",      44,  5,  6,  7,  5, 1, 0, 0,  7, ".136", ".159", ".191", 17,  2, 1, 0,  16,  22,  4,  ".905", 0),
        ("Robinson",    78, 20, 22, 11, 17, 4, 0, 1, 29, ".282", ".372", ".429", 13, 20, 0, 2,  22,   3, 10,  ".714", 0),
        ("Schnurman",   25,  4,  5,  4,  4, 1, 0, 0,  6, ".200", ".240", ".394",  2,  5, 3, 0,  95,   5,  1,  ".990", 0),
        ("Sutton",      27,  9,  8,  6,  5, 2, 0, 1, 13, ".296", ".481", ".472",  5,  5, 4, 0,   2,   2,  1,  ".984", 0),
        ("Stevens",     39,  7,  9,  7,  6, 3, 0, 0, 12, ".231", ".308", ".348", 13,  7, 0, 0,   5,  20, 34,  ".844", 0),
        ("Toglia",      80, 25, 24, 21,  9, 6, 5, 4, 52, ".300", ".650", ".429", 15, 15, 3, 0,   1,  54,  8,  ".984",17),
        ("Werner",      64, 12, 15, 19, 12, 1, 3, 0, 29, ".234", ".453", ".395", 10, 14, 3, 2,   0,  28,  0, "1.000", 0),
        ("TEAM",       678,146,165,126,100, 44, 14, 10, 268, ".245", ".395", ".369",155,174,20, 27, 395, 210, 44,  ".931",17),
    ]
    write_table(
        ws2,
        title="2015 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=[14] + [7] * (len(hit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Pitching ──────────────────────────────────────────
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF/AB", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Fredrickson",  1, 1, 0, "11.2", 50,  3,  9, 10,  5, 1,  6, 3, 0, 0, 0, "0.625"),
        ("Gearhart",     1, 2, 0, "11.0", 62, 15, 19,  6,  5, 9, 16, 3, 0, 0, 1, "5.727"),
        ("Harrison",     3, 0, 3, "14.0", 61,  4, 15, 14,  3, 4, 12, 3, 0, 0, 0, "2.000"),
        ("Henckel",      6, 1, 1, "53.0", 209,  9, 20, 67, 20, 6, 18, 2, 0, 0, 3, "0.792"),
        ("Jones",        1, 0, 0, "14.0", 59,  7,  8,  7,  4, 3,  4, 4, 0, 0, 1, "1.500"),
        ("Morford",      1, 2, 0, "20.2", 92, 14, 20, 25,  8, 8, 16, 3, 1, 0, 1, "2.772"),
        ("Parker",       3, 0, 0, "11.0", 49,  3,  9, 13,  5, 1,  8, 1, 0, 0, 0, "0.636"),
        ("Robinson",     2, 0, 0, "14.1", 67, 11, 14, 12,  7, 3,  9, 3, 1, 1, 1, "1.489"),
        ("Sutton",       0, 0, 0, "1.0",   4,  0,  0,  0,  0, 0,  1, 0, 0, 0, 0, "0.000"),
        ("Toglia",       3, 0, 2, "34.0", 134, 6, 22, 35,  7, 4, 16, 6, 0, 0, 1, "0.824"),
        ("TEAM",        21, 7, 6, "177.2", 787, 72, 136, 189, 65, 39, 105, 28, 2, 1, 8, "1.541"),
    ]
    write_table(
        ws3,
        title="2015 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    rec_rows = [
        ("Highest Average",        "Jon Burghardt",                 ".359", "30+ AB (28-78)"),
        ("Most Plate Appearances", "Michael Toglia, Jon Burghardt",  99,    "PA = AB + BB + HBP + SAC"),
        ("Lowest K Ratio",         "RJ Green",                      "3.3%", "(2-60)"),
        ("Most Doubles",           "Jon Burghardt",                   7,    ""),
        ("Most Triples",           "Michael Toglia",                  5,    ""),
        ("Most Home Runs",         "Michael Toglia, Jon Burghardt",   4,    ""),
        ("Most Walks",             "Colton Robinson",                20,    ""),
        ("Most Stolen Bases",      "Wyatt Elledge, RJ Green",         4,    ""),
        ("Most Total Bases",       "Michael Toglia",                 52,    ""),
        ("Most RBIs",              "Jon Burghardt",                  25,    ""),
        ("Best On-Base Avg.",      "Jon Burghardt",                 ".479", "30+ AB"),
        ("Longest Hitting Streak", "Jon Burghardt",                  13,    ""),
        ("Most Runs Scored",       "Michael Toglia",                 25,    ""),
        ("Most Wins Pitching",     "Matthew Henckel",                 6,    ""),
        ("Most Innings Pitched",   "Matthew Henckel",                53,    ""),
        ("Most K's",               "Matthew Henckel",                67,    ""),
        ("Lowest ERA",             "Matthew Henckel",              "0.792", "21+ IP"),
        ("Most Saves",             "Alex Harrison",                   3,    ""),
    ]
    write_table(
        ws4,
        title="2015 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 32, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
