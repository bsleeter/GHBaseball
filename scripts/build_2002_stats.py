#!/usr/bin/env python3
"""Build 2002 Gig Harbor Varsity season stats xlsx from pages in
Historical/2002/."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _shared_xlsx_builder import (
    write_table, write_readme, write_roster_sheet,
)
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2002" / "2002_Season_Stats.xlsx"


def build():
    wb = Workbook()

    ws_n = wb.active
    ws_n.title = "README"
    write_readme(ws_n, 2002, [
        ("Individual Records", "Clean typed source. Verify before publishing.", "NEEDS MANUAL UPDATE"),
        ("Team Highlights", "Clean typed source. Values preserved as printed.", "NEEDS MANUAL UPDATE"),
        ("Schedule", "No per-game schedule photo provided. Season record 10-8 comes from stats header (19 games — may include a tie or forfeit).", "NOT PROVIDED"),
        ("Team Batting", "Season-to-date stats thru 19 games transcribed. Some cells had reading challenges.", "NEEDS SPOT-CHECK"),
        ("Team Pitching", "Transcribed. Team: 10-8, 4.57 ERA, 103 K.", "READABLE"),
        ("Roster", "Clean transcription.", "READABLE"),
    ])

    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Jeff Huiatt",         12),
        (2,  "Tyler Kullman",       12),
        (3,  "Sam Rosendahl",       12),
        (4,  "Brett Shearer",       12),
        (5,  "Ryan Emmett",         11),
        (6,  "Chris Henning",       11),
        (7,  "David Hunt",          11),
        (8,  "David Jackson",       11),
        (9,  "Alex Medeiros",       11),
        (10, "Matt Stock",          11),
        (11, "Kevin Owens",         10),
        (12, "Kevin Bogue",         10),
        (13, "Grant Goodman",       10),
        (14, "Brandon Shurick",     10),
        (15, "Matt Pleau",           9),
        (16, "Matt Shearer",         9),
        (17, "Matt Schweitzer",      9),
    ]
    write_roster_sheet(
        ws_r,
        title="2002 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson"],
    )

    ws2 = wb.create_sheet("Team Batting")
    bat_headers = [
        "Player", "Inn", "AB", "R", "H", "2B", "3B", "HR", "BB", "RBI",
        "SAC", "K", "HBP", "SLG", "OBP", "AVG",
    ]
    bat_rows = [
        ("B. Shearer",   "97.0",  42, 18, 17, 6, 0, 1,  8, 14, 2,  7, 3, ".619", ".545", ".405"),
        ("Bogue",        "50.0",  15,  3,  3, 0, 0, 0,  1,  1, 0,  7, 0, ".200", ".267", ".200"),
        ("Cummings",      "1.0",   1,  0,  0, 0, 0, 0,  0,  0, 0,  1, 0, ".000", ".000", ".000"),
        ("Dorland",       "3.0",   4,  1,  1, 0, 0, 0,  0,  0, 0,  2, 0, ".250", ".250", ".250"),
        ("Emmett",       "93.0",  42,  8, 14, 4, 0, 0,  3,  7, 0,  7, 0, ".429", ".378", ".333"),
        ("Finley",        "1.0",   0,  0,  0, 0, 0, 0,  1,  0, 0,  0, 0, ".000", "1.000", ".000"),
        ("Goodman",      "39.0",  15,  4,  2, 2, 0, 0,  2,  0, 0,  5, 0, ".267", ".235", ".133"),
        ("Henning",      "21.0",   6,  0,  0, 0, 0, 0,  4,  0, 1,  1, 0, ".000", ".400", ".000"),
        ("Huiatt",       "65.0",  32,  8, 11, 1, 0, 0,  2,  5, 0,  4, 1, ".375", ".400", ".344"),
        ("Hunt",         "78.0",  40,  7,  9, 0, 1, 0,  4,  3, 0,  9, 1, ".275", ".311", ".225"),
        ("Jackson",      "76.0",  40,  9, 14, 2, 2, 0,  3, 10, 0,  9, 1, ".500", ".409", ".350"),
        ("Knox",          "2.0",   1,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Kullman",      "43.0",   7,  1,  1, 0, 0, 0,  0,  2, 0,  6, 0, ".143", ".143", ".143"),
        ("M. Shearer",   "72.0",  30,  9,  8, 0, 0, 0,  8,  6, 0,  5, 1, ".267", ".436", ".267"),
        ("Medeiros",     "81.0",  43,  9, 13, 3, 1, 0,  3, 10, 0,  9, 4, ".419", ".400", ".302"),
        ("Owens",       "111.0",  46, 15, 15, 1, 1, 1, 13, 15, 0,  2, 3, ".457", ".619", ".326"),
        ("Pleau",        "57.0",  23,  6,  4, 2, 0, 0,  4,  7, 1, 13, 0, ".261", ".286", ".174"),
        ("Rosendahl",   "123.0",  54, 12, 15, 2, 3, 0,  4,  7, 0,  7, 1, ".426", ".339", ".278"),
        ("Schweitzer",   "29.0",  15,  4,  6, 3, 0, 1,  1,  3, 1,  4, 0, ".800", ".471", ".400"),
        ("Shurick",      "38.0",  18,  5,  7, 3, 0, 0,  0,  1, 0,  7, 1, ".556", ".421", ".389"),
        ("Stock",       "110.0",  55, 10, 14, 3, 1, 0,  1,  8, 1, 14, 5, ".364", ".377", ".255"),
        ("Weyhrauch",     "7.0",   4,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("Zylkowski",     "1.0",   0,  0,  0, 0, 0, 0,  0,  0, 0,  0, 0, ".000", ".000", ".000"),
        ("TEAM",       "1197.0", 511,128,145,29, 6, 3, 73,101,11,108,13, ".362", ".400", ".284"),
    ]
    write_table(
        ws2,
        title="2002 Gig Harbor Varsity — Team Batting (Season to Date, 19 games)",
        headers=bat_headers,
        rows=bat_rows,
        team_row_index=len(bat_rows) - 1,
        col_widths=[14, 8] + [6] * (len(bat_headers) - 2),
        left_align_cols={1},
    )

    ws3 = wb.create_sheet("Team Pitching")
    pit_headers = [
        "Player", "IP", "BF", "H", "R", "ER", "BB", "WP", "HBP",
        "K", "W", "L", "SV", "OBA", "ERA",
    ]
    pit_rows = [
        ("Bogue",      "7.0",   33, 14,  7,  7,  1, 0, 0,  0, 0, 1, 0, ".438",  "7.00"),
        ("Emmett",    "30.2",  120, 34, 20, 16, 20, 0, 7, 30, 4, 0, 0, ".283",  "3.65"),
        ("Goodman",    "0.1",    3,  3,  3,  0,  0, 1, 0,  0, 0, 0, 0, "1.000",  "0.00"),
        ("Henning",   "19.2",   84, 22, 16, 11, 14, 0, 1, 18, 2, 0, 0, ".262",  "3.91"),
        ("Hunt",       "8.2",   42, 18, 11,  7,  5, 0, 1,  2, 1, 0, 0, ".429",  "5.65"),
        ("Jackson",    "7.0",   28,  8,  3,  1,  6, 0, 0,  6, 1, 1, 0, ".286",  "1.00"),
        ("Kullman",   "37.0",  130, 30, 27, 16, 24, 0, 4, 32, 3, 1, 1, ".231",  "3.03"),
        ("Rosendahl", "12.0",   55, 19, 13, 13,  5, 0, 4,  7, 0, 5, 1, ".345",  "7.58"),
        ("Zylkowski",  "0.2",    3,  0,  0,  0,  0, 0, 0,  0, 0, 0, 0, ".000",  "0.00"),
        ("TEAM",    "122.2",  507,144,105, 80, 76, 0,18,103,10, 8, 3, ".284",  "4.57"),
    ]
    write_table(
        ws3,
        title="2002 Gig Harbor Varsity — Team Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    ws_h = wb.create_sheet("Team Highlights")
    highlight_rows = [
        ("Most Runs (Game)",              21, "vs. Bremerton"),
        ("Most Runs (Inning)",            11, "vs. Mt. Tahoma"),
        ("Most Runs Allowed (Game)",      16, "vs. Port Angeles"),
        ("Widest Margin of Victory",      19, "vs. Bremerton"),
        ("One Run Games Lost",             2, "vs. Eatonville, Olympic"),
        ("One Run Games Won",              1, "vs. South Kitsap"),
        ("Most K's (by GH Pitchers)",     10, "vs. Foss & Olympic"),
        ("Most K's (by Opponents)",       12, "vs. South Kitsap"),
        ("Most Hits (Game)",              15, "vs. North Kitsap"),
        ("Most Hits Allowed (Game)",      14, "vs. Port Angeles"),
        ("Most Walks (Game for GH)",      13, "vs. Mt. Tahoma"),
        ("Most Walks Allowed",             8, "vs. Mt. Tahoma"),
        ("Most Singles (Game)",           11, "vs. North Kitsap"),
        ("Most Doubles (Game)",            4, "vs. North Kitsap & Bremerton"),
        ("Most Triples (Game)",            2, "vs. Olympic"),
        ("Most Home Runs (Game)",          1, "vs. W.F. West, PHS, Bremerton"),
        ("Most Steals (Game)",             9, "vs. Central Kitsap"),
        ("Longest Winning Streak",         5, ""),
        ("Longest Losing Streak",          3, ""),
        ("Total Team Runs Scored",       126, ""),
        ("Total Team Hits",              145, ""),
        ("Total Team Extra-Base Hits",    38, ""),
        ("Total Team Home Runs",           3, ""),
        ("Total Team Stolen Bases",       59, ""),
        ("Total Team Walks",              73, ""),
        ("Total Team HBP",                13, ""),
    ]
    write_table(
        ws_h,
        title="2002 Gig Harbor Varsity — Team Highlights",
        headers=["Highlight", "Value", "Context"],
        rows=highlight_rows,
        col_widths=[34, 10, 44],
        left_align_cols={1, 3},
    )

    ws4 = wb.create_sheet("Individual Records")
    rec_rows = [
        ("Highest Average",        "Brett Shearer",                     ".405", "30+ AB"),
        ("Most At Bats",           "Sam Rosendahl",                      54,    ""),
        ("Most HBP",               "Brett Shearer, Kevin Owens",          3,    ""),
        ("Lowest K Ratio",         "Kevin Owens",                       "3.2%", "(2-62)"),
        ("Most Doubles",           "Brett Shearer",                       6,    ""),
        ("Most Triples",           "Sam Rosendahl",                       3,    ""),
        ("Most Home Runs",         "Brett Shearer, Kevin Owens, Matt Schweitzer", 1, ""),
        ("Most Walks",             "Matt Stock",                         14,    ""),
        ("Most Stolen Bases",      "David Jackson",                      10,    ""),
        ("Most Total Bases",       "Brett Shearer, Kevin Owens",         37,    ""),
        ("Most RBIs",              "Kevin Owens",                        15,    ""),
        ("Most Wins Pitching",     "Ryan Emmett",                         4,    ""),
        ("Most Innings Pitched",   "Tyler Kullman",                      37,    ""),
        ("Most K's",               "Tyler Kullman",                      32,    ""),
        ("Lowest ERA",             "Tyler Kullman",                    "3.03",  ""),
        ("Most Saves",             "David Hunt",                          3,    ""),
        ("Best On-Base Avg.",      "Brett Shearer",                   "54.5%",  "30+ AB"),
        ("Longest Hitting Streak", "Ryan Emmett",                         8,    ""),
        ("Most Runs Scored",       "Brett Shearer",                      18,    ""),
    ]
    write_table(
        ws4,
        title="2002 Gig Harbor Varsity — Individual Records",
        headers=["Record", "Holder", "Value", "Qualifier / Note"],
        rows=rec_rows,
        col_widths=[26, 50, 12, 22],
        left_align_cols={2, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
