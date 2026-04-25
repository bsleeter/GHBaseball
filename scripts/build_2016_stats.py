#!/usr/bin/env python3
"""Build 2016 Gig Harbor Varsity season stats xlsx from scanned records
in Historical/2016/. Individual Records have been re-derived directly
from the 2016 hitting and pitching stats (the original records page
was hand-annotated mid-2017 and unreliable)."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "Historical" / "2016" / "2016_Season_Stats.xlsx"

NAVY = "1B2A4A"
CAROLINA = "4B9CD3"
LIGHT = "F4F6F9"
WHITE = "FFFFFF"
FLAG = "FFF3D6"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell):
    cell.font = Font(name="Arial", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_data(cell, is_even=False, bold=False, left_align=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=NAVY)
    if is_even:
        cell.fill = PatternFill("solid", start_color=LIGHT)
    cell.alignment = Alignment(
        horizontal="left" if left_align else "center", vertical="center"
    )
    cell.border = BORDER


def style_team_row(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", start_color=CAROLINA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_note(cell):
    cell.font = Font(name="Arial", size=9, italic=True, color="8B5A00")
    cell.fill = PatternFill("solid", start_color=FLAG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_roster_sheet(ws, title, rows, coaches):
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["#", "Player", "Grade"], start=1):
        cell = ws.cell(row=3, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[3].height = 20
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=4 + i, column=c, value=val)
            style_data(cell, is_even=(i % 2 == 1), left_align=(c == 2))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    last = 4 + len(rows) + 2
    ws.cell(row=last, column=1, value="COACHES").font = Font(
        name="Arial", size=10, bold=True, color=NAVY
    )
    for i, coach in enumerate(coaches):
        ws.cell(row=last + 1 + i, column=1, value=coach).font = Font(
            name="Arial", size=10, color=NAVY
        )
    ws.freeze_panes = "A4"


def write_table(ws, title, headers, rows, team_row_index=None, col_widths=None,
                left_align_cols=None, note=None):
    left_align_cols = left_align_cols or set()
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 22

    start_row = 3
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c = ws.cell(row=2, column=1, value=note)
        style_note(c)
        ws.row_dimensions[2].height = 36
        start_row = 4

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        style_header(cell)
    ws.row_dimensions[start_row].height = 20

    r = start_row + 1
    for i, row in enumerate(rows):
        is_team = team_row_index is not None and i == team_row_index
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if is_team:
                style_team_row(cell)
            else:
                style_data(cell, is_even=(i % 2 == 1), left_align=(c in left_align_cols))
        r += 1

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


def build():
    wb = Workbook()

    # ─── Sheet 1: Schedule ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Schedule"
    schedule_rows = [
        ("Curtis",       "3-2"),
        ("Puyallup",     "0-11"),
        ("G. Kapowsin",  "17-12"),
        ("C. Kitsap",    "3-5"),
        ("Rogers",       "0-3"),
        ("S. Kitsap",    "5-0"),
        ("Bellarmine",   "5-0"),
        ("Bellarmine",   "12-1"),
        ("Yelm",         "11-0 (9 inn)"),
        ("Yelm",         "4-2"),
        ("Capital",      "9-3"),
        ("Timberline",   "2-0"),
        ("Timberline",   "3-0"),
        ("S. Kitsap",    "4-5"),
        ("Peninsula",    "1-6"),
        ("Olympia",      "7-3"),
        ("Olympia",      "10-4"),
        ("Stadium",      "17-5"),
        ("Stadium",      "0-6"),
        ("Wilson",       "13-3"),
        ("Tahoma",       "6-0"),
        ("Fed Way",      ""),          # score not legible
        ("Kentwood",     "1-3"),
        ("Mt. Si",       "1-6"),       # state tournament
        ("Kentwood",     "3-7"),        # state tournament
        ("Season Record", "17-8"),
    ]
    write_table(
        ws,
        title="2016 Gig Harbor Varsity — Schedule",
        headers=["Opponent", "Result (GH-Opp)"],
        rows=schedule_rows,
        team_row_index=len(schedule_rows) - 1,
        col_widths=[22, 22],
        left_align_cols={1},
        note=(
            "Note: Schedule transcribed from the main stats sheet photo. "
            "Final 2 entries (Mt. Si, Kentwood) were state tournament games. "
            "One Fed Way score was not legible. 17-8 implies 25 games total."
        ),
    )

    # ─── Sheet 2: Roster ────────────────────────────────────────────
    ws_r = wb.create_sheet("Roster")
    roster_rows = [
        (1,  "Colton Robinson",     12),
        (2,  "Mike Toglia",         12),
        (3,  "Jon Burghardt",       12),
        (4,  "Neal Hassan",         12),
        (5,  "Jeremy Schnurman",    12),
        (6,  "Patrick Fletcher",    12),
        (7,  "Andrew Parker",       12),
        (8,  "Drew Gallinger",      12),
        (9,  "Kale Wong",           12),
        (10, "Jake Kein",           12),
        (11, "Jack Nordi",          11),
        (12, "RJ Green",            11),
        (13, "Chad Stevens",        11),
        (14, "Cameron MacIntosh",   11),
        (15, "Alex Harrison",       11),
        (16, "Avery Jones",         11),
        (17, "Jacob Bonnell",       11),
        (18, "Logan Gerling",       11),
        (19, "Patrick Fredrickson", 11),
        (20, "Jordan Haworth",      10),
    ]
    _write_roster_sheet(
        ws_r,
        title="2016 Gig Harbor Varsity — Roster",
        rows=roster_rows,
        coaches=["Pete Jansen", "Larry Carlson", "Larry Roehr"],
    )

    # ─── Sheet 3: Hitting & Fielding ────────────────────────────────
    ws2 = wb.create_sheet("Hitting & Fielding")
    hit_headers = [
        "Player", "AB", "R", "H", "RBI", "1B", "2B", "3B", "HR", "TB",
        "AVG", "SLG", "OBP", "SO", "BB", "HBP", "SAC", "PO", "A", "E",
        "FLD%", "SB",
    ]
    hit_rows = [
        ("Bonnell",      1,   0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   2,   0,  0, "1.000", 0),
        ("Burghardt",   79,   3, 24, 18, 16, 5, 1, 0, 36, ".304", ".456", ".402",  6, 13, 0, 0,  13,  20,  2,  ".943", 0),
        ("Fletcher",     6,   1,  2,  0,  1, 1, 0, 0,  3, ".333", ".500", ".333",  2,  0, 0, 0,   0,   0,  0,  ".000", 0),
        ("Fredrickson",  0,   0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   3,   6,  1,  ".900", 0),
        ("Gallinger",    1,   0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   1,   2,  0, "1.000", 0),
        ("Gerling",      6,   0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   1,   1,  2,  ".500", 0),
        ("Green",       81,  14, 26, 10, 24, 1, 1, 0, 29, ".321", ".370", ".415",  4, 13, 1, 0,  28,   2,  2,  ".938", 9),
        ("Harrison",     5,   0,  0,  0,  0, 0, 0, 0,  0, ".000", ".000", ".000",  0,  0, 0, 0,   1,   0,  0, "1.000", 0),
        ("Hassan",      64,  18, 16, 10, 12, 3, 1, 0, 22, ".250", ".344", ".333",  8,  6, 2, 1,  25,   0,  5,  ".833", 2),
        ("Haworth",     15,   2,  6,  1,  7, 0, 0, 0,  6, ".400", ".467", ".438",  6,  1, 0, 0,   1,   0,  0,  ".000", 0),
        ("Jones",        5,   2,  2,  1,  0, 1, 0, 0,  0, ".400", ".400", ".400",  1,  0, 0, 0,   0,   5,  0,  ".000", 0),
        ("Kein",         3,   0,  1,  2,  0, 0, 1, 0,  3, ".308", ".538", ".400",  2,  0, 0, 0,   0,   1,  0,  ".000", 0),
        ("MacIntosh",   71,  17, 21, 15, 17, 2, 1, 1, 28, ".296", ".394", ".412", 13, 11, 3, 1,  16,  42,  7,  ".892", 6),
        ("Nordi",        8,  14,  3,  0,  1, 2, 0, 0,  5, ".375", ".625", ".440",  1,  0, 0, 0,   4,   2,  1,  ".857", 0),
        ("Parker",      37,   6,  8, 11,  3, 2, 0, 0, 12, ".216", ".405", ".293", 11,  4, 0, 0,  14,   2,  0,  ".941", 0),
        ("Robinson",    88,  20, 26, 12, 17, 3, 0, 2, 32, ".294", ".338", ".473",  8, 19, 4, 2, 120,  16,  2,  ".783", 4),
        ("Schnurman",   54,   5, 18,  8, 16, 2, 0, 0, 20, ".333", ".370", ".449", 15, 15, 2, 1, 132,  10,  2,  ".986", 0),
        ("Stevens",     72,  15, 21, 12, 14, 6, 1, 0, 30, ".292", ".417", ".400", 16, 12, 2, 2,  30,  41,  4,  ".947", 5),
        ("Toglia",      76,  22, 24, 26, 15, 5, 1, 3, 41, ".316", ".566", ".475", 17, 22, 7, 1,  36,  14,  1,  ".981", 5),
        ("Wong",        13,   6,  3,  1,  3, 0, 0, 0,  3, ".231", ".231", ".333",  8,  2, 0, 0,   1,   0,  0,  ".875", 0),
        ("TEAM",       665, 148,196,124,149,30, 7, 5,270, ".295", ".406", ".415",114,115,18,10, 320, 176, 38,  ".929",38),
    ]
    write_table(
        ws2,
        title="2016 Gig Harbor Varsity — Hitting & Fielding",
        headers=hit_headers,
        rows=hit_rows,
        team_row_index=len(hit_rows) - 1,
        col_widths=[14] + [7] * (len(hit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 4: Pitching ──────────────────────────────────────────
    ws3 = wb.create_sheet("Pitching")
    pit_headers = [
        "Player", "W", "L", "SV", "IP", "BF/AB", "R", "H", "SO", "BB", "ER",
        "1B", "2B", "3B", "HR", "HBP", "ERA",
    ]
    pit_rows = [
        ("Bonnell",     1, 1, 0, "8.0",   36,  6,  7,  6,  1,  3,   7,  0, 0, 0, 1, "2.625"),
        ("Fredrickson", 3, 2, 0, "20.1",  85,  9, 16, 21,  5,  5,  15,  1, 0, 0, 2, "1.721"),
        ("Gallinger",   1, 1, 0, "22.1",  101, 12, 28, 21,  7, 10,  25,  3, 0, 0, 2, "3.134"),
        ("Gerling",     3, 0, 2, "18.1",  83, 14, 18, 26,  6,  5,  17,  1, 0, 0, 1, "1.909"),
        ("Harrison",    0, 0, 1, "13.0",  58,  7, 11, 19,  6,  6,   8,  3, 0, 0, 1, "3.230"),
        ("Jones",       3, 2, 0, "34.1",  138, 14, 24, 37,  9, 11,  19,  2, 1, 2, 3, "2.242"),
        ("Morford",     0, 1, 0, "1.0",   10,  5,  2,  0,  1,  1,   1,  1, 0, 0, 2, "7.000"),
        ("Parker",      1, 1, 0, "7.0",   35,  4, 10, 10,  3,  4,   8,  2, 0, 0, 0, "4.000"),
        ("Robinson",    0, 0, 0, "0.1",    4,  2,  3,  0,  0,  0,   0,  2, 1, 0, 0, "0.000"),
        ("Toglia",      5, 0, 0, "42.0",  151,  8, 19, 52, 10,  6,  15,  3, 0, 1, 0, "1.000"),
        ("TEAM",       17, 8, 3, "166.2", 701, 81,138,191, 48, 51, 117, 17, 1, 3,12, "2.142"),
    ]
    write_table(
        ws3,
        title="2016 Gig Harbor Varsity — Pitching",
        headers=pit_headers,
        rows=pit_rows,
        team_row_index=len(pit_rows) - 1,
        col_widths=[14] + [7] * (len(pit_headers) - 1),
        left_align_cols={1},
    )

    # ─── Sheet 5: Individual Records ────────────────────────────────
    ws4 = wb.create_sheet("Individual Records")
    rec_headers = ["Record", "Holder", "Value", "Qualifier / Note"]
    # Records re-derived directly from the 2016 Hitting & Fielding and
    # Pitching sheets, replacing the conflicting hand-annotated values from
    # the original records sheet. Mike Toglia (senior) led nearly every
    # offensive AND pitching category — a remarkable two-way season.
    rec_rows = [
        ("Highest Average",        "Jeremy Schnurman",                    ".333", "30+ AB (18-54)"),
        ("Most At Bats",           "Colton Robinson",                      88,    ""),
        ("Most HBP",               "Mike Toglia",                           7,    ""),
        ("Lowest K Ratio",         "RJ Green",                            "4.9%", "(4-81) · 30+ AB"),
        ("Most Doubles",           "Chad Stevens",                          6,    ""),
        ("Most Triples",           "7 players tied",                        1,    "Burghardt, Hassan, Green, MacIntosh, Stevens, Toglia, Kein"),
        ("Most Home Runs",         "Mike Toglia",                           3,    ""),
        ("Most Walks",             "Mike Toglia",                          22,    ""),
        ("Most Stolen Bases",      "RJ Green",                              9,    ""),
        ("Most Total Bases",       "Mike Toglia",                          41,    ""),
        ("Most RBIs",              "Mike Toglia",                          26,    ""),
        ("Best On-Base Avg.",      "Mike Toglia",                         ".475", "30+ AB"),
        ("Longest Hitting Streak", "RJ Green",                             "—",   "Source value illegible (game-by-game data not available)"),
        ("Most Runs Scored",       "Mike Toglia",                          22,    ""),
        ("Most Wins Pitching",     "Mike Toglia",                           5,    "5-0 record"),
        ("Most Innings Pitched",   "Mike Toglia",                       "42.0",   ""),
        ("Most K's",               "Mike Toglia",                          52,    ""),
        ("Lowest ERA",             "Mike Toglia",                       "1.000",  "21+ IP"),
        ("Most Saves",             "Logan Gerling",                         2,    ""),
    ]
    write_table(
        ws4,
        title="2016 Gig Harbor Varsity — Individual Records",
        headers=rec_headers,
        rows=rec_rows,
        col_widths=[26, 28, 10, 56],
        left_align_cols={2, 4},
        note=(
            "Records derived from the 2016 Hitting & Fielding and Pitching "
            "sheets in this workbook (replacing the original hand-annotated "
            "records page that had been edited mid-2017). Longest Hitting "
            "Streak cannot be derived from cumulative stats — needs game-log "
            "verification."
        ),
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
